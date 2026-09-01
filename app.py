from flask import Flask, request, jsonify, render_template, send_file, redirect,url_for, flash, jsonify
import azure.cognitiveservices.speech as speechsdk
from azure.storage.blob import BlobServiceClient,BlobClient,ContainerClient
from werkzeug.utils import secure_filename
import base64
import io
from io import BytesIO
from PIL import Image
import os
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from flask_cors import CORS
from datetime import datetime
from azure.storage.blob import ContentSettings
from dotenv import load_dotenv
from office365.sharepoint.client_context import ClientContext
from office365.runtime.auth.authentication_context import AuthenticationContext
from office365.sharepoint.lists.list import List
from office365.sharepoint.listitems.listitem import ListItem
import psycopg2
from werkzeug.security import generate_password_hash, check_password_hash
from flask import session
import secrets
from pydub import AudioSegment
import tempfile
import traceback
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image
import base64
import uuid
import json
import requests
from datetime import datetime
import pytz
from fpdf import FPDF
import io
from tempfile import NamedTemporaryFile
from supabase import create_client
from psycopg2 import pool as pg_pool
import time
from datetime import timezone
from contextlib import contextmanager
from datetime import timedelta
import secrets
import unicodedata
import re
from datetime import datetime, timedelta, timezone
from openai import OpenAI
import tempfile
from datetime import date
from flask import Response
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from api_movil import api_movil


connection_pool = None


SUPABASE_URL = os.environ.get("SUPABASE_URL")
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_KEY")
supabase_client = create_client(SUPABASE_URL, SUPABASE_KEY)


# Configurar zona horaria
tijuana_tz = pytz.timezone('America/Tijuana')
fecha_hora_tijuana = datetime.now(tijuana_tz)

UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

SYNCHRO_FORM_DEFINITION_ID = 'e4bQKVghekuuA8Y6dmHKWHlOJyH5vilDm9vLfuTg2mg'

DATABASE_URL = os.environ.get("DATABASE_URL")

POSTGRES_CONFIG = DATABASE_URL

# Configuración PostgreSQL
#POSTGRES_CONFIG = {
    #"host": "localhost",
    #"database": "Bitacora",
    #"user": "postgres",  # Normalmente 'postgres' por defecto
    #"password": "Daniel2030#",
    #"port": "5432"  # Puerto predeterminado de PostgreSQL
#}

SYNCHRO_CONFIG = {
    'client_id': 'service-o5fkAjNrOy3DBriRDwK4aA3Ud',
    'client_secret': 'VTkTyFi36+pUdJ/drZ5chOEhJufMuAZGofF9fzgg/SOUOkrPhPOZERxsq07FpleSZ0bBIRPJVjOua+bR4Exe3Q==',
    'token_url': 'https://ims.bentley.com/connect/token',
    'forms_url': 'https://api.bentley.com/forms',
    'itwin_id': '29d0867b-2158-4b7a-ae03-c63a7661ca58',
    'form_id': 'e4bQKVghekuuA8Y6dmHKWPFDh67WqydKr1vfz4Z0oAs'  # Formulario 1.09-00001
}

# Configura SharePoint (modifica con tus datos)
SHAREPOINT_SITE_URL = "https://iacsas.sharepoint.com/sites/Pruebasproyectossantiago"
LIST_NAME = "Proyectos"  # Nombre de la biblioteca
LIST_NAME_REGISTROS = "RegistrosBitacora"
SHAREPOINT_USER = "santiago.giraldo@iac.com.co"
SHAREPOINT_PASSWORD = "rwrwerwer"


# Cargar variables de entorno
#load_dotenv('config/settings.env')  # Ruta relativa al archivo .env

app = Flask(__name__,template_folder='templates')
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=12)
app.config['SESSION_COOKIE_SECURE'] = True
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

# Cliente de OpenAI para Whisper
openai_client = OpenAI(api_key=os.getenv('OPENAI_API_KEY'))

@app.before_request
def make_session_permanent():
    session.permanent = True
def init_pool():
    global connection_pool

    connection_pool = pg_pool.ThreadedConnectionPool(
        minconn=2,
        maxconn=10,
        dsn=os.environ.get('DATABASE_URL')
    )

    print(
        f"[POOL] Inicializado | "
        f"PID={os.getpid()} | "
        f"pool={id(connection_pool)} | "
        f"minconn=2 | maxconn=10"
    )


@contextmanager
def db_connection():
    conn = None
    cursor = None
    conn_ok = False

    try:
        for intento in range(2):
            try:
                conn = connection_pool.getconn()

                print(
                    f"[POOL] db_connection | "
                    f"pool={id(connection_pool)} | "
                    f"usadas={len(connection_pool._used)} | "
                    f"disponibles={len(connection_pool._pool)}"
                )

                if conn.closed:
                    connection_pool.putconn(conn, close=True)
                    conn = None
                    continue

                cursor = conn.cursor()

                # Comprobar conexión real
                cursor.execute("SELECT 1")
                cursor.fetchone()

                # Conexión válida
                break

            except psycopg2.OperationalError:
                if cursor:
                    try:
                        cursor.close()
                    except Exception:
                        pass
                    cursor = None

                if conn:
                    try:
                        connection_pool.putconn(conn, close=True)
                    except Exception:
                        pass
                    conn = None

                if intento == 1:
                    raise

        if conn is None or cursor is None:
            raise psycopg2.OperationalError(
                "No se pudo obtener una conexión válida"
            )

        empresa_id = session.get('empresa_id', 1)

        cursor.execute(
            "SET app.empresa_id = %s",
            (empresa_id,)
        )

        yield conn, cursor

        conn.commit()
        conn_ok = True

    except Exception:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        raise

    finally:
        if cursor:
            try:
                cursor.close()
            except Exception:
                pass

        if conn:
            try:
                if conn_ok and not conn.closed:
                    connection_pool.putconn(conn)
                else:
                    connection_pool.putconn(conn, close=True)
            except Exception:
                pass
#app.secret_key = secrets.token_hex(16)  # Clave secreta para sesiones
app.secret_key = os.environ.get('SECRET_KEY', 'bitacora-iac-2026-fallback')
#app.secret_key = '78787878tyg8987652vgdfdf3445'
CORS(app)

'''
from flask_mail import Mail, Message
import random, string

# Configuración de correo (ajusta con tu cuenta SMTP)
app.config['MAIL_SERVER']   = 'smtp.gmail.com'
app.config['MAIL_PORT']     = 587
app.config['MAIL_USE_TLS']  = True
app.config['MAIL_USERNAME'] = 'muneragacias@gmail.com'      # ← tu correo
app.config['MAIL_PASSWORD'] = 'rxghrdeqoupdkaex'         # ← contraseña de app Gmail
app.config['MAIL_DEFAULT_SENDER'] = 'muneragacias@gmail.com'

mail = Mail(app)
'''
import random, string
import msal
import requests

# ── Configuración de correo vía Microsoft Graph (Outlook / Microsoft 365) ──
GRAPH_TENANT_ID     = os.environ.get('GRAPH_TENANT_ID')
GRAPH_CLIENT_ID     = os.environ.get('GRAPH_CLIENT_ID')
GRAPH_CLIENT_SECRET = os.environ.get('GRAPH_CLIENT_SECRET')
GRAPH_SENDER_EMAIL  = os.environ.get('GRAPH_SENDER_EMAIL', 'bitacora.notificaciones@iaclatam.com')

_graph_app = msal.ConfidentialClientApplication(
    GRAPH_CLIENT_ID,
    authority=f'https://login.microsoftonline.com/{GRAPH_TENANT_ID}',
    client_credential=GRAPH_CLIENT_SECRET
)

def _obtener_token_graph():
    """Pide un access token a Azure AD (con caché interno de msal)."""
    resultado = _graph_app.acquire_token_silent(
        scopes=['https://graph.microsoft.com/.default'], account=None
    )
    if not resultado:
        resultado = _graph_app.acquire_token_for_client(
            scopes=['https://graph.microsoft.com/.default']
        )
    if 'access_token' not in resultado:
        raise Exception(f"No se pudo autenticar con Microsoft Graph: {resultado.get('error_description')}")
    return resultado['access_token']


def enviar_correo(destinatarios, asunto, cuerpo_html):
    """
    Envía un correo usando Microsoft Graph API, desde GRAPH_SENDER_EMAIL.
    destinatarios: string o lista de strings con los correos.
    """
    if isinstance(destinatarios, str):
        destinatarios = [destinatarios]

    token = _obtener_token_graph()

    payload = {
        "message": {
            "subject": asunto,
            "body": {"contentType": "HTML", "content": cuerpo_html},
            "toRecipients": [{"emailAddress": {"address": d}} for d in destinatarios]
        },
        "saveToSentItems": "true"
    }

    resp = requests.post(
        f'https://graph.microsoft.com/v1.0/users/{GRAPH_SENDER_EMAIL}/sendMail',
        headers={
            'Authorization': f'Bearer {token}',
            'Content-Type': 'application/json'
        },
        json=payload,
        timeout=15
    )

    if resp.status_code != 202:
        raise Exception(f"Error enviando correo ({resp.status_code}): {resp.text}")


projects = []

# Conecta con el servicio de Blob Storage de Azure
connection_string = "DefaultEndpointsProtocol=https;AccountName=registrobitacora;AccountKey=ZyHZAOvOBijiOfY3BR3ZEDZsCAHOu3swEPnS+D7AacR2Yr94HS+jBMa2/20sJpZ71decGXYHQxE2+AStBWI/wA==;EndpointSuffix=core.windows.net"
container_name = "registros"


# Inicializa el cliente de BlobServiceClient
blob_service_client = BlobServiceClient.from_connection_string(connection_string)

def generar_password_temporal(longitud=10):
    caracteres = string.ascii_letters + string.digits
    return ''.join(random.choices(caracteres, k=longitud))


@app.route('/upload_foto', methods=['POST'])
def upload_foto():
    if 'user_id' not in session:
        return jsonify({"error": "No autorizado"}), 401
    try:
        data = request.json
        file_data = data.get('file_data', '')

        if not file_data:
            return jsonify({"error": "No se recibió imagen"}), 400

        if ',' in file_data:
            header, b64 = file_data.split(',', 1)
            if 'png' in header:
                ext, mime = 'png', 'image/png'
            elif 'webp' in header:
                ext, mime = 'webp', 'image/webp'
            else:
                ext, mime = 'jpg', 'image/jpeg'
        else:
            b64, ext, mime = file_data, 'jpg', 'image/jpeg'

        imagen_bytes = base64.b64decode(b64)
        nombre_archivo = f"{uuid.uuid4()}.{ext}"
        ruta = f"registros/{nombre_archivo}"

        supabase_client.storage.from_('fotos-bitacora').upload(
            ruta,
            imagen_bytes,
            {"content-type": mime}
        )

        url_publica = f"{SUPABASE_URL}/storage/v1/object/public/fotos-bitacora/{ruta}"
        return jsonify({"url": url_publica}), 200

    except Exception as e:
        print(f"Error subiendo foto: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/invitar-empresa', methods=['POST'])
def invitar_empresa():
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    if session.get('empresa_id') != 1:
        return jsonify({'error': 'No tienes permisos'}), 403

    data     = request.get_json()
    email    = data.get('email', '').strip()
    contacto = data.get('contacto', '').strip() or 'Cliente'

    if not email:
        return jsonify({'error': 'El correo es obligatorio'}), 400

    try:
        token     = secrets.token_urlsafe(32)
        expira_en = datetime.now(timezone.utc) + timedelta(days=7)
        link      = f"https://bitacoraiac.onrender.com/registroEmpresa?token={token}"

        with db_connection() as (conn, cursor):
            cursor.execute("""
                INSERT INTO tokens_registro (token, email, expira_en)
                VALUES (%s, %s, %s)
            """, (token, email, expira_en))

        cuerpo_html = f"""
        <div style="font-family:'DM Sans',Arial,sans-serif;max-width:560px;margin:auto;background:#ffffff;border-radius:16px;overflow:hidden;box-shadow:0 4px 24px rgba(0,0,0,0.08);">

            <!-- Header -->
            <div style="background:#0f0f0f;padding:32px 40px;text-align:center;">
                <img src="https://bitacoraiac.onrender.com/static/LogoBlancoIAC.png"
                     style="height:52px;border-radius:10px;margin-bottom:16px;" alt="IAC">
                <h1 style="color:#FFAF33;font-size:22px;margin:0;font-weight:800;letter-spacing:-0.5px;">
                    Bitácora IAC
                </h1>
                <p style="color:#9ca3af;font-size:13px;margin:8px 0 0;">
                    Plataforma de gestión de proyectos en campo
                </p>
            </div>

            <!-- Cuerpo -->
            <div style="padding:40px;">
                <p style="font-size:16px;color:#1a1a1a;margin:0 0 8px;">
                    Hola, <strong>{contacto}</strong>
                </p>
                <p style="font-size:15px;color:#4b5563;line-height:1.6;margin:0 0 28px;">
                    <strong>IAC — Ingeniería Asistida por Computador</strong> te ha invitado a registrar
                    tu empresa en <strong>Bitácora IAC</strong>, la plataforma para gestión de
                    actividades y contactos en campo.
                </p>

                <!-- Pasos -->
                <div style="background:#f9fafb;border-radius:12px;padding:24px;margin-bottom:28px;">
                    <p style="font-size:13px;font-weight:700;color:#6b7280;letter-spacing:0.05em;margin:0 0 16px;">
                        ¿QUÉ INCLUYE TU CUENTA?
                    </p>
                    <div style="display:flex;align-items:center;gap:12px;margin-bottom:12px;">
                        <div style="width:32px;height:32px;border-radius:8px;background:#fff8ee;display:flex;align-items:center;justify-content:center;flex-shrink:0;">
                            <span style="color:#FFAF33;font-size:16px;">🏢</span>
                        </div>
                        <span style="font-size:14px;color:#374151;">Espacio exclusivo para tu empresa</span>
                    </div>
                    <div style="display:flex;align-items:center;gap:12px;margin-bottom:12px;">
                        <div style="width:32px;height:32px;border-radius:8px;background:#fff8ee;display:flex;align-items:center;justify-content:center;flex-shrink:0;">
                            <span style="color:#FFAF33;font-size:16px;">👥</span>
                        </div>
                        <span style="font-size:14px;color:#374151;">Invita a todo tu equipo de trabajo</span>
                    </div>
                    <div style="display:flex;align-items:center;gap:12px;">
                        <div style="width:32px;height:32px;border-radius:8px;background:#fff8ee;display:flex;align-items:center;justify-content:center;flex-shrink:0;">
                            <span style="color:#FFAF33;font-size:16px;">🎙️</span>
                        </div>
                        <span style="font-size:14px;color:#374151;">Registro por voz y captura de evidencias</span>
                    </div>
                </div>

                <!-- Botón CTA -->
                <div style="text-align:center;margin-bottom:28px;">
                    <a href="{link}"
                       style="display:inline-block;background:#FFAF33;color:#ffffff;
                              padding:16px 40px;border-radius:10px;text-decoration:none;
                              font-size:16px;font-weight:700;letter-spacing:0.02em;">
                        Registrar mi empresa →
                    </a>
                </div>

                <!-- Nota de expiración -->
                <div style="background:#fff8ee;border:1px solid #fed7aa;border-radius:8px;padding:14px 16px;margin-bottom:24px;">
                    <p style="font-size:13px;color:#92400e;margin:0;">
                        ⏳ <strong>Este enlace expira en 7 días.</strong>
                        Si necesitas uno nuevo, contacta a IAC.
                    </p>
                </div>

                <p style="font-size:12px;color:#9ca3af;line-height:1.5;margin:0;">
                    Si no esperabas este correo, puedes ignorarlo de forma segura.
                    El enlace solo funciona una vez y expira automáticamente.
                </p>
            </div>

            <!-- Footer -->
            <div style="background:#f9fafb;padding:24px 40px;border-top:1px solid #e5e7eb;text-align:center;">
                <p style="font-size:12px;color:#6b7280;margin:0;">
                    © 2026 IAC — Ingeniería Asistida por Computador
                </p>
                <p style="font-size:12px;color:#9ca3af;margin:6px 0 0;">
                    <a href="https://iac.com.co" style="color:#FFAF33;text-decoration:none;">iac.com.co</a>
                </p>
            </div>

        </div>
        """
        enviar_correo(
            destinatarios=email,
            asunto='Invitación para registrar tu empresa en Bitácora IAC',
            cuerpo_html=cuerpo_html
        )

        return jsonify({'success': True})

    except Exception as e:
        print(f"Error enviando invitación empresa: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/invitar-usuarios', methods=['POST'])
def invitar_usuarios():
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'No autenticado'}), 401
    data = request.get_json()
    resultado, codigo = invitar_usuarios_core(data, session['user_id'], session.get('empresa_id'))
    return jsonify(resultado), codigo


def invitar_usuarios_core(data, admin_user_id, empresa_id, _reintento=0):
    """Lógica pura de invitación. No usa session ni request.
    La llaman la ruta web y el blueprint del APK."""
    personas = data.get('personas', [])

    if not personas:
        return {'success': False, 'error': 'No se recibieron datos'}, 400

    try:
        with db_connection() as (conn, cursor):

            cursor.execute("""
                SELECT name, empresa_id FROM usuario WHERE user_id = %s
            """, (admin_user_id,))
            admin        = cursor.fetchone()
            admin_nombre = admin[0] if admin else 'El administrador'

            enviados = []
            omitidos = []

            for p in personas:
                nombre   = p.get('nombre', '')
                apellido = p.get('apellido', '')
                correo   = p.get('correo', '')
                cargo    = p.get('cargo', 'Sin asignar')
                rol      = p.get('rol', 'viewer')

                if not nombre or not apellido or not correo:
                    omitidos.append(correo or 'sin correo')
                    continue

                cursor.execute(
                    "SELECT user_id FROM usuario WHERE email = %s", (correo,)
                )
                if cursor.fetchone():
                    omitidos.append(correo)
                    continue

                password_temp = generar_password_temporal()
                hashed        = generate_password_hash(password_temp)

                cursor.execute("""
                    INSERT INTO usuario
                        (name, apellido, email, password, cargo, rol, empresa_id, estado)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, 'pendiente')
                """, (nombre, apellido, correo, hashed, cargo, rol, empresa_id))

                try:
                    cuerpo_html = f"""
                    <div style="font-family:Arial,sans-serif;max-width:480px;margin:auto;
                                padding:32px;background:#fff;border-radius:12px;border:1px solid #eee;">
                        <h2 style="color:#1A1A2E;margin:0 0 8px;">Has sido invitado</h2>
                        <p style="color:#555;font-size:15px;margin:0 0 24px;">
                            <strong>{admin_nombre}</strong> te ha invitado a unirte a Bitácora App.
                        </p>
                        <div style="background:#F5F6FA;border-radius:10px;padding:20px;margin-bottom:24px;">
                            <p style="margin:0 0 8px;color:#888;font-size:13px;">TUS CREDENCIALES</p>
                            <p style="margin:0 0 4px;font-size:15px;">
                                <strong>Correo:</strong> {correo}
                            </p>
                            <p style="margin:0;font-size:15px;">
                                <strong>Contraseña temporal:</strong>
                                <span style="font-family:monospace;background:#fff;padding:2px 8px;
                                             border-radius:4px;border:1px solid #ddd;">
                                    {password_temp}
                                </span>
                            </p>
                        </div>
                        <p style="color:#e09a1f;font-size:13px;margin:0 0 24px;">
                            ⚠️ Cambia tu contraseña después de ingresar por primera vez.
                        </p>
                        <a href="https://bitacora.iaclatam.com"
                           style="display:block;text-align:center;background:#FBAF33;color:#fff;
                                  padding:14px;border-radius:8px;text-decoration:none;
                                  font-weight:bold;font-size:16px;">
                            Ingresar a la app
                        </a>
                    </div>
                    """
                    enviar_correo(
                        destinatarios=correo,
                        asunto='Invitación — Bitácora App',
                        cuerpo_html=cuerpo_html
                    )
                except Exception as mail_err:
                    print(f"Error enviando correo a {correo}: {mail_err}")

                enviados.append(correo)

            mensaje = f'Invitación enviada a {len(enviados)} persona(s).'
            if omitidos:
                mensaje += f' {len(omitidos)} omitido(s) por ya existir.'

            return {
                'success':  True,
                'enviados': len(enviados),
                'omitidos': omitidos,
                'mensaje':  mensaje
            }, 200

    except (psycopg2.OperationalError, psycopg2.InterfaceError) as e:
        print(f"Error de conexión en invitar_usuarios: {e} (intento {_reintento + 1})")
        if _reintento < 2:
            return invitar_usuarios_core(data, admin_user_id, empresa_id, _reintento=_reintento + 1)
        return {
            'success': False,
            'error':   'No se pudo conectar a la base de datos. Intenta enviar la invitación de nuevo.'
        }, 503

    except Exception as e:
        print(f"Error en invitar_usuarios: {e}")
        return {'success': False, 'error': 'Ocurrió un error al procesar la invitación. Intenta de nuevo.'}, 500


@app.route('/api/opciones-dinamicas')
def opciones_dinamicas():
    """Devuelve las opciones de un campo de selección dinámica: los valores
    ya registrados en otro formulario del mismo proyecto."""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    proyecto_id   = request.args.get('proyecto_id')
    formulario_id = request.args.get('formulario_id')
    campo         = request.args.get('campo')

    if not all([proyecto_id, formulario_id, campo]):
        return jsonify({'error': 'Faltan parámetros'}), 400

    try:
        with db_connection() as (conn, cursor):
            cursor.execute("""
                SELECT rf.id, rf.respuestas->>%s AS etiqueta
                FROM respuestas_formulario rf
                JOIN formularios f ON f.id = rf.formulario_id
                WHERE rf.id_proyecto   = %s
                  AND rf.formulario_id = %s
                  AND f.empresa_id     = %s
                  AND COALESCE(rf.respuestas->>%s, '') <> ''
                ORDER BY rf.created_at DESC
            """, (campo, proyecto_id, formulario_id,
                  session.get('empresa_id'), campo))

            vistos, opciones = set(), []
            for rid, etiqueta in cursor.fetchall():
                clave = etiqueta.strip().lower()
                if clave in vistos:
                    continue          # el mismo frente puede tener varias aperturas
                vistos.add(clave)
                opciones.append({'value': etiqueta.strip(), 'codigo': rid})

            opciones.sort(key=lambda o: o['value'])
            return jsonify(opciones)

    except Exception as e:
        print(f"Error en opciones-dinamicas: {e}")
        return jsonify({'error': 'Error consultando opciones'}), 500


# ── BI: CRUD de tableros ─────────────────────────────────────────────

@app.route('/api/bi/tableros', methods=['GET'])
def bi_listar_tableros():
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        with db_connection() as (conn, cursor):
            cursor.execute("""
                SELECT id, nombre, created_at
                FROM tableros_bi
                WHERE empresa_id = %s
                ORDER BY created_at DESC
            """, (session.get('empresa_id'),))
            tableros = [{'id': r[0], 'nombre': r[1], 'created_at': str(r[2])}
                        for r in cursor.fetchall()]
        return jsonify({'tableros': tableros})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/bi/campos')
def bi_campos():
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    formulario_id = request.args.get('formulario_id', '')
    if not formulario_id:
        return jsonify({'error': 'Falta formulario_id'}), 400

    try:
        with db_connection() as (conn, cursor):
            cursor.execute("""
                SELECT campos FROM formularios
                WHERE id = %s AND empresa_id = %s
            """, (formulario_id, session.get('empresa_id')))
            row = cursor.fetchone()
            if not row:
                return jsonify({'error': 'Formulario no encontrado'}), 404

            campos_raw = row[0] or []

            todos_ids = [
                item['id'] for item in campos_raw
                if isinstance(item, dict) and 'id' in item
            ]

            campos_info = {}
            if todos_ids:
                cursor.execute("""
                    SELECT id, nombre, tipo, opciones FROM campos_globales WHERE id = ANY(%s)
                """, (todos_ids,))
                for r in cursor.fetchall():
                    opciones = r[3] or []
                    # Detectar si tiene valor asociado (opciones con izquierda/derecha)
                    con_valor = (
                        len(opciones) > 0 and
                        isinstance(opciones[0], dict) and
                        'izquierda' in opciones[0]
                    )
                    campos_info[str(r[0])] = {
                        'nombre':    r[1],
                        'tipo':      r[2],
                        'con_valor': con_valor
                    }

        sueltos = []
        grupos  = []
        grupo_actual = None

        for item in campos_raw:
            if not isinstance(item, dict):
                continue

            if item.get('tipo') == 'grupo':
                if grupo_actual and grupo_actual['campos']:
                    grupos.append(grupo_actual)
                grupo_actual = {
                    'gid':    item.get('gid'),
                    'nombre': item.get('nombre'),
                    'campos': []
                }
            elif 'id' in item:
                cid  = str(item['id'])
                info = campos_info.get(cid)
                if not info:
                    continue
                campo = {
                    'id':        cid,
                    'nombre':    info['nombre'],
                    'tipo':      info['tipo'],
                    'con_valor': info['con_valor']   # ← nuevo
                }
                if grupo_actual is not None:
                    grupo_actual['campos'].append(campo)
                else:
                    sueltos.append(campo)

        if grupo_actual and grupo_actual['campos']:
            grupos.append(grupo_actual)

        return jsonify({'sueltos': sueltos, 'grupos': grupos})

    except Exception as e:
        print(f"Error en bi_campos: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/bi/tableros', methods=['POST'])
def bi_crear_tablero():
    if 'user_id' not in session or session.get('user_rol') != 'admin':
        return jsonify({'error': 'No autorizado'}), 401
    try:
        data   = request.get_json()
        nombre = data.get('nombre', '').strip()
        if not nombre:
            return jsonify({'error': 'El nombre es obligatorio'}), 400
        with db_connection() as (conn, cursor):
            cursor.execute("""
                INSERT INTO tableros_bi (empresa_id, nombre)
                VALUES (%s, %s) RETURNING id
            """, (session.get('empresa_id'), nombre))
            nuevo_id = cursor.fetchone()[0]
        return jsonify({'success': True, 'id': nuevo_id})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/bi/tableros/<int:tablero_id>', methods=['DELETE'])
def bi_eliminar_tablero(tablero_id):
    if 'user_id' not in session or session.get('user_rol') != 'admin':
        return jsonify({'error': 'No autorizado'}), 401
    try:
        with db_connection() as (conn, cursor):
            cursor.execute("""
                DELETE FROM tableros_bi
                WHERE id = %s AND empresa_id = %s
            """, (tablero_id, session.get('empresa_id')))
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── BI: CRUD de visualizaciones ──────────────────────────────────────

@app.route('/api/bi/tableros/<int:tablero_id>/visualizaciones', methods=['GET'])
def bi_listar_visualizaciones(tablero_id):
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        with db_connection() as (conn, cursor):
            # Verificar que el tablero pertenece a la empresa
            cursor.execute("""
                SELECT id FROM tableros_bi
                WHERE id = %s AND empresa_id = %s
            """, (tablero_id, session.get('empresa_id')))
            if not cursor.fetchone():
                return jsonify({'error': 'Tablero no encontrado'}), 404

            cursor.execute("""
                SELECT id, tipo, titulo, config, posicion
                FROM visualizaciones_bi
                WHERE tablero_id = %s
                ORDER BY posicion ASC
            """, (tablero_id,))
            vizs = [{'id': r[0], 'tipo': r[1], 'titulo': r[2],
                     'config': r[3], 'posicion': r[4]}
                    for r in cursor.fetchall()]
        return jsonify({'visualizaciones': vizs})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/bi/tableros/<int:tablero_id>/visualizaciones', methods=['POST'])
def bi_guardar_visualizaciones(tablero_id):
    if 'user_id' not in session or session.get('user_rol') != 'admin':
        return jsonify({'error': 'No autorizado'}), 401
    try:
        data           = request.get_json()
        visualizaciones = data.get('visualizaciones', [])
        with db_connection() as (conn, cursor):
            # Verificar pertenencia
            cursor.execute("""
                SELECT id FROM tableros_bi
                WHERE id = %s AND empresa_id = %s
            """, (tablero_id, session.get('empresa_id')))
            if not cursor.fetchone():
                return jsonify({'error': 'Tablero no encontrado'}), 404

            # Reemplazar todas las visualizaciones del tablero
            cursor.execute("DELETE FROM visualizaciones_bi WHERE tablero_id = %s", (tablero_id,))
            for i, viz in enumerate(visualizaciones):
                cursor.execute("""
                    INSERT INTO visualizaciones_bi (tablero_id, tipo, titulo, config, posicion)
                    VALUES (%s, %s, %s, %s, %s)
                """, (tablero_id, viz['tipo'], viz['titulo'],
                      json.dumps(viz.get('config', {})), i))
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── BI: datos para una visualización ─────────────────────────────────
@app.route('/api/bi/datos', methods=['POST'])
def bi_datos():
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        data       = request.get_json()
        datasets   = data.get('datasets', [])
        filtros    = data.get('filtros', {})
        desde      = filtros.get('desde', '')
        hasta      = filtros.get('hasta', '')
        proyecto   = filtros.get('proyecto_id', '')
        empresa_id = session.get('empresa_id')
        acumulado  = {}

        def buscar_valor_en_bloque(bloque, ids_posibles):
            """Para campos numéricos — devuelve el valor crudo."""
            for cid in ids_posibles:
                val = bloque.get(cid)
                if val is not None:
                    return val
            return None

        def buscar_label_en_bloque(bloque, ids_posibles, parte='valor'):
            """Para campos de agrupación — respeta si el admin eligió código o causa."""
            for cid in ids_posibles:
                if parte == 'causa':
                    texto = bloque.get(cid + '_codigo')
                    if texto:
                        return texto
                elif parte == 'codigo':
                    val = bloque.get(cid)
                    if val is not None:
                        return val
                else:
                    texto = bloque.get(cid + '_codigo') or bloque.get(cid)
                    if texto is not None:
                        return texto
            return None

        def obtener_label(resp, bloque, ids_posibles, parte, label_raiz=''):
            """Obtiene el label de agrupación desde raíz o desde el bloque."""
            if label_raiz:
                return label_raiz
            if bloque:
                return str(buscar_label_en_bloque(bloque, ids_posibles, parte) or 'Sin dato')
            # Nivel raíz sin bloque
            for cid in ids_posibles:
                if parte == 'causa':
                    v = resp.get(cid + '_codigo')
                elif parte == 'codigo':
                    v = resp.get(cid)
                else:
                    v = resp.get(cid + '_codigo') or resp.get(cid)
                if v:
                    return str(v)
            return 'Sin dato'

        for ds in datasets:
            formulario_id           = ds.get('formulario_id')
            campo_valor             = str(ds.get('campo_valor', ''))
            campo_valor_nombre      = ds.get('campo_valor_nombre', '')
            campo_agrupacion        = str(ds.get('campo_agrupacion', ''))
            campo_agrupacion_nombre = ds.get('campo_agrupacion_nombre', '')
            parte_agrupacion        = ds.get('parte_agrupacion', 'valor')
            agregacion              = ds.get('agregacion', 'suma')
            es_grupo                = ds.get('es_grupo', False)
            gid                     = ds.get('gid', '')
            campo_fecha_id          = str(ds.get('campo_fecha_id', ''))
            es_tabla  = ds.get('es_tabla', False)
            columnas  = ds.get('columnas', [])
            es_temporal             = ds.get('es_temporal', False)
            granularidad            = ds.get('granularidad', 'dia')
            es_agrupado             = ds.get('es_agrupado', False)
            campo_serie             = str(ds.get('campo_serie', ''))
            campo_serie_nombre      = ds.get('campo_serie_nombre', '')

            print(f"[BI-TABLA] es_tabla: {es_tabla}")
            print(f"[BI-TABLA] columnas: {columnas}")
            print(f"[BI-TABLA] formulario_id: {formulario_id}")
            print(f"[BI-TABLA] registros encontrados: {len(registros) if 'registros' in dir() else 'aún no consultado'}")

            if not formulario_id or (not campo_valor and not es_tabla):
                continue

            with db_connection() as (conn, cursor):
                cursor.execute("""
                    SELECT id FROM formularios
                    WHERE id = %s AND empresa_id = %s
                """, (formulario_id, empresa_id))
                if not cursor.fetchone():
                    continue

                query  = "SELECT rf.respuestas FROM respuestas_formulario rf WHERE rf.formulario_id = %s"
                params = [formulario_id]

                if proyecto:
                    query += " AND rf.id_proyecto = %s"
                    params.append(proyecto)
                if desde and campo_fecha_id:
                    query += f" AND NULLIF(rf.respuestas->>'{campo_fecha_id}', '')::date >= %s"
                    params.append(desde)
                if hasta and campo_fecha_id:
                    query += f" AND NULLIF(rf.respuestas->>'{campo_fecha_id}', '')::date <= %s"
                    params.append(hasta)

                cursor.execute(query, params)
                registros = [r[0] for r in cursor.fetchall() if isinstance(r[0], dict)]

                # Resolver IDs posibles para campo de agrupación
                ids_posibles_agrupacion = [campo_agrupacion]
                if campo_agrupacion_nombre:
                    cursor.execute("""
                        SELECT id FROM campos_globales
                        WHERE nombre = %s ORDER BY id DESC
                    """, (campo_agrupacion_nombre,))
                    ids_encontrados = [str(r[0]) for r in cursor.fetchall()]
                    if ids_encontrados:
                        ids_posibles_agrupacion = list(dict.fromkeys(
                            [campo_agrupacion] + ids_encontrados
                        ))

                # Resolver IDs posibles para campo de valor
                ids_posibles_valor = [campo_valor]
                if campo_valor_nombre:
                    cursor.execute("""
                        SELECT id FROM campos_globales
                        WHERE nombre = %s ORDER BY id DESC
                    """, (campo_valor_nombre,))
                    ids_encontrados = [str(r[0]) for r in cursor.fetchall()]
                    if ids_encontrados:
                        ids_posibles_valor = list(dict.fromkeys(
                            [campo_valor] + ids_encontrados
                        ))

                # Resolver IDs posibles para campo de serie (barras agrupadas)
                ids_posibles_serie = [campo_serie] if campo_serie else []
                if campo_serie_nombre and es_agrupado:
                    cursor.execute("""
                        SELECT id FROM campos_globales
                        WHERE nombre = %s ORDER BY id DESC
                    """, (campo_serie_nombre,))
                    ids_encontrados = [str(r[0]) for r in cursor.fetchall()]
                    if ids_encontrados:
                        ids_posibles_serie = list(dict.fromkeys(
                            ([campo_serie] if campo_serie else []) + ids_encontrados
                        ))

            # ── Modo tabla: devuelve filas completas ──
            if es_tabla and columnas:
                filas = []
                for resp in registros:
                    fila = {}
                    tiene_datos = False
                    for col in columnas:
                        cid      = str(col.get('campo_id', ''))
                        es_grupo = col.get('es_grupo', False)
                        gid_col  = col.get('gid', '')
                        tipo_col = col.get('campo_tipo', '')

                        if es_grupo and gid_col:
                            # Para campos de grupo toma el primer bloque
                            bloques = (resp.get('__repeticiones') or {}).get(gid_col, [])
                            if bloques:
                                val = bloques[0].get(cid + '_codigo') or bloques[0].get(cid) or ''
                                fila[cid] = str(val) if val else '—'
                                tiene_datos = True
                            else:
                                fila[cid] = '—'
                        else:
                            val = resp.get(cid + '_codigo') or resp.get(cid) or ''
                            fila[cid] = str(val) if val else '—'
                            if val:
                                tiene_datos = True

                    if tiene_datos:
                        filas.append(fila)

                return jsonify({
                    'filas':    filas,
                    'columnas': [{'campo_id': str(c.get('campo_id','')),
                                'nombre':   c.get('campo_nombre',''),
                                'tipo':     c.get('campo_tipo','')} for c in columnas],
                    'total':    len(filas)
                })

            # Procesar cada registro
            for resp in registros:
                # ── Modo temporal: agrupar por fecha (gráfico de línea) ──
                if es_temporal and campo_fecha_id:
                    fecha_raw = resp.get(campo_fecha_id, '')
                    if not fecha_raw:
                        continue
                    try:
                        from datetime import datetime
                        fecha = datetime.strptime(fecha_raw[:10], '%Y-%m-%d')
                        if granularidad == 'mes':
                            label = fecha.strftime('%Y-%m')
                        elif granularidad == 'semana':
                            label = f"{fecha.isocalendar()[0]}-S{fecha.isocalendar()[1]:02d}"
                        else:
                            label = fecha.strftime('%Y-%m-%d')
                    except:
                        continue

                    if es_grupo and gid:
                        for bloque in (resp.get('__repeticiones') or {}).get(gid, []):
                            val_raw = buscar_valor_en_bloque(bloque, ids_posibles_valor)
                            try:    val = float(val_raw or 0)
                            except: continue
                            acumulado[label] = acumulado.get(label, 0) + val
                    else:
                        val_raw = resp.get(campo_valor)
                        try:    val = float(val_raw or 0)
                        except: continue
                        acumulado[label] = acumulado.get(label, 0) + val

                # ── Modo agrupado: barras con múltiples series ──
                elif es_agrupado and ids_posibles_serie:
                    if es_grupo and gid:
                        label_raiz = ''
                        for cid in ids_posibles_agrupacion:
                            if parte_agrupacion == 'causa':
                                label_raiz = str(resp.get(cid + '_codigo') or '')
                            elif parte_agrupacion == 'codigo':
                                label_raiz = str(resp.get(cid) or '')
                            else:
                                label_raiz = str(resp.get(cid + '_codigo') or resp.get(cid) or '')
                            if label_raiz:
                                break

                        for bloque in (resp.get('__repeticiones') or {}).get(gid, []):
                            label = obtener_label(resp, bloque, ids_posibles_agrupacion,
                                                  parte_agrupacion, label_raiz)
                            serie = str(
                                buscar_label_en_bloque(bloque, ids_posibles_serie, 'valor')
                                or resp.get(campo_serie + '_codigo')
                                or resp.get(campo_serie)
                                or 'Sin dato'
                            )
                            val_raw = buscar_valor_en_bloque(bloque, ids_posibles_valor)
                            try:    val = float(val_raw or 0)
                            except: continue

                            if label not in acumulado:
                                acumulado[label] = {}
                            acumulado[label][serie] = acumulado[label].get(serie, 0) + val
                    else:
                        label = obtener_label(resp, None, ids_posibles_agrupacion, parte_agrupacion)
                        serie = str(
                            resp.get(campo_serie + '_codigo')
                            or resp.get(campo_serie)
                            or 'Sin dato'
                        )
                        val_raw = resp.get(campo_valor)
                        try:    val = float(val_raw or 0)
                        except: continue

                        if label not in acumulado:
                            acumulado[label] = {}
                        acumulado[label][serie] = acumulado[label].get(serie, 0) + val

                # ── Modo normal: una sola serie ──
                else:
                    if es_grupo and gid:
                        label_raiz = ''
                        for cid in ids_posibles_agrupacion:
                            if parte_agrupacion == 'causa':
                                label_raiz = str(resp.get(cid + '_codigo') or '')
                            elif parte_agrupacion == 'codigo':
                                label_raiz = str(resp.get(cid) or '')
                            else:
                                label_raiz = str(resp.get(cid + '_codigo') or resp.get(cid) or '')
                            if label_raiz:
                                break

                        for bloque in (resp.get('__repeticiones') or {}).get(gid, []):
                            label = obtener_label(resp, bloque, ids_posibles_agrupacion,
                                                  parte_agrupacion, label_raiz)
                            val_raw = buscar_valor_en_bloque(bloque, ids_posibles_valor)
                            try:    val = float(val_raw or 0)
                            except: continue

                            if agregacion == 'promedio':
                                if label not in acumulado:
                                    acumulado[label] = {'suma': 0, 'n': 0}
                                acumulado[label]['suma'] += val
                                acumulado[label]['n']    += 1
                            else:
                                acumulado[label] = acumulado.get(label, 0) + val
                    else:
                        label = obtener_label(resp, None, ids_posibles_agrupacion, parte_agrupacion)
                        val_raw = resp.get(campo_valor)
                        try:    val = float(val_raw or 0)
                        except: continue

                        if agregacion == 'conteo':
                            acumulado[label] = acumulado.get(label, 0) + 1
                        elif agregacion == 'promedio':
                            if label not in acumulado:
                                acumulado[label] = {'suma': 0, 'n': 0}
                            acumulado[label]['suma'] += val
                            acumulado[label]['n']    += 1
                        else:
                            acumulado[label] = acumulado.get(label, 0) + val

        # ── Calcular promedios finales (modo normal) ──
        if any(isinstance(v, dict) and 'suma' in v for v in acumulado.values()):
            acumulado = {
                k: round(v['suma'] / v['n'], 2)
                for k, v in acumulado.items()
                if isinstance(v, dict) and 'suma' in v and v['n']
            }

        # ── Respuesta para modo agrupado ──
        if any(ds.get('es_agrupado') for ds in datasets):
            labels = sorted(acumulado.keys())
            series_nombres = []
            for v in acumulado.values():
                if isinstance(v, dict):
                    for s in v.keys():
                        if s not in series_nombres:
                            series_nombres.append(s)

            series = [
                {
                    'nombre':  s,
                    'valores': [round(acumulado.get(label, {}).get(s, 0), 2) for label in labels]
                }
                for s in series_nombres
            ]
            total = round(sum(v for serie in series for v in serie['valores']), 2)
            return jsonify({'labels': labels, 'series': series, 'total': total})

        # ── Respuesta para modo normal y temporal ──
        if any(ds.get('es_temporal') for ds in datasets):
            ordenado = sorted(acumulado.items(), key=lambda x: x[0])
        else:
            ordenado = sorted(acumulado.items(), key=lambda x: x[1], reverse=True)

        labels  = [x[0] for x in ordenado]
        valores = [round(x[1], 2) for x in ordenado]
        return jsonify({'labels': labels, 'valores': valores, 'total': round(sum(valores), 2)})

    except Exception as e:
        print(f"Error en bi_datos: {e}")
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ── Exportar tablero BI a PDF ───────────────────────────────────
@app.route('/api/bi/tableros/<int:tablero_id>/exportar-pdf', methods=['POST'])
def bi_exportar_pdf(tablero_id):
    if 'user_id' not in session or session.get('user_rol') != 'admin':
        return jsonify({'error': 'No autorizado'}), 401

    try:
        data      = request.get_json() or {}
        bloques   = data.get('bloques') or []
        titulo    = (data.get('titulo') or 'Reporte analítico').strip()
        subtitulo = (data.get('subtitulo') or '').strip()
        notas     = (data.get('notas') or '').strip()
        filtros   = (data.get('filtros') or '').strip()

        if not bloques:
            return jsonify({'error': 'No se recibió ninguna visualización'}), 400

        # ── Datos de marca de la empresa ──
        empresa_nombre = ''
        logo_url       = None
        color_primario = '#FFAF33'

        with db_connection() as (conn, cursor):
            cursor.execute("""
                SELECT nombre, logo_url, color_primario
                FROM empresas WHERE id = %s
            """, (session.get('empresa_id'),))
            emp = cursor.fetchone()
            if emp:
                empresa_nombre = emp[0] or ''
                logo_url       = emp[1]
                color_primario = emp[2] or '#FFAF33'

        rgb = _hex_a_rgb(color_primario)

        # ── Descargar el logo a un archivo temporal (FPDF necesita ruta) ──
        logo_path = None
        if logo_url:
            try:
                r = requests.get(logo_url, timeout=8)
                if r.ok:
                    tmp = NamedTemporaryFile(delete=False, suffix='.png')
                    # Normaliza a PNG con fondo blanco: FPDF no maneja
                    # transparencia ni algunos formatos exóticos.
                    im = Image.open(BytesIO(r.content)).convert('RGBA')
                    fondo = Image.new('RGBA', im.size, (255, 255, 255, 255))
                    fondo.alpha_composite(im)
                    fondo.convert('RGB').save(tmp.name, 'PNG')
                    logo_path = tmp.name
            except Exception as e:
                print(f"No se pudo cargar el logo para el PDF: {e}")

        ahora = datetime.now(pytz.timezone('America/Bogota'))

        pdf = _PDFTablero(
            titulo_tablero=titulo,
            logo_path=logo_path,
            rgb=rgb,
            empresa=empresa_nombre
        )
        pdf.set_auto_page_break(auto=True, margin=20)
        pdf.alias_nb_pages()

        # ── Portada ──
        pdf.portada(titulo, subtitulo, empresa_nombre, filtros, ahora,
                    session.get('user_nombre') or '')

        # ── Visualizaciones ──
        
        # Las tablas van al final, como anexo: así el reporte mantiene
        # orientación vertical de principio a fin y solo rota en los anexos.
        graficos = [b for b in bloques if b.get('tipo') != 'tabla']
        tablas   = [b for b in bloques if b.get('tipo') == 'tabla']

        pdf.add_page()
        for bloque in graficos:
            tipo = bloque.get('tipo')
            if tipo == 'tarjetas':
                pdf.bloque_tarjetas(bloque.get('titulo', ''), bloque.get('tarjetas') or [])
            elif bloque.get('imagen'):
                pdf.bloque_grafico(bloque.get('titulo', ''), bloque['imagen'])

        # ── Notas (antes de los anexos) ──
        if notas:
            pdf.bloque_notas(notas)

        # ── Anexo: tablas de datos ──
        for bloque in tablas:
            pdf.bloque_tabla(bloque.get('titulo', ''),
                             bloque.get('columnas') or [],
                             bloque.get('filas') or [])

        salida = pdf.output(dest='S')
        if isinstance(salida, str):
            salida = salida.encode('latin-1')

        if logo_path:
            try:
                os.remove(logo_path)
            except OSError:
                pass

        return send_file(
            BytesIO(salida),
            mimetype='application/pdf',
            as_attachment=False,
            download_name=f"{titulo[:40]}.pdf"
        )

    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': f'Error generando el PDF: {e}'}), 500

def _hex_a_rgb(hex_color):
    """'#FFAF33' → (255, 175, 51). Devuelve el naranja por defecto si falla."""
    try:
        h = (hex_color or '').lstrip('#')
        return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))
    except Exception:
        return (255, 175, 51)


class _PDFTablero(FPDF):
    """PDF del tablero BI con encabezado y pie de marca."""

    def __init__(self, titulo_tablero='', logo_path=None, rgb=(255, 175, 51), empresa=''):
        super().__init__(orientation='P', unit='mm', format='A4')
        self.titulo_tablero = titulo_tablero
        self.logo_path      = logo_path
        self.rgb            = rgb
        self.empresa        = empresa
        self.es_portada     = False
        self.orientacion_actual = 'P'

    # ── Encabezado / pie automáticos ──
    def header(self):
        if self.es_portada:
            return
        if self.logo_path:
            try:
                self.image(self.logo_path, 15, 8, 16)
            except Exception:
                pass
        self.set_font('Helvetica', 'B', 9)
        self.set_text_color(90, 90, 90)
        self.set_xy(34, 11)
        self.cell(0, 5, self._txt(self.titulo_tablero), 0, 1, 'L')

        # Franja de color
        self.set_fill_color(*self.rgb)
        self.rect(15, 22, self.w - 30, 0.8, 'F')
        self.ln(8)

    def footer(self):
        if self.es_portada:
            return
        self.set_y(-14)
        self.set_font('Helvetica', '', 8)
        self.set_text_color(150, 150, 150)
        self.cell(0, 5, self._txt(self.empresa), 0, 0, 'L')
        self.cell(0, 5, f'Pagina {self.page_no()} de {{nb}}', 0, 0, 'R')

    # ── Secciones ──
    def portada(self, titulo, subtitulo, empresa, filtros, ahora, autor):
        self.es_portada = True
        self.add_page()

        # Franja superior de color
        self.set_fill_color(*self.rgb)
        self.rect(0, 0, 210, 6, 'F')

        y = 60
        if self.logo_path:
            try:
                self.image(self.logo_path, 85, y, 40)
                y += 48
            except Exception:
                y += 8

        self.set_xy(20, y)
        self.set_font('Helvetica', 'B', 22)
        self.set_text_color(30, 30, 30)
        self.multi_cell(170, 10, self._txt(titulo), 0, 'C')

        if subtitulo:
            self.ln(2)
            self.set_x(20)
            self.set_font('Helvetica', '', 12)
            self.set_text_color(110, 110, 110)
            self.multi_cell(170, 7, self._txt(subtitulo), 0, 'C')

        self.ln(14)
        self.set_x(20)
        self.set_font('Helvetica', 'B', 11)
        self.set_text_color(*self.rgb)
        self.cell(170, 6, self._txt(empresa), 0, 1, 'C')

        # Metadatos
        self.ln(10)
        self.set_font('Helvetica', '', 10)
        self.set_text_color(110, 110, 110)
        for linea in [
            f"Filtros aplicados: {filtros}" if filtros else '',
            f"Generado: {ahora.strftime('%d/%m/%Y %I:%M %p')}",
            f"Por: {autor}" if autor else ''
        ]:
            if linea:
                self.set_x(20)
                self.cell(170, 6, self._txt(linea), 0, 1, 'C')

        self.es_portada = False

    def _asegurar_vertical(self):
        """Los gráficos y tarjetas van en vertical. Si la tabla anterior
        dejó la página en horizontal, se abre una nueva vertical."""
        if self.orientacion_actual != 'P':
            self.add_page(orientation='P')
            self.orientacion_actual = 'P'

    def _ajustar(self, texto, ancho_mm):
        """Recorta el texto a lo que quepa en el ancho dado, con puntos suspensivos."""
        t = self._txt(texto)
        if self.get_string_width(t) <= ancho_mm - 2:
            return t
        while t and self.get_string_width(t + '...') > ancho_mm - 2:
            t = t[:-1]
        return t + '...' if t else ''

    def bloque_grafico(self, titulo, imagen_b64):
        self._asegurar_vertical()
        """Inserta un gráfico. Salta de página si no cabe completo."""
        ALTO_IMG   = 78    # mm
        ALTO_TOTAL = ALTO_IMG + 14

        if self.get_y() + ALTO_TOTAL > 270:
            self.add_page()

        self.set_font('Helvetica', 'B', 12)
        self.set_text_color(30, 30, 30)
        self.cell(0, 7, self._txt(titulo), 0, 1, 'L')
        self.ln(1)

        try:
            b64 = imagen_b64.split(',', 1)[1] if ',' in imagen_b64 else imagen_b64
            tmp = NamedTemporaryFile(delete=False, suffix='.png')
            tmp.write(base64.b64decode(b64))
            tmp.close()
            self.image(tmp.name, x=15, w=180)
            os.remove(tmp.name)
        except Exception as e:
            print(f"Error insertando gráfico en el PDF: {e}")
            self.set_font('Helvetica', 'I', 9)
            self.set_text_color(160, 160, 160)
            self.cell(0, 6, 'No se pudo renderizar esta visualizacion', 0, 1)

        self.ln(6)

    def bloque_tarjetas(self, titulo, tarjetas):
        self._asegurar_vertical()
        """Dibuja las tarjetas resumen como recuadros nativos, en filas de 3."""
        if not tarjetas:
            return

        COLS, ANCHO, ALTO, GAP = 3, 58.0, 20.0, 3.0
        n_filas    = (len(tarjetas) + COLS - 1) // COLS
        alto_total = 12 + n_filas * (ALTO + GAP)

        if self.get_y() + alto_total > 268:
            self.add_page()

        self.set_font('Helvetica', 'B', 12)
        self.set_text_color(30, 30, 30)
        self.cell(0, 7, self._txt(titulo), 0, 1, 'L')
        self.ln(2)

        x0 = 15.0
        y0 = self.get_y()

        for i, t in enumerate(tarjetas):
            col, fila = i % COLS, i // COLS
            x = x0 + col * (ANCHO + GAP)
            y = y0 + fila * (ALTO + GAP)

            # Fondo gris claro y franja del color de la empresa
            self.set_fill_color(249, 250, 251)
            self.rect(x, y, ANCHO, ALTO, 'F')
            self.set_fill_color(*self.rgb)
            self.rect(x, y, 1.6, ALTO, 'F')

            # Etiqueta (truncada si no cabe)
            label = self._txt(t.get('label', ''))
            self.set_xy(x + 5, y + 4)
            self.set_font('Helvetica', '', 8)
            self.set_text_color(120, 120, 120)
            while label and self.get_string_width(label) > ANCHO - 8:
                label = label[:-1]
            self.cell(ANCHO - 8, 4, label, 0, 0, 'L')

            # Valor
            self.set_xy(x + 5, y + 10)
            self.set_font('Helvetica', 'B', 14)
            self.set_text_color(30, 30, 30)
            self.cell(ANCHO - 8, 6, self._txt(t.get('valor', '')), 0, 0, 'L')

        self.set_y(y0 + n_filas * (ALTO + GAP) + 5)

    def bloque_tabla(self, titulo, columnas, filas):
        """Dibuja la tabla con TODAS las columnas seleccionadas y TODAS las
        filas que el usuario tiene filtradas. Las columnas se estrechan si son
        muchas (el usuario decide si quita alguna); las filas paginan solas."""
        if not columnas or not filas:
            return

        # Más de 6 columnas: hoja horizontal para ganar 77mm de ancho
        orientacion = 'L' if len(columnas) > 6 else 'P'
        self.add_page(orientation=orientacion)
        self.orientacion_actual = orientacion

        util = self.w - 30

        # Ancho proporcional al contenido más largo (encabezado o celdas)
        pesos = []
        for col in columnas:
            cid = col.get('campo_id')
            muestras = [str(col.get('nombre', ''))]
            muestras += [str(f.get(cid, '') or '') for f in filas[:60]]
            pesos.append(min(max(len(m) for m in muestras), 40))

        total_peso = sum(pesos) or 1
        anchos = [p / total_peso * util for p in pesos]

        # Totales de columnas numéricas (mismo criterio que la tabla en pantalla)
        totales = {}
        for col in columnas:
            if col.get('tipo') == 'numero':
                cid = col.get('campo_id')
                suma = 0.0
                for f in filas:
                    try:
                        suma += float(str(f.get(cid, '') or '0').replace(',', '.'))
                    except ValueError:
                        pass
                totales[cid] = suma

        def encabezado():
            self.set_font('Helvetica', 'B', 7.5)
            self.set_fill_color(55, 58, 64)
            self.set_text_color(255, 255, 255)
            for col, w in zip(columnas, anchos):
                self.cell(w, 7, self._ajustar(col.get('nombre', ''), w), 1, 0, 'C', True)
            self.ln()

        self.set_font('Helvetica', 'B', 12)
        self.set_text_color(30, 30, 30)
        self.cell(0, 7, self._txt(titulo), 0, 1, 'L')
        self.set_font('Helvetica', '', 8)
        self.set_text_color(120, 120, 120)
        self.cell(0, 5, self._txt(f"{len(filas)} registros"), 0, 1, 'L')
        self.ln(1)

        encabezado()

        self.set_font('Helvetica', '', 7.5)
        alto_fila = 5.5
        alterno = False

        for fila in filas:
            # Salto de página conservando la orientación y repitiendo encabezado
            if self.get_y() + alto_fila > self.h - 18:
                self.add_page(orientation=orientacion)
                encabezado()
                self.set_font('Helvetica', '', 7.5)

            self.set_fill_color(248, 249, 250) if alterno else self.set_fill_color(255, 255, 255)
            self.set_text_color(50, 50, 50)

            for col, w in zip(columnas, anchos):
                valor = fila.get(col.get('campo_id'), '')
                valor = '' if valor is None else str(valor)
                alineacion = 'R' if col.get('tipo') == 'numero' else 'L'
                self.cell(w, alto_fila, self._ajustar(valor, w), 1, 0, alineacion, True)
            self.ln()
            alterno = not alterno

        # Fila de totales
        if totales:
            if self.get_y() + alto_fila > self.h - 18:
                self.add_page(orientation=orientacion)
                encabezado()
            self.set_font('Helvetica', 'B', 7.5)
            self.set_fill_color(*self.rgb)
            self.set_text_color(255, 255, 255)
            for col, w in zip(columnas, anchos):
                cid = col.get('campo_id')
                if cid in totales:
                    txt = f"{totales[cid]:,.2f}".rstrip('0').rstrip('.')
                    self.cell(w, alto_fila + 1, self._ajustar(txt, w), 1, 0, 'R', True)
                else:
                    self.cell(w, alto_fila + 1, '', 1, 0, 'C', True)
            self.ln()

        self.set_text_color(50, 50, 50)
        self.ln(5)

    def bloque_notas(self, notas):
        if self.get_y() > 230:
            self.add_page()
        self.ln(4)
        self.set_font('Helvetica', 'B', 12)
        self.set_text_color(30, 30, 30)
        self.cell(0, 7, 'Notas y conclusiones', 0, 1)
        self.set_fill_color(*self.rgb)
        self.rect(15, self.get_y(), 30, 0.6, 'F')
        self.ln(4)
        self.set_font('Helvetica', '', 10)
        self.set_text_color(70, 70, 70)
        self.multi_cell(180, 6, self._txt(notas))

    # ── Utilidad ──
    @staticmethod
    def _txt(s):
        """FPDF clásico solo maneja latin-1: descarta lo que no pueda codificar."""
        return str(s or '').encode('latin-1', 'ignore').decode('latin-1')


# ══════════════════════════════════════════════════════════════════
#  INFORME DE AVANCE DE OBRA
#  Todo el mapeo de ids vive aquí: si se renombra o recrea un campo,
#  este es el único punto a tocar.
# ══════════════════════════════════════════════════════════════════
EMPRESA_INFORME_OBRA = 14
INFORME_OBRA_CFG = {
    # Campos del tipo de proyecto (proyectos.datos_tipo)
    'proyecto': {
        'nombre':     '127',
        'estado':     '128',
        'objeto':     '135',
        'alcance':    '136',
        'contratista':'138',
        'contrato_no':'139',
    },
    # Formulario 14 — Apertura de frente
    'apertura': {
        'formulario_id':     14,
        'frente':            '140',
        'ubicacion':         '141',
        'tramo':             '142',
        'tipo_elemento':     '129',
        'tipo_intervencion': '130',
        'fecha_ini_prog':    '143',
        'fecha_fin_prog':    '145',
    },
    # Formulario 18 — Corte mensual del frente
    'corte': {
        'formulario_id':   18,
        'frente_codigo':   '180_codigo',
        'sub_etapa':       '131',
        'estado_frente':   '132',
        'fecha_ini_real':  '144',
        'fecha_fin_real':  '181',
        'dias_atraso':     '159',
        'pct_fin_prog':    '165',
        'pct_fin_ejec':    '166',
        'ejec_fin_prog':   '167',
        'ejec_fin_real':   '168',
        'grupo_meta':      'gmthh3i4n58jb',
        'meta':            '160',
        'meta_unidad':     '148',
        'meta_prog':       '161',
        'meta_ejec':       '162',
        'meta_obs':        '163',
    },
    # Formulario 15 — Actividad ejecutada
    'actividad': {
        'formulario_id':  15,
        'frente_codigo':  '180_codigo',
        'descripcion':    '171',
        'fecha':          '172',
        'responsable':    '173',
        'grupo_material': 'gmthgmyqsg5lc',
        'mat_pavimento':  '149',
        'mat_uso':        '169',
        'mat_material':   '151',
        'mat_cantidad':   '170',
        'mat_unidad':     '148',
    },
    # Formulario 17 — Novedad / Imprevisto
    'novedad': {
        'formulario_id': 17,
        'frente_codigo': '180_codigo',
        'tipo':          '153',
        'impacto':       '154',
        'estado':        '155',
        'descripcion':   '174',
        'fecha':         '175',
        'observacion':   '176',
    },
}


def _io_num(valor, decimales=2):
    """Convierte a float un valor de texto del JSONB. 0.0 si no es numérico."""
    try:
        return round(float(str(valor or '0').strip().replace(',', '.')), decimales)
    except (ValueError, TypeError):
        return 0.0


def _io_money(valor):
    """Formatea un valor monetario al estilo del informe: $ 51.127.264.471,00"""
    n = _io_num(valor)
    entero = f"{n:,.2f}"
    # Cambia separadores al formato colombiano
    entero = entero.replace(',', '@').replace('.', ',').replace('@', '.')
    return f"$ {entero}"


def _io_fecha(valor):
    """Normaliza una fecha del JSONB a dd/mm/aaaa. Cadena vacía si no aplica."""
    raw = str(valor or '').strip()
    if not raw:
        return ''
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(raw[:10], fmt).strftime('%d/%m/%Y')
        except ValueError:
            continue
    return raw

INFORME_OBRA_EVIDENCIA_IDS = ['164', '133', '98']   # ids de campos de imagen
INFORME_OBRA_MAX_FOTOS     = 40                     # tope por informe (RAM en Render)


def _io_urls(valor):
    """El campo de imagen guarda una URL suelta, un array JSON, o —cuando
    psycopg2 ya deserializó el JSONB— una lista de Python."""
    if not valor:
        return []

    # Ya viene como lista/tupla desde el JSONB
    if isinstance(valor, (list, tuple)):
        return [str(u).strip() for u in valor if str(u or '').strip()]

    raw = str(valor).strip()
    if not raw:
        return []

    if raw.startswith('['):
        try:
            datos = json.loads(raw)
            if isinstance(datos, list):
                return [str(u).strip() for u in datos if str(u or '').strip()]
        except (ValueError, TypeError):
            # Respaldo: extraer URLs aunque el formato no sea JSON válido
            return re.findall(r'https?://[^\s"\',\]]+', raw)
        return []

    return [raw]


def _io_evidencias_de(bloque):
    """Extrae todas las URLs de imagen de un dict de respuestas."""
    urls = []
    for cid in INFORME_OBRA_EVIDENCIA_IDS:
        urls += _io_urls(bloque.get(cid))
    return urls


def _io_imagen_temp(url, ancho_px=900):
    """Descarga una imagen y la redimensiona antes de insertarla.
    Redimensionar es obligatorio: fotos de celular de 4 MB agotarían
    los 512 MB de RAM del plan de Render."""
    try:
        r = requests.get(url, timeout=12)
        if not r.ok:
            return None
        im = Image.open(BytesIO(r.content))
        if im.mode in ('RGBA', 'P', 'LA'):
            fondo = Image.new('RGB', im.size, (255, 255, 255))
            fondo.paste(im.convert('RGBA'), mask=im.convert('RGBA').split()[-1])
            im = fondo
        else:
            im = im.convert('RGB')

        if im.width > ancho_px:
            alto = int(im.height * ancho_px / im.width)
            im = im.resize((ancho_px, alto), Image.LANCZOS)

        tmp = NamedTemporaryFile(delete=False, suffix='.jpg')
        im.save(tmp.name, 'JPEG', quality=80)
        ratio = im.height / im.width
        im.close()
        return tmp.name, ratio
    except Exception as e:
        print(f"[INFORME] No se pudo cargar la imagen: {e}")
        return None

class _PDFInformeObra(FPDF):
    """Informe de avance de contrato de obra, con marca de la empresa."""

    def __init__(self, contrato_no='', logo_path=None, rgb=(255, 175, 51), empresa=''):
        super().__init__(orientation='P', unit='mm', format='A4')
        self.contrato_no = contrato_no
        self.logo_path   = logo_path
        self.rgb         = rgb
        self.empresa     = empresa
        self.set_auto_page_break(auto=True, margin=18)

    # ── Encabezado y pie ──
    def header(self):
        if self.logo_path:
            try:
                self.image(self.logo_path, 15, 8, 15)
            except Exception:
                pass
        self.set_font('Helvetica', 'B', 9)
        self.set_text_color(90, 90, 90)
        self.set_xy(33, 11)
        self.cell(0, 5, self._txt('INFORME DE AVANCE CONTRATO DE OBRA EN EJECUCION'), 0, 1)
        self.set_font('Helvetica', '', 8)
        self.set_xy(33, 16)
        self.cell(0, 4, self._txt(self.contrato_no), 0, 1)
        self.set_fill_color(*self.rgb)
        self.rect(15, 23, 180, 0.8, 'F')
        self.ln(9)

    def footer(self):
        self.set_y(-13)
        self.set_font('Helvetica', '', 8)
        self.set_text_color(150, 150, 150)
        self.cell(0, 5, self._txt(self.empresa), 0, 0, 'L')
        self.cell(0, 5, f'Pagina {self.page_no()} de {{nb}}', 0, 0, 'R')

    # ── Utilidades ──
    @staticmethod
    def _txt(s):
        return str(s or '').encode('latin-1', 'ignore').decode('latin-1')

    def _espacio(self, alto):
        """Salta de página si el bloque no cabe completo."""
        if self.get_y() + alto > self.h - 22:
            self.add_page()

    def titulo_seccion(self, texto):
        self._espacio(14)
        self.ln(3)
        self.set_fill_color(*self.rgb)
        self.set_text_color(255, 255, 255)
        self.set_font('Helvetica', 'B', 10)
        self.cell(180, 7, '  ' + self._txt(texto), 0, 1, 'L', True)
        self.set_text_color(40, 40, 40)
        self.ln(2)

    def campo_largo(self, etiqueta, valor):
        if not valor:
            return
        self._espacio(16)
        self.set_font('Helvetica', 'B', 9)
        self.cell(0, 5, self._txt(etiqueta), 0, 1)
        self.set_font('Helvetica', '', 9)
        self.set_text_color(70, 70, 70)
        self.multi_cell(180, 4.6, self._txt(valor))
        self.set_text_color(40, 40, 40)
        self.ln(2)

    def tabla_pares(self, filas, ancho_etq=70):
        """Tabla de dos columnas: etiqueta | valor."""
        self.set_font('Helvetica', '', 9)
        for etq, val in filas:
            self._espacio(8)
            self.set_font('Helvetica', 'B', 9)
            self.cell(ancho_etq, 6, self._txt(etq), 1, 0, 'L')
            self.set_font('Helvetica', '', 9)
            self.cell(180 - ancho_etq, 6, self._txt(val), 1, 1, 'L')

    def _lineas_necesarias(self, texto, ancho):
        """Cuántas líneas ocupa el texto en el ancho dado, con la fuente actual."""
        t = self._txt(texto).strip()
        if not t:
            return 1
        disponible = ancho - 2
        lineas, actual = 1, ''
        for palabra in t.split():
            tentativa = (actual + ' ' + palabra).strip()
            if self.get_string_width(tentativa) <= disponible:
                actual = tentativa
            else:
                # Palabra sola más ancha que la celda: se parte igual
                if not actual:
                    actual = palabra
                    continue
                lineas += 1
                actual = palabra
        return lineas

    def tabla(self, encabezados, anchos, filas, alineaciones=None):
        """Tabla con alto de fila variable: las celdas con texto largo
        envuelven en varias líneas en lugar de recortarse."""
        if not filas:
            return
        alineaciones = alineaciones or ['L'] * len(encabezados)
        ALTO_LINEA = 4.2

        def pintar_encabezado():
            self.set_font('Helvetica', 'B', 8)
            self.set_fill_color(60, 60, 60)
            self.set_text_color(255, 255, 255)
            for h, w in zip(encabezados, anchos):
                self.cell(w, 6.5, self._recortar(h, w), 1, 0, 'C', True)
            self.ln()
            self.set_text_color(40, 40, 40)
            self.set_font('Helvetica', '', 8)

        self._espacio(20)
        pintar_encabezado()

        for fila in filas:
            valores = [str(v) if v is not None else '' for v in fila]

            # El alto lo marca la celda que más líneas necesita
            n_lineas = max(self._lineas_necesarias(v, w) for v, w in zip(valores, anchos))
            alto = max(n_lineas * ALTO_LINEA, 5.5)

            if self.get_y() + alto > self.h - 22:
                self.add_page()
                pintar_encabezado()

            y0 = self.get_y()
            x  = self.l_margin

            for valor, w, al in zip(valores, anchos, alineaciones):
                # Recuadro completo primero, para que el borde cubra todo el alto
                self.rect(x, y0, w, alto)
                self.set_xy(x + 1, y0 + 0.6)
                self.multi_cell(w - 2, ALTO_LINEA, self._txt(valor), 0, al)
                x += w

            self.set_xy(self.l_margin, y0 + alto)

        self.ln(3)

    def _recortar(self, texto, ancho):
        t = self._txt(texto)
        if self.get_string_width(t) <= ancho - 2:
            return t
        while t and self.get_string_width(t + '...') > ancho - 2:
            t = t[:-1]
        return t + '...' if t else ''

    def bloque_evidencias(self, evidencias):
        """Registro fotográfico: dos por fila, con leyenda de origen."""
        if not evidencias:
            self.set_font('Helvetica', 'I', 9)
            self.set_text_color(150, 150, 150)
            self.cell(0, 6, self._txt('Sin evidencias fotograficas en el periodo.'), 0, 1)
            self.set_text_color(40, 40, 40)
            return

        ANCHO, GAP = 86.0, 8.0
        col = 0
        y_fila = self.get_y()
        alto_fila = 0

        for ev in evidencias[:INFORME_OBRA_MAX_FOTOS]:
            descargada = _io_imagen_temp(ev['url'])
            if not descargada:
                continue
            ruta, ratio = descargada

            alto_img = min(ANCHO * ratio, 62)
            alto_bloque = alto_img + 13

            if col == 0:
                self._espacio(alto_bloque)
                y_fila = self.get_y()
                alto_fila = 0

            x = 15 + col * (ANCHO + GAP)
            try:
                self.image(ruta, x=x, y=y_fila, w=ANCHO, h=alto_img)
            except Exception as e:
                print(f"[INFORME] Error insertando imagen: {e}")
            finally:
                try:
                    os.remove(ruta)      # se libera de inmediato
                except OSError:
                    pass

            # Leyenda: origen, fecha, frente
            self.set_xy(x, y_fila + alto_img + 1)
            self.set_font('Helvetica', 'B', 7)
            self.set_text_color(*self.rgb)
            self.cell(ANCHO, 3.5, self._recortar(ev['origen'], ANCHO), 0, 2)
            self.set_font('Helvetica', '', 7)
            self.set_text_color(110, 110, 110)
            self.cell(ANCHO, 3.5, self._recortar(ev['detalle'], ANCHO), 0, 2)
            if ev.get('nota'):
                self.cell(ANCHO, 3.5, self._recortar(ev['nota'], ANCHO), 0, 2)
            self.set_text_color(40, 40, 40)

            alto_fila = max(alto_fila, alto_bloque)
            col += 1
            if col == 2:
                col = 0
                self.set_y(y_fila + alto_fila + 5)

        if col == 1:
            self.set_y(y_fila + alto_fila + 5)

        total = len(evidencias)
        if total > INFORME_OBRA_MAX_FOTOS:
            self.set_font('Helvetica', 'I', 8)
            self.set_text_color(150, 150, 150)
            self.cell(0, 5, self._txt(
                f'Se muestran {INFORME_OBRA_MAX_FOTOS} de {total} evidencias.'), 0, 1)
            self.set_text_color(40, 40, 40)

@app.route('/api/informe-obra/<int:proyecto_id>')
def informe_obra(proyecto_id):
    """Genera el informe de avance de obra en PDF.
    Filtros: actividades y novedades por su campo de fecha propio;
    los cortes por created_at, porque sus fechas son de ejecución de obra."""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    # El informe está construido sobre la estructura de formularios de
    # Ingeniarcol (ids fijos en INFORME_OBRA_CFG), así que no aplica a
    # otras empresas.
    if session.get('empresa_id') != EMPRESA_INFORME_OBRA:
        return jsonify({'error': 'Función no disponible para esta empresa'}), 403

    desde     = request.args.get('desde') or None
    hasta     = request.args.get('hasta') or None
    planeado  = _io_num(request.args.get('planeado'))
    ejecutado = _io_num(request.args.get('ejecutado'))
    variacion = round(ejecutado - planeado, 2)

    C = INFORME_OBRA_CFG

    try:
        with db_connection() as (conn, cursor):
            # ── Datos del contrato (tipo de proyecto) ──
            cursor.execute("""
                SELECT p.nombre_proyecto, p.datos_tipo, e.nombre, e.logo_url, e.color_primario
                FROM proyectos p
                JOIN empresas e ON e.id = p.empresa_id
                WHERE p.id = %s AND p.empresa_id = %s
            """, (proyecto_id, session.get('empresa_id')))
            fila = cursor.fetchone()
            if not fila:
                return jsonify({'error': 'Proyecto no encontrado'}), 404

            nombre_proy, datos_tipo, empresa_nombre, logo_url, color = fila
            datos_tipo = datos_tipo or {}
            cp = C['proyecto']

            # ── Frentes: todos los del proyecto, sin filtro de fecha ──
            cursor.execute("""
                SELECT id, respuestas FROM respuestas_formulario
                WHERE id_proyecto = %s AND formulario_id = %s
                ORDER BY created_at
            """, (proyecto_id, C['apertura']['formulario_id']))
            frentes = [{'id': r[0], 'r': r[1] or {}} for r in cursor.fetchall()]

            # ── Cortes: filtrados por created_at ──
            sql = """SELECT respuestas, created_at FROM respuestas_formulario
                     WHERE id_proyecto = %s AND formulario_id = %s"""
            par = [proyecto_id, C['corte']['formulario_id']]
            if desde:
                sql += " AND created_at::date >= %s"; par.append(desde)
            if hasta:
                sql += " AND created_at::date <= %s"; par.append(hasta)
            sql += " ORDER BY created_at ASC"
            cursor.execute(sql, par)
            cortes_raw = cursor.fetchall()

            # ── Actividades y novedades: por su campo de fecha propio ──
            def por_fecha(cfg):
                s = """SELECT respuestas FROM respuestas_formulario
                       WHERE id_proyecto = %s AND formulario_id = %s"""
                p = [proyecto_id, cfg['formulario_id']]
                if desde:
                    s += " AND COALESCE(NULLIF(respuestas->>%s,''),'9999-12-31')::date >= %s"
                    p += [cfg['fecha'], desde]
                if hasta:
                    s += " AND COALESCE(NULLIF(respuestas->>%s,''),'1900-01-01')::date <= %s"
                    p += [cfg['fecha'], hasta]
                s += " ORDER BY created_at"
                cursor.execute(s, p)
                return [r[0] or {} for r in cursor.fetchall()]

            actividades = por_fecha(C['actividad'])
            novedades   = por_fecha(C['novedad'])

        # ── Agrupar por frente usando 180_codigo (id del registro de apertura) ──
        cc = C['corte']
        cortes_por_frente = {}
        for resp, _ in cortes_raw:
            resp = resp or {}
            cortes_por_frente[str(resp.get(cc['frente_codigo'], ''))] = resp  # el último gana

        ca = C['actividad']
        act_por_frente = {}
        for a in actividades:
            act_por_frente.setdefault(str(a.get(ca['frente_codigo'], '')), []).append(a)

        # ── Construir el PDF ──
        rgb  = _hex_a_rgb(color or '#FFAF33')
        logo = _io_descargar_logo(logo_url)
        contrato_no = datos_tipo.get(cp['contrato_no'], '')

        pdf = _PDFInformeObra(contrato_no, logo, rgb, empresa_nombre or '')
        pdf.alias_nb_pages()
        pdf.add_page()

        # Encabezado del contrato
        pdf.set_font('Helvetica', 'B', 14)
        pdf.multi_cell(180, 7, pdf._txt(datos_tipo.get(cp['nombre'], nombre_proy)), 0, 'C')
        pdf.ln(2)
        ahora = datetime.now(pytz.timezone('America/Bogota'))
        rango = f"Del {_io_fecha(desde)} al {_io_fecha(hasta)}" if (desde and hasta) else 'Periodo completo'
        pdf.set_font('Helvetica', '', 9)
        pdf.set_text_color(110, 110, 110)
        pdf.cell(180, 5, pdf._txt(f"{rango}  ·  Generado {ahora.strftime('%d/%m/%Y %I:%M %p')}"), 0, 1, 'C')
        pdf.set_text_color(40, 40, 40)
        pdf.ln(3)

        pdf.campo_largo('OBJETO DEL CONTRATO', datos_tipo.get(cp['objeto'], ''))
        pdf.campo_largo('ALCANCE', datos_tipo.get(cp['alcance'], ''))

        pdf.tabla_pares([
            ('CONTRATO No.',     contrato_no),
            ('CONTRATISTA',      datos_tipo.get(cp['contratista'], '')),
            ('ESTADO DEL CONTRATO', datos_tipo.get(cp['estado'], '')),
            ('% PLANEADO',       f"{planeado:.2f}%"),
            ('% EJECUTADO',      f"{ejecutado:.2f}%"),
            ('% VARIACION',      f"{variacion:+.2f}%"),
        ])

        # ── Un bloque por frente ──
        pdf.titulo_seccion('AVANCES POR FRENTE / COMPONENTE')
        ap = C['apertura']

        evidencias = []

        for fr in frentes:
            r  = fr['r']
            co = cortes_por_frente.get(str(fr['id']), {})

            pdf._espacio(40)
            pdf.ln(2)
            pdf.set_font('Helvetica', 'B', 11)
            pdf.cell(0, 6, pdf._txt(r.get(ap['frente'], 'Frente sin nombre')), 0, 1)
            pdf.set_font('Helvetica', '', 8)
            pdf.set_text_color(110, 110, 110)
            ubic = ' · '.join(x for x in [r.get(ap['ubicacion'], ''), r.get(ap['tramo'], '')] if x)
            if ubic:
                pdf.cell(0, 4.5, pdf._txt(ubic), 0, 1)
            pdf.set_text_color(40, 40, 40)
            pdf.ln(1)

            pdf.tabla(
                ['TIPO DE ELEMENTO', 'TIPO DE INTERVENCION', 'SUB ETAPA ACTUAL', 'ESTADO ACTUAL'],
                [45, 55, 40, 40],
                [[r.get(ap['tipo_elemento'], ''), r.get(ap['tipo_intervencion'], ''),
                  co.get(cc['sub_etapa'], ''), co.get(cc['estado_frente'], '')]]
            )

            # Fechas y avances lado a lado
            pdf._espacio(34)
            y0 = pdf.get_y()
            pdf.set_font('Helvetica', 'B', 8)
            pdf.cell(88, 5.5, 'FECHAS', 1, 0, 'C')
            pdf.cell(4, 5.5, '', 0, 0)
            pdf.cell(88, 5.5, 'AVANCES', 1, 1, 'C')

            fechas = [
                ('Inicio programada',      _io_fecha(r.get(ap['fecha_ini_prog']))),
                ('Inicio real',            _io_fecha(co.get(cc['fecha_ini_real']))),
                ('Finalizacion programada',_io_fecha(r.get(ap['fecha_fin_prog']))),
                ('Finalizacion real',      _io_fecha(co.get(cc['fecha_fin_real']))),
                ('Dias atraso (-) / adelanto (+)', co.get(cc['dias_atraso'], '')),
            ]
            avances = [
                ('% Financiero programado acum.', f"{_io_num(co.get(cc['pct_fin_prog'])):.2f}" if co else ''),
                ('% Financiero ejecutado acum.',  f"{_io_num(co.get(cc['pct_fin_ejec'])):.2f}" if co else ''),
                ('Ejecucion financiera programada', _io_money(co.get(cc['ejec_fin_prog'])) if co else ''),
                ('Ejecucion financiera real',       _io_money(co.get(cc['ejec_fin_real'])) if co else ''),
                ('', ''),
            ]
            pdf.set_font('Helvetica', '', 7.5)
            for (e1, v1), (e2, v2) in zip(fechas, avances):
                yf = pdf.get_y()
                pdf.cell(50, 5, pdf._recortar(e1, 50), 1, 0)
                pdf.cell(38, 5, pdf._recortar(str(v1), 38), 1, 0)
                pdf.cell(4, 5, '', 0, 0)
                pdf.cell(52, 5, pdf._recortar(e2, 52), 1, 0)
                pdf.cell(36, 5, pdf._recortar(str(v2), 36), 1, 1, 'R')
            pdf.ln(3)

            if not co:
                pdf.set_font('Helvetica', 'I', 8)
                pdf.set_text_color(150, 150, 150)
                pdf.cell(0, 5, pdf._txt('Sin corte mensual registrado en el periodo.'), 0, 1)
                pdf.set_text_color(40, 40, 40)
                pdf.ln(2)

            # Materiales: se agregan de todas las actividades del frente
            acumulado = {}
            for a in act_por_frente.get(str(fr['id']), []):
                for b in (a.get('__repeticiones') or {}).get(ca['grupo_material'], []):
                    clave = (b.get(ca['mat_pavimento'], ''), b.get(ca['mat_uso'], ''),
                             b.get(ca['mat_material'], ''), b.get(ca['mat_unidad'], ''))
                    if not any(clave):
                        continue
                    acumulado[clave] = acumulado.get(clave, 0) + _io_num(b.get(ca['mat_cantidad']))

            if acumulado:
                pdf.set_font('Helvetica', 'B', 9)
                pdf.cell(0, 5.5, 'MATERIALES', 0, 1)
                pdf.tabla(
                    ['TIPO DE PAVIMENTO', 'USO DEL MATERIAL', 'MATERIAL', 'CANTIDAD', 'UNIDAD'],
                    [42, 42, 46, 25, 25],
                    [[k[0], k[1], k[2], f"{v:g}", k[3]] for k, v in sorted(acumulado.items())],
                    ['L', 'L', 'L', 'R', 'L']
                )

            # Metas físicas del corte
            metas = (co.get('__repeticiones') or {}).get(cc['grupo_meta'], []) if co else []
            if metas:
                pdf.set_font('Helvetica', 'B', 9)
                pdf.cell(0, 5.5, 'AVANCES POR META FISICA', 0, 1)
                pdf.tabla(
                    ['META FISICA', 'UNIDAD', 'PROGRAMADO', 'EJECUTADO', 'OBSERVACION'],
                    [58, 22, 25, 25, 50],
                    [[m.get(cc['meta'], ''), m.get(cc['meta_unidad'], ''),
                      f"{_io_num(m.get(cc['meta_prog'])):g}", f"{_io_num(m.get(cc['meta_ejec'])):g}",
                      m.get(cc['meta_obs'], '')] for m in metas],
                    ['L', 'L', 'R', 'R', 'L']
                )

            # Actividades del frente
            acts = act_por_frente.get(str(fr['id']), [])
            if acts:
                pdf.set_font('Helvetica', 'B', 9)
                pdf.cell(0, 5.5, 'ACTIVIDADES EJECUTADAS', 0, 1)
                pdf.tabla(
                    ['FECHA', 'DESCRIPCION', 'RESPONSABLE'],
                    [25, 110, 45],
                    [[_io_fecha(a.get(ca['fecha'])), a.get(ca['descripcion'], ''),
                      a.get(ca['responsable'], '')] for a in acts]
                )

            nombre_frente = r.get(ap['frente'], 'Frente sin nombre')

            # Evidencias de actividades
            for a in acts:
                for u in _io_evidencias_de(a):
                    evidencias.append({
                        'url':     u,
                        'origen':  'EVIDENCIA DE ACTIVIDAD',
                        'detalle': f"{_io_fecha(a.get(ca['fecha']))} · {nombre_frente}",
                        'nota':    a.get(ca['descripcion'], '')
                    })

            # Evidencias del avance físico (dentro del corte)
            for m in metas:
                for u in _io_evidencias_de(m):
                    evidencias.append({
                        'url':     u,
                        'origen':  'EVIDENCIA DE AVANCE FISICO',
                        'detalle': nombre_frente,
                        'nota':    m.get(cc['meta'], '')
                    })

        # ── Novedades ──
        cn = C['novedad']
        pdf.titulo_seccion('NOVEDADES')
        if novedades:
            pdf.tabla(
                ['FECHA', 'TIPO', 'IMPACTO', 'ESTADO', 'DESCRIPCION'],
                [22, 32, 22, 28, 76],
                [[_io_fecha(n.get(cn['fecha'])), n.get(cn['tipo'], ''), n.get(cn['impacto'], ''),
                  n.get(cn['estado'], ''), n.get(cn['descripcion'], '')] for n in novedades]
            )
        else:
            pdf.set_font('Helvetica', 'I', 9)
            pdf.set_text_color(150, 150, 150)
            pdf.cell(0, 6, pdf._txt('Sin novedades registradas en el periodo.'), 0, 1)
            pdf.set_text_color(40, 40, 40)

        # ── Registro fotográfico ──
        for n in novedades:
            for u in _io_evidencias_de(n):
                evidencias.append({
                    'url':     u,
                    'origen':  'EVIDENCIA DE NOVEDAD',
                    'detalle': f"{_io_fecha(n.get(cn['fecha']))} · {n.get(cn['tipo'], '')}",
                    'nota':    n.get(cn['descripcion'], '')
                })

        print(f"[INFORME] Evidencias recolectadas: {len(evidencias)}")
        for ev in evidencias:
            print(f"   - {ev['origen']} | {ev['url'][:70]}")
        pdf.titulo_seccion('REGISTRO FOTOGRAFICO')
        pdf.bloque_evidencias(evidencias)

        salida = pdf.output(dest='S')
        if isinstance(salida, str):
            salida = salida.encode('latin-1')

        if logo:
            try:
                os.remove(logo)
            except OSError:
                pass

        return send_file(BytesIO(salida), mimetype='application/pdf',
                         as_attachment=False,
                         download_name=f"Informe_{contrato_no or proyecto_id}.pdf")

    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': f'Error generando el informe: {e}'}), 500


def _io_descargar_logo(logo_url):
    """Descarga el logo y lo normaliza a PNG sobre fondo blanco.
    FPDF no maneja transparencia."""
    if not logo_url:
        return None
    try:
        r = requests.get(logo_url, timeout=8)
        if not r.ok:
            return None
        tmp = NamedTemporaryFile(delete=False, suffix='.png')
        im = Image.open(BytesIO(r.content)).convert('RGBA')
        fondo = Image.new('RGBA', im.size, (255, 255, 255, 255))
        fondo.alpha_composite(im)
        fondo.convert('RGB').save(tmp.name, 'PNG')
        return tmp.name
    except Exception as e:
        print(f"[INFORME] No se pudo cargar el logo: {e}")
        return None

# ── Utilidad: generar slug ──────────────────────────────────────
def generar_slug(nombre_empresa):
    # Normalizar: quitar tildes, minúsculas, reemplazar espacios
    nfkd = unicodedata.normalize('NFD', nombre_empresa)
    sin_tildes = ''.join(c for c in nfkd if not unicodedata.combining(c))
    slug = sin_tildes.lower()
    slug = re.sub(r'[^a-z0-9\s-]', '', slug)
    slug = re.sub(r'\s+', '-', slug.strip())
    return slug
 

# ── Formulario Lote: GET ────────────────────────────────────────
@app.route('/formulario-lote')
def formulario_lote():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    project_id = request.args.get('project_id')
    if not project_id:
        return redirect(url_for('registros'))

    color_primario   = '#FFAF33'
    color_secundario = '#E3E3E3'
    try:
        with db_connection() as (conn, cursor):
            cursor.execute("""
                SELECT color_primario, color_secundario
                FROM empresas WHERE id = %s
            """, (session.get('empresa_id'),))
            row = cursor.fetchone()
            if row:
                color_primario   = row[0] or '#FFAF33'
                color_secundario = row[1] or '#E3E3E3'
    except Exception as e:
        print(f"Error en formulario-lote GET: {e}")

    return render_template('formLote.html',
                           project_id=project_id,
                           color_primario=color_primario,
                           color_secundario=color_secundario)


# ── Formulario Lote: POST ───────────────────────────────────────
@app.route('/api/registros-lote', methods=['POST'])
def crear_registro_lote():
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    try:
        data = request.get_json()

        campos_requeridos = [
            'project_id', 'fecha_produccion', 'turno', 'planta_linea',
            'orden_produccion', 'codigo_producto', 'numero_lote',
            'cantidad_programada', 'cantidad_real_producida',
            'responsable_produccion', 'supervisor'
        ]
        for campo in campos_requeridos:
            if not data.get(campo):
                return jsonify({'error': f'El campo {campo} es obligatorio'}), 400

        with db_connection() as (conn, cursor):
            cursor.execute("""
                INSERT INTO registros_lote (
                    proyecto_id, empresa_id, user_id,
                    fecha_produccion, turno, planta_linea,
                    orden_produccion, codigo_producto, numero_lote,
                    cantidad_programada, cantidad_real_producida,
                    responsable_produccion, supervisor
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            """, (
                data['project_id'],
                session.get('empresa_id'),
                session['user_id'],
                data['fecha_produccion'],
                data['turno'],
                data['planta_linea'],
                data['orden_produccion'],
                data['codigo_producto'],
                data['numero_lote'],
                data['cantidad_programada'],
                data['cantidad_real_producida'],
                data['responsable_produccion'],
                data['supervisor']
            ))
            nuevo_id = cursor.fetchone()[0]
            return jsonify({'success': True, 'id': nuevo_id})

    except Exception as e:
        print(f"Error en registro lote: {e}")
        return jsonify({'error': str(e)}), 500

 
@app.route('/api/interpretar-respuesta', methods=['POST'])
def interpretar_respuesta():
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    try:
        data          = request.get_json()
        pregunta      = data.get('pregunta', '')
        respuesta     = data.get('respuesta', '')
        campo_id      = data.get('campo_id', '')
        campo_nombre  = data.get('campo_nombre', '')
        campo_tipo    = data.get('campo_tipo', 'texto_corto')

        fecha_hoy = date.today().strftime('%Y-%m-%d')

        prompt = f"""El sistema le preguntó al usuario: "{pregunta}"
El usuario respondió: "{respuesta}"

El campo a llenar es: "{campo_nombre}" (tipo: {campo_tipo})
Fecha actual: {fecha_hoy}

Extrae el valor que corresponde al campo. Devuelve SOLO un JSON:
{{"valor": "el valor extraído"}}

Reglas:
- Si el tipo es fecha, devuelve en formato YYYY-MM-DD.
- Si el tipo es numero/moneda/porcentaje, devuelve solo el número.
- Si el tipo es booleano, devuelve true o false.
- Si no puedes extraer un valor claro, devuelve {{"valor": ""}}
- No inventes datos. Solo extrae lo que el usuario dijo."""

        response = openai_client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "user", "content": prompt}
            ],
            temperature=0.1,
            max_tokens=200
        )

        respuesta_texto = response.choices[0].message.content.strip()
        respuesta_texto = respuesta_texto.replace('```json', '').replace('```', '').strip()
        resultado = json.loads(respuesta_texto)

        print(f"[INTERPRETAR] Campo: {campo_nombre} → Valor: {resultado.get('valor', '')}")

        return jsonify({'success': True, 'valor': resultado.get('valor', '')})

    except Exception as e:
        print(f"[INTERPRETAR] Error: {e}")
        return jsonify({'error': str(e), 'success': False}), 500

# ── Ruta GET: mostrar formulario ────────────────────────────────
@app.route('/registroEmpresa', methods=['GET'])
def registro_page():
    token = request.args.get('token', '')
 
    if not token:
        return render_template('registroEmpresa.html', token_valido=False, token='', email_sugerido='')
 
    try:
        with db_connection() as (conn, cursor):
            cursor.execute("""
                SELECT id, email, usado, expira_en
                FROM tokens_registro
                WHERE token = %s
            """, (token,))
            row = cursor.fetchone()
 
        if not row:
            return render_template('registroEmpresa.html', token_valido=False, token='', email_sugerido='')
 
        _, email_sugerido, usado, expira_en = row
 
        if usado:
            return render_template('registroEmpresa.html', token_valido=False, token='', email_sugerido='')
 
        # Verificar expiración
        ahora = datetime.now(timezone.utc)
        if expira_en.tzinfo is None:
            expira_en = expira_en.replace(tzinfo=timezone.utc)
 
        if ahora > expira_en:
            return render_template('registroEmpresa.html', token_valido=False, token='', email_sugerido='')
 
        return render_template('registroEmpresa.html',
                               token_valido=True,
                               token=token,
                               email_sugerido=email_sugerido or '')
 
    except Exception as e:
        print(f"Error validando token: {e}")
        return render_template('registroEmpresa.html', token_valido=False, token='', email_sugerido='')
 
 
# ── Ruta POST: procesar registro ────────────────────────────────
@app.route('/registroEmpresa', methods=['POST'])
def registro_post():
    data     = request.get_json()
    token    = data.get('token', '')
    nombre   = data.get('nombre', '').strip()
    apellido = data.get('apellido', '').strip()
    email    = data.get('email', '').strip()
    password = data.get('password', '')
    cargo    = data.get('cargo', '').strip()
    empresa  = data.get('empresa', '').strip()

    if not all([token, nombre, apellido, email, password, empresa]):
        return jsonify({'error': 'Faltan campos obligatorios'}), 400

    if len(password) < 8:
        return jsonify({'error': 'La contraseña debe tener al menos 8 caracteres'}), 400

    try:
        with db_connection() as (conn, cursor):

            # 1. Validar token
            cursor.execute("""
                SELECT id, usado, expira_en
                FROM tokens_registro
                WHERE token = %s
            """, (token,))
            row = cursor.fetchone()

            if not row:
                return jsonify({'error': 'Token inválido'}), 400

            token_id, usado, expira_en = row

            if usado:
                return jsonify({'error': 'Este enlace ya fue utilizado'}), 400

            ahora = datetime.now(timezone.utc)
            if expira_en.tzinfo is None:
                expira_en = expira_en.replace(tzinfo=timezone.utc)

            if ahora > expira_en:
                return jsonify({'error': 'Este enlace ha expirado'}), 400

            # 2. Verificar que el email no exista
            cursor.execute(
                "SELECT user_id FROM usuario WHERE email = %s", (email,)
            )
            if cursor.fetchone():
                return jsonify({'error': 'Este correo ya está registrado'}), 400

            # 3. Generar slug único para la empresa
            slug_base = generar_slug(empresa)
            slug      = slug_base
            contador  = 1
            while True:
                cursor.execute(
                    "SELECT id FROM empresas WHERE slug = %s", (slug,)
                )
                if not cursor.fetchone():
                    break
                slug = f"{slug_base}-{contador}"
                contador += 1

            # 4. Crear empresa
            cursor.execute("""
                INSERT INTO empresas (nombre, slug)
                VALUES (%s, %s)
                RETURNING id
            """, (empresa, slug))
            empresa_id = cursor.fetchone()[0]

            # 5. Crear usuario admin
            hashed = generate_password_hash(password)
            cursor.execute("""
                INSERT INTO usuario
                    (name, apellido, email, password, cargo, rol, empresa_id, estado)
                VALUES (%s, %s, %s, %s, %s, 'admin', %s, 'activo')
                RETURNING user_id
            """, (nombre, apellido, email, hashed, cargo or 'Administrador', empresa_id))

            nuevo_user_id = cursor.fetchone()[0]

            session['user_id']    = nuevo_user_id
            session['user_rol']   = 'admin'
            session['empresa_id'] = empresa_id
            session['user_name']  = nombre

            # 6. Marcar token como usado
            cursor.execute("""
                UPDATE tokens_registro
                SET usado = TRUE
                WHERE id = %s
            """, (token_id,))

        return jsonify({'success': True})

    except Exception as e:
        print(f"Error en registro: {e}")
        return jsonify({'error': str(e)}), 500
 
 
# ── Ruta: generar token de invitación (solo admins IAC) ─────────
@app.route('/generar-token-registro', methods=['POST'])
def generar_token_registro():
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
 
    # Solo el admin de IAC (empresa_id = 1) puede generar tokens
    if session.get('empresa_id') != 1:
        return jsonify({'error': 'No tienes permisos para esta acción'}), 403
 
    data  = request.get_json()
    email = data.get('email', '').strip()
 
    if not email:
        return jsonify({'error': 'El correo es obligatorio'}), 400
 
    try:
        token     = secrets.token_urlsafe(32)
        expira_en = datetime.now(timezone.utc) + timedelta(days=7)
 
        with db_connection() as (conn, cursor):
            cursor.execute("""
                INSERT INTO tokens_registro (token, email, expira_en)
                VALUES (%s, %s, %s)
            """, (token, email, expira_en))
 
        link = f"https://bitacoraiac.onrender.com/registroEmpresa?token={token}"
 
        return jsonify({
            'success': True,
            'token':   token,
            'link':    link,
            'expira':  expira_en.strftime('%d/%m/%Y %H:%M')
        })
 
    except Exception as e:
        print(f"Error generando token: {e}")
        return jsonify({'error': str(e)}), 500


# ========================================
# OBTENER TOKEN DE BENTLEY
# ========================================
def obtener_token_synchro():
    """Obtiene token de acceso de Bentley IMS"""
    try:
        payload = {
            'grant_type': 'client_credentials',
            'client_id': SYNCHRO_CONFIG['client_id'],
            'client_secret': SYNCHRO_CONFIG['client_secret'],
            'scope': 'itwin-platform'
        }
        
        response = requests.post(SYNCHRO_CONFIG['token_url'], data=payload, timeout=10)
        
        if response.status_code == 200:
            token = response.json().get('access_token')
            print("✅ Token obtenido")
            return token
        else:
            print(f"❌ Error obteniendo token: {response.status_code}")
            return None
            
    except Exception as e:
        print(f"❌ Excepción: {str(e)}")
        return None

# ========================================
# ENVIAR ACTIVIDADES A SYNCHRO
# ========================================
def enviar_actividades_synchro(token, data):
    """Envía todas las actividades al formulario de Synchro"""
    try:
        headers = {
            'Authorization': f'Bearer {token}',
            'Accept': 'application/vnd.bentley.itwin-platform.v2+json',
            'Prefer': 'return=representation',
            'Content-Type': 'application/json'
        }
        
        # 1. Obtener formulario actual
        url_form = f"{SYNCHRO_CONFIG['forms_url']}/{SYNCHRO_CONFIG['form_id']}"
        response = requests.get(url_form, headers=headers, timeout=10)
        
        if response.status_code != 200:
            return {'success': False, 'error': f'No se pudo obtener formulario: {response.status_code}'}
        
        form_actual = response.json().get('form', {})
        props = form_actual.get('properties', {})
        
        # 2. Actualizar propiedades básicas
        props['Codigo Proyecto'] = data['codigo_proyecto']
        props['Contratista'] = data['contratista']
        props['Contrato'] = data['contrato']
        
        # 3. Sección 1: Actividades finalizadas
        actividades_finalizadas = props.get('Actividades finalizadas', [])
        for act in data.get('actividades_finalizadas', []):
            nueva_act = {
                'id': str(uuid.uuid4()),
                '__x00cd__tem': act['item'],
                'Descripci__x00f3__n': act['descripcion'],
                'Observaciones__x0020__actividades__x': act['observaciones']
            }
            actividades_finalizadas.append(nueva_act)
        props['Actividades finalizadas'] = actividades_finalizadas
        
        # 4. Sección 2: Actividades pendientes por culminar
        actividades_pendientes = props.get('Actividades pendientes', [])
        for act in data.get('actividades_pendientes', []):
            nueva_act = {
                'id': str(uuid.uuid4()),
                '__x00cd__tem__x0020__Pendiente': act['item'],
                'Descripci__x00f3__n__x0020__pendient': act['descripcion'],
                'Pendiente__x0020__generado': act.get('pendiente_generado', ''),
                'Observaciones__x0020__pendientes': act['observaciones']
            }
            actividades_pendientes.append(nueva_act)
        props['Actividades pendientes'] = actividades_pendientes
        
        # 5. Sección 3: Actividades pendientes por facturar
        # Nota: Necesitarás el nombre exacto de este campo en Synchro
        # Por ahora lo dejo como ejemplo
        if 'actividades_facturar' in data and data['actividades_facturar']:
            actividades_facturar = props.get('Actividades pendientes por facturar', [])
            for act in data['actividades_facturar']:
                nueva_act = {
                    'id': str(uuid.uuid4()),
                    '__x00cd__tem': act['item'],
                    'Descripci__x00f3__n': act['descripcion'],
                    'Cantidad_contractual': act['cantidad_contractual'],
                    'Cantidad_facturada': act['cantidad_facturada'],
                    'Cantidad_pendiente': act['cantidad_pendiente'],
                    'Observaci__x00f3__n': act['observacion']
                }
                actividades_facturar.append(nueva_act)
            props['Actividades pendientes por facturar'] = actividades_facturar
        
        # 6-8. Secciones de documentación (similar estructura)
        # Agregar según los nombres exactos de los campos en Synchro
        
        # 9. Enviar actualización
        cambios = {'properties': props}
        response_update = requests.patch(url_form, headers=headers, json=cambios, timeout=15)
        
        if response_update.status_code == 200:
            print("✅ Formulario actualizado en Synchro")
            return {'success': True, 'form_id': SYNCHRO_CONFIG['form_id']}
        else:
            error_msg = response_update.text
            print(f"❌ Error actualizando: {error_msg}")
            return {'success': False, 'error': error_msg}
            
    except Exception as e:
        print(f"❌ Excepción: {str(e)}")
        return {'success': False, 'error': str(e)}

# ========================================
# SUBIR ATTACHMENTS A SYNCHRO
# ========================================
def subir_attachments_synchro(token, fotos, videos):
    """Sube fotos y videos como adjuntos al formulario"""
    try:
        headers = {
            'Authorization': f'Bearer {token}',
            'Accept': 'application/vnd.bentley.itwin-platform.v2+json'
        }
        
        url_attachments = f"{SYNCHRO_CONFIG['forms_url']}/{SYNCHRO_CONFIG['form_id']}/attachments"
        
        contador = 0
        
        # Subir fotos
        for i, foto_base64 in enumerate(fotos[:10]):  # Máximo 10 fotos
            try:
                if ',' in foto_base64:
                    foto_base64 = foto_base64.split(',')[1]
                
                foto_bytes = base64.b64decode(foto_base64)
                
                files = {
                    'file': (f'foto_{i+1}.jpg', io.BytesIO(foto_bytes), 'image/jpeg')
                }
                
                data = {
                    'caption': f'Foto {i+1} - Evidencia'
                }
                
                response = requests.post(url_attachments, headers=headers, files=files, data=data, timeout=30)
                
                if response.status_code == 201:
                    contador += 1
                    print(f"✅ Foto {i+1} subida")
                else:
                    print(f"⚠️ Error subiendo foto {i+1}")
                    
            except Exception as e:
                print(f"⚠️ Error procesando foto {i+1}: {str(e)}")
                continue
        
        # Subir videos
        for i, video_base64 in enumerate(videos[:5]):  # Máximo 5 videos
            try:
                if ',' in video_base64:
                    video_base64 = video_base64.split(',')[1]
                
                video_bytes = base64.b64decode(video_base64)
                
                files = {
                    'file': (f'video_{i+1}.webm', io.BytesIO(video_bytes), 'video/webm')
                }
                
                data = {
                    'caption': f'Video {i+1} - Evidencia'
                }
                
                response = requests.post(url_attachments, headers=headers, files=files, data=data, timeout=60)
                
                if response.status_code == 201:
                    contador += 1
                    print(f"✅ Video {i+1} subido")
                else:
                    print(f"⚠️ Error subiendo video {i+1}")
                    
            except Exception as e:
                print(f"⚠️ Error procesando video {i+1}: {str(e)}")
                continue
        
        return contador
        
    except Exception as e:
        print(f"❌ Error en subir_attachments: {str(e)}")
        return 0


@app.route('/guardar-configuracion', methods=['POST'])
def guardar_configuracion():
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = None
    try:
        data             = request.get_json()
        color_primario   = data.get('color_primario', '#FBAF33')
        color_secundario = data.get('color_secundario', '#E3E3E3')
        logo             = data.get('logo', '')

        conn, cursor = get_db_connection()
        cursor.execute("""
            UPDATE empresas
            SET color_primario   = %s,
                color_secundario = %s,
                logo_url         = %s
            WHERE id = %s
        """, (color_primario, color_secundario, logo, session.get('empresa_id')))
        conn.commit()

        return jsonify({'success': True})
    except Exception as e:
        print(f"Error guardando configuración: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            cursor.close()
            connection_pool.putconn(conn)

@app.route('/guardar-formulario', methods=['POST'])
def guardar_formulario():
    """Recibe datos del frontend y los envía a Synchro"""
    try:
        data = request.json
        print("📥 Datos recibidos del frontend")
        
        # Validar que al menos venga UNA sección con datos
        secciones_con_datos = sum([
            1 if data.get('actividades_finalizadas') else 0,
            1 if data.get('actividades_pendientes') else 0,
            1 if data.get('actividades_facturar') else 0,
            1 if data.get('documentacion_seguridad') else 0,
            1 if data.get('documentacion_ambiental') else 0,
            1 if data.get('documentacion_calidad') else 0
        ])
        
        if secciones_con_datos == 0:
            return jsonify({
                'success': False,
                'error': 'Debes llenar al menos una sección del formulario'
            }), 400
        
        print(f"✅ Validación OK: {secciones_con_datos} sección(es) con datos")
        
        # 1. Obtener token
        token = obtener_token_synchro()
        if not token:
            return jsonify({
                'success': False,
                'error': 'No se pudo obtener token de Synchro'
            }), 500
        
        # 2. Enviar actividades a Synchro
        resultado = enviar_actividades_synchro(token, data)
        if not resultado['success']:
            return jsonify(resultado), 500
        
        # 3. Subir fotos/videos si existen
        fotos = data.get('fotos', [])
        videos = data.get('videos', [])
        attachments_subidos = 0
        
        if fotos or videos:
            attachments_subidos = subir_attachments_synchro(token, fotos, videos)
        
        # 4. Retornar éxito
        return jsonify({
            'success': True,
            'mensaje': 'Registro guardado en Synchro exitosamente',
            'form_id': SYNCHRO_CONFIG['form_id'],
            'attachments_subidos': attachments_subidos,
            'secciones_guardadas': secciones_con_datos
        }), 200
        
    except Exception as e:
        print(f"❌ Error en /guardar-formulario: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

'''
def create_user(nombre, apellido, email, password, cargo, rol, empresa):
    try:
        conn, cursor = get_db_connection()
        
        hashed_password = generate_password_hash(password)
        
        cursor.execute(
            """INSERT INTO usuario (name, apellido, email, password, cargo, rol, empresa)
               VALUES (%s, %s, %s, %s, %s, %s, %s) RETURNING user_id""",
            (nombre, apellido, email, hashed_password, cargo, rol, empresa)
        )
        
        user_id = cursor.fetchone()[0]
        conn.commit()
        return user_id
    except psycopg2.Error as e:
        print(f"Error al crear usuario: {e}")
        return None
    finally:
        if conn:
            cursor.close()
            connection_pool.putconn(conn)
'''

def verify_user(email, password):
    conn = None
    try:
        conn = connection_pool.getconn()
        cursor = conn.cursor()
        # Sin SET empresa_id porque aún no hay sesión
        cursor.execute(
            "SELECT user_id, password, rol, empresa_id, name, estado FROM usuario WHERE email = %s",
            (email,)
        )
        user = cursor.fetchone()
        if user and check_password_hash(user[1], password):
            return {
                'user_id':    user[0],
                'rol':        user[2],
                'empresa_id': user[3],
                'name':       user[4],
                'estado':     user[5]
            }
        return None
    except Exception as e:
        print(f"Error al verificar usuario: {e}")
        return None
    finally:
        if conn:
            cursor.close()
            connection_pool.putconn(conn)

def insert_registro_bitacora(respuestas, id_proyecto, fotos=None, videos=None):
    """
    Inserta un nuevo registro de bitácora, junto con sus fotos y videos asociados
    y sus descripciones, en la base de datos.
    """
    conn = None  # Definimos conn aquí para asegurarnos de que exista en el bloque finally
    try:
        conn, cursor = get_db_connection()

        # CAMBIO 1: Simplificamos el INSERT principal.
        # - Eliminamos la columna 'foto_base64' que ya es obsoleta.
        # - Cambiamos los nombres de las claves para que coincidan con tu formulario.
        cursor.execute("""
            INSERT INTO registrosbitacoraeqing (
                zona_intervencion, -- Mapeado desde "Tipo de informe"
                items,             -- Mapeado desde "Sede"
                metros_lineales,   -- Mapeado desde "Repuestos utilizados"
                proximas_tareas,   -- Mapeado desde "Repuestos a cotizar"
                id_proyecto
            )
            VALUES (%s, %s, %s, %s, %s)
            RETURNING id_registro
        """, (
            respuestas.get('zona_intervencion'),
            respuestas.get('items'),
            respuestas.get('metros_lineales'),
            respuestas.get('proximas_tareas'),
            id_proyecto,
        ))
        id_registro = cursor.fetchone()[0]

        # CAMBIO 2: Actualizamos el bucle para que maneje objetos (archivo + descripción).
        # Ahora esperamos una lista de diccionarios, no solo una lista de strings.
        for foto_obj in fotos or []:
            file_data = foto_obj.get('file_data')
            description = foto_obj.get('description')
            cursor.execute(
                """INSERT INTO fotos_registro 
                   (id_registro, imagen_base64, description) 
                   VALUES (%s, %s, %s)""",
                (id_registro, file_data, description)
            )

        # CAMBIO 3: Hacemos lo mismo para los videos.
        for video_obj in videos or []:
            file_data = video_obj.get('file_data')
            description = video_obj.get('description')
            cursor.execute(
                """INSERT INTO videos_registro 
                   (id_registro, video_base64, description) 
                   VALUES (%s, %s, %s)""",
                (id_registro, file_data, description)
            )

        conn.commit()
        print(f"Registro {id_registro} guardado exitosamente en PostgreSQL.")

    except psycopg2.Error as e: # MEJORA: Capturamos el error específico de psycopg2 para más detalles
        print(f"Error de base de datos al guardar en PostgreSQL: {e}")
        # Opcional: podrías querer que la función devuelva un error
        # raise e 
    except Exception as e:
        print(f"Error general al guardar en PostgreSQL: {str(e)}")
        # raise e
    finally:
        if conn:
            cursor.close()
            connection_pool.putconn(conn)

def create_project(user_id, nombre, fecha_inicio, fecha_fin, director, ubicacion, coordenadas, cliente, numero_proyecto):
    try:
        conn, cursor = get_db_connection()
        
        cursor.execute(
            """INSERT INTO proyectos (nombre_proyecto, fecha_inicio, fecha_fin, director_obra, ubicacion, coordenadas, user_id, cliente, numero_proyecto)
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING id_proyecto""",
            (nombre, fecha_inicio, fecha_fin, director, ubicacion, coordenadas, user_id, cliente, numero_proyecto)
        )
        
        project_id = cursor.fetchone()[0]
        conn.commit()
        return project_id
    except psycopg2.Error as e:
        print(f"Error al crear proyecto: {e}")
        return None
    finally:
        if conn:
            cursor.close()
            connection_pool.putconn(conn)

def get_db_connection():
    conn = None
    cursor = None

    try:
        conn = connection_pool.getconn()

        print(
            f"[POOL] getconn | "
            f"PID={os.getpid()} | "
            f"pool={id(connection_pool)} | "
            f"usadas={len(connection_pool._used)} | "
            f"disponibles={len(connection_pool._pool)}"
        )

        # Verificar si la conexión está marcada como cerrada
        if conn.closed:
            connection_pool.putconn(conn, close=True)
            conn = connection_pool.getconn()

        cursor = conn.cursor()

        # Verificar que la conexión PostgreSQL/SSL
        # realmente siga funcionando
        cursor.execute("SELECT 1")
        cursor.fetchone()

        empresa_id = session.get('empresa_id', 1)

        cursor.execute(
            "SET app.empresa_id = %s",
            (empresa_id,)
        )

        print(f"[POOL] Conexiones en uso: {len(connection_pool._used)}")

        return conn, cursor

    except Exception as e:
        print(f"[POOL] Error obteniendo conexión: {e}")

        # Cerrar cursor si llegó a crearse
        if cursor:
            try:
                cursor.close()
            except Exception:
                pass

        # La conexión fallida NO debe volver al pool
        if conn:
            try:
                connection_pool.putconn(conn, close=True)
            except Exception:
                pass

        raise

'''
def get_db_connection():
    conn = psycopg2.connect(POSTGRES_CONFIG)
    cursor = conn.cursor()
    empresa_id = session.get('empresa_id', 1)
    cursor.execute("SET app.empresa_id = %s", (empresa_id,))
    return conn, cursor
'''

def get_user_projects(user_id):
    try:
        with db_connection() as (conn, cursor):
            cursor.execute("""
                SELECT 
                    p.id, 
                    p.nombre_proyecto, 
                    p.fecha_inicio, 
                    p.cliente, 
                    p.user_id,
                    p.estado,
                    COUNT(c.id) as total_registros,
                    MAX(rf.created_at) as ultima_actividad
                FROM proyectos p
                INNER JOIN proyecto_usuarios pu ON pu.id_proyecto = p.id
                LEFT JOIN contactos c ON c.id_proyecto = p.id
                LEFT JOIN respuestas_formulario rf ON rf.id_proyecto = p.id
                WHERE pu.user_id = %s 
                GROUP BY p.id, p.nombre_proyecto, 
                        p.fecha_inicio, p.cliente, p.user_id, p.estado
                ORDER BY COALESCE(MAX(rf.created_at), p.fecha_inicio) DESC
            """, (user_id,))
            
            projects = []
            project_ids = []
            for row in cursor.fetchall():
                projects.append({
                    'id_proyecto':      row[0],
                    'name':             row[1],
                    'fecha_inicio':     row[2].strftime('%Y-%m-%d') if row[2] else '',
                    'cliente':          row[3],
                    'user_id':          row[4],
                    'estado':           row[5] or 'En Curso',
                    'total_registros':  row[6],
                    'ultima_actividad': row[7].strftime('%Y-%m-%d %H:%M') if row[7] else None,
                    'formularios':      []
                })
                project_ids.append(row[0])

            # Cargar formularios activados con estadísticas
            if project_ids:
                cursor.execute("""
                    SELECT 
                        pfa.proyecto_id,
                        f.id,
                        f.nombre,
                        COUNT(rf.id) FILTER (WHERE rf.created_at::date = CURRENT_DATE) AS registros_hoy,
                        MAX(rf.created_at) AS ultimo_uso
                    FROM proyecto_formularios_activos pfa
                    INNER JOIN formularios f ON f.id = pfa.formulario_id
                    LEFT JOIN respuestas_formulario rf 
                        ON rf.formulario_id = pfa.formulario_id 
                        AND rf.id_proyecto = pfa.proyecto_id
                    WHERE pfa.proyecto_id = ANY(%s)
                      AND pfa.user_id = %s
                    GROUP BY pfa.proyecto_id, f.id, f.nombre
                    ORDER BY MAX(rf.created_at) DESC NULLS LAST, f.nombre ASC
                """, (project_ids, user_id))

                # Trackear el último formulario usado por cada proyecto
                ultimo_por_proyecto = {}

                for row in cursor.fetchall():
                    proyecto_id   = row[0]
                    formulario_id = row[1]
                    nombre        = row[2]
                    registros_hoy = row[3]
                    ultimo_uso    = row[4]

                    # Determinar subtexto
                    if ultimo_uso and proyecto_id not in ultimo_por_proyecto:
                        subtexto = '⭐ Último usado'
                        ultimo_por_proyecto[proyecto_id] = formulario_id
                        es_ultimo = True
                    elif registros_hoy > 0:
                        subtexto = f'{registros_hoy} {"registro" if registros_hoy == 1 else "registros"} hoy'
                        es_ultimo = False
                    elif ultimo_uso:
                        subtexto = f'Último uso: {ultimo_uso.strftime("%d/%m/%Y")}'
                        es_ultimo = False
                    else:
                        subtexto = 'Sin registros aún'
                        es_ultimo = False

                    proyecto = next((p for p in projects if p['id_proyecto'] == proyecto_id), None)
                    if proyecto:
                        proyecto['formularios'].append({
                            'id':        formulario_id,
                            'nombre':    nombre,
                            'subtexto':  subtexto,
                            'es_ultimo': es_ultimo
                        })

            return projects
    except Exception as e:
        print(f"Error al obtener proyectos: {e}")
        return []

# Función para subir archivos a Azure Blob Storage
def upload_to_blob(file_name, data, content_type):
    try:
        blob_client = blob_service_client.get_blob_client(container=container_name, blob=file_name)
        blob_client.upload_blob(data, blob_type="BlockBlob", content_settings={"content_type": content_type})
        print(f"Archivo {file_name} subido con éxito.")
    except Exception as e:
        print(f"Error al subir {file_name}: {e}")
        raise


def get_speech_config():
    speech_key = '999fcb4d3f34436ab454ec47920febe0'
    service_region = 'centralus'
    speech_config = speechsdk.SpeechConfig(subscription=speech_key, region=service_region)
    speech_config.speech_recognition_language = "es-CO"
    speech_config.speech_synthesis_language = "es-CO"
    speech_config.speech_synthesis_voice_name = "es-CO-GonzaloNeural"
    speech_config.set_property(speechsdk.PropertyId.SpeechServiceConnection_EndSilenceTimeoutMs, "8000")

    # Esto le pide a Azure que formatee el texto, convirtiendo palabras como "cinco" a "5".
    speech_config.set_property(speechsdk.PropertyId.SpeechServiceResponse_PostProcessingOption, "TrueText")

    return speech_config

def synthesize_speech(text):
    speech_config = get_speech_config()
    audio_config = speechsdk.audio.AudioOutputConfig(use_default_speaker=True)
    synthesizer = speechsdk.SpeechSynthesizer(speech_config=speech_config, audio_config=audio_config)
    result = synthesizer.speak_text_async(text).get()
    return result.reason == speechsdk.ResultReason.SynthesizingAudioCompleted

#Obtener los proyectos desde Azure Blob Storage
def get_projects_from_blob():
    projects = []
    try:
        # Obtener el cliente del contenedor
        container_client = blob_service_client.get_container_client(container_name)
        
        # Listar los blobs en el directorio de proyectos
        blobs = list(container_client.list_blobs(name_starts_with="Proyectos/"))
        
        for blob in blobs:
            if blob.name.endswith('.txt'):
                # Obtener el cliente del blob
                blob_client = blob_service_client.get_blob_client(container=container_name, blob=blob.name)
                
                # Descargar el contenido del blob
                content = blob_client.download_blob().readall().decode('utf-8')
                
                # Extraer información del proyecto
                project_info = {}
                for line in content.strip().split('\n'):
                    line = line.strip()
                    if line:
                        parts = line.split(':', 1)
                        if len(parts) == 2:
                            key = parts[0].strip()
                            value = parts[1].strip()
                            project_info[key] = value
                
                # Extraer el nombre del proyecto del nombre del archivo
                file_name = blob.name.split('/')[-1]
                project_name = file_name.replace('proyecto_', '').replace('.txt', '')
                
                # Crear un objeto de proyecto
                project = {
                    'name': project_info.get('Nombre del Proyecto', project_name),
                    'date': project_info.get('Fecha de Inicio', 'Fecha no disponible'),
                    'blob_name': blob.name,
                    # Añadir más campos según sea necesario
                }
                
                projects.append(project)
                
    except Exception as e:
        print(f"Error al obtener proyectos del Blob Storage: {e}")
    
    return projects

@app.after_request
def add_header(response):
    response.headers["ngrok-skip-browser-warning"] = "true"
    return response

@app.route('/')
def principalscreen():
    return render_template('PrincipalScreen.html')

@app.route('/paginaprincipal')
def paginaprincipal():
    project_id = request.args.get('project_id')
    project_name = request.args.get('project')

    if not project_id:
        return redirect(url_for('history')) # <--- AQUÍ ES DONDE TE ESTÁ MANDANDO

    try:
        conn, cursor = get_db_connection()
        
        # ASEGÚRATE DE QUE ESTE QUERY USE LA TABLA NUEVA
        cursor.execute('SELECT * FROM proyectos WHERE id_proyecto = %s', (project_id,))
        proyecto = cursor.fetchone()
        
        if not proyecto:
            # Si el ID existe en la URL pero no en la tabla, te manda a history
            return redirect(url_for('history')) 

        return render_template('paginaprincipal.html', project_id=project_id, project_name=project_name)
    except Exception as e:
        print(f"Error: {e}")
        return redirect(url_for('history'))

'''
@app.route('/registro', methods=['GET', 'POST'])
def registro():
    if request.method == 'POST':
        nombre = request.form.get('nombre')
        apellido = request.form.get('apellido')
        email = request.form.get('email')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        empresa = request.form.get('empresa')
        cargo = request.form.get('cargo')
        rol = request.form.get('rol')
        
        if password != confirm_password:
            flash('Las contraseñas no coinciden', 'error')
            return redirect(url_for('registro'))
        
        user_id = create_user(nombre, apellido, email, password, cargo, rol, empresa)
        if user_id:
            flash('Registro exitoso. Por favor inicie sesión.', 'success')
            return redirect(url_for('principalscreen'))
        else:
            flash('Error al registrar el usuario', 'error')
    
    return render_template('registro.html')
'''

@app.route('/login', methods=['POST'])
def login():
    t0 = time.time()
    email    = request.form.get('email')
    password = request.form.get('password')

    if not email or not password:
        return jsonify({'error': 'Por favor ingrese ambos campos'}), 400

    t1 = time.time()
    usuario = verify_user(email, password)
    print(f"verify_user tardó: {time.time() - t1:.2f}s")

    if usuario:
        session['user_id']    = usuario['user_id']
        session['user_rol']   = usuario['rol']
        session['empresa_id'] = usuario['empresa_id']
        session['user_name']  = usuario['name']

        print(f"Login total tardó: {time.time() - t0:.2f}s")

        # Si es pendiente, redirigir a cambiar contraseña
        if usuario.get('estado') == 'pendiente':
            return redirect(url_for('cambiar_password_page'))

        return redirect(url_for('registros'))
    else:
        return jsonify({'error': 'Credenciales incorrectas'}), 401


@app.route('/check-session')
def check_session():
    if 'user_id' not in session:
        return jsonify({'redirect': '/'})
    
    try:
        with db_connection() as (conn, cursor):
            cursor.execute(
                "SELECT estado FROM usuario WHERE user_id = %s",
                (session['user_id'],)
            )
            row = cursor.fetchone()
            estado = row[0] if row else 'activo'

        if estado == 'pendiente':
            return jsonify({'redirect': '/cambiar-password'})
        return jsonify({'redirect': '/registros'})

    except Exception as e:
        return jsonify({'redirect': '/registros'})
        
"""
@app.route('/index')
def index():
    return render_template('index.html')
"""

@app.route('/index')
def index():
    if 'user_id' not in session:
        return redirect(url_for('principalscreen'))
    
    project_id = request.args.get('project_id')
    project_info = None
    
    if project_id:
        try:
            conn, cursor = get_db_connection()
            # Consultamos las columnas exactas de tu tabla según las imágenes
            cursor.execute("""
                SELECT nombre_proyecto, cliente, contratista, orden_de_trabajo, ubicacion 
                FROM proyectos 
                WHERE id = %s
            """, (project_id,))
            row = cursor.fetchone()
            if row:
                project_info = {
                    'nombre': row[0],
                    'cliente': row[1],
                    'contratista': row[2],
                    'orden_de_trabajo': row[3],
                    'ubicacion': row[4]
                }
            conn.close()
        except Exception as e:
            print(f"Error al obtener info del proyecto: {e}")

    return render_template('index.html', project=project_info)

def obtener_token():
    """Obtiene un token de autenticación de Bentley."""
    try:
        payload = {
            'grant_type': 'client_credentials',
            'client_id': SYNCHRO_CONFIG['client_id'],
            'client_secret': SYNCHRO_CONFIG['client_secret'],
            'scope': 'itwin-platform'
        }
        response = requests.post(SYNCHRO_CONFIG['token_url'], data=payload)
        
        if response.status_code == 200:
            return response.json()['access_token']
        else:
            print(f"Error al obtener token (código {response.status_code}): {response.text}")
            return None
    except Exception as e:
        print(f"Excepción al obtener token: {str(e)}")
        return None
    
def obtener_id_por_numero(token, numero):
    """Busca un formulario por su número y retorna su ID y el objeto 'form' completo."""
    headers = {
        'Authorization': f'Bearer {token}',
        'Accept': 'application/vnd.bentley.itwin-platform.v2+json',
        'Prefer': 'return=representation'
    }
    
    url = SYNCHRO_CONFIG['forms_url']
    params = {
        'iTwinId': SYNCHRO_CONFIG['itwin_id'],
        '$top': 50  # Obtener de 50 en 50
    }
    
    while True:
        try:
            response = requests.get(url, headers=headers, params=params)
            
            if response.status_code != 200:
                print(f"Error al buscar formulario (código {response.status_code}): {response.text}")
                return None, None

            data = response.json()
            forms_data = data.get('forms', data) # A veces la respuesta no viene anidada
            
            forms_list = forms_data.get('formDataInstances', [])
            
            for form in forms_list:
                if form.get('number') == numero:
                    # ¡Encontrado! Retorna el ID y el objeto
                    return form.get('id'), form
            
            # Lógica de paginación
            next_link_data = forms_data.get('_links', {}).get('next')
            if not next_link_data:
                break # No hay más páginas
            
            # Extraer el 'continuationToken' para la siguiente página
            next_href = next_link_data.get('href', '')
            if 'continuationToken=' in next_href:
                params['continuationToken'] = next_href.split('continuationToken=')[-1]
                params.pop('$top', None) # Ya no es necesario
            else:
                break # No se pudo encontrar el token de paginación
                
        except Exception as e:
            print(f"Excepción al buscar formulario: {str(e)}")
            return None, None

    # Si sale del bucle sin encontrarlo
    print(f"No se encontró ningún formulario con el número: {numero}")
    return None, None

@app.route('/formulario-synchro')
def formulario_synchro():
    if 'user_id' not in session:
        return redirect(url_for('principalscreen'))
    # Asumiendo que has guardado el archivo como 'indexFormulario.html' en tu carpeta 'templates'
    return render_template('indexFormulario.html')

@app.route('/get-synchro-form-data')
def get_synchro_data():
    form_number = request.args.get('form_number')
    if not form_number:
        return jsonify({'error': 'Falta form_number'}), 400

    token = obtener_token() # (Tu función de Synchro)
    if not token:
        return jsonify({'error': 'No se pudo obtener el token'}), 500

    form_id, form_data = obtener_id_por_numero(token, form_number) # (Tu función de Synchro)

    if not form_id:
        return jsonify({'error': 'Formulario no encontrado'}), 404

    # Devuelve el 'number' y las 'properties'
    return jsonify({
        'id': form_id,
        'number': form_data.get('number'),
        'properties': form_data.get('properties', {})
    })

@app.route('/update-synchro-form', methods=['POST'])
def update_synchro_data():
    try:
        data = request.json
        form_number = data.get('form_number')
        new_properties = data.get('properties')
        # Aquí también puedes manejar data.get('media')

        if not form_number or not new_properties:
            return jsonify({'error': 'Faltan datos (form_number, properties)'}), 400

        token = obtener_token()
        if not token:
            return jsonify({'error': 'No se pudo obtener el token'}), 500

        form_id, form = obtener_id_por_numero(token, form_number)
        if not form_id:
            return jsonify({'error': 'Formulario no encontrado'}), 404

        props_actuales = form.get('properties', {})
        
        for section_name, new_items in new_properties.items():
            if not new_items: 
                continue

            lista_actual = props_actuales.get(section_name, [])
            
            # Generar UUIDs para los nuevos items ANTES de agregarlos
            for item in new_items:
                item['id'] = str(uuid.uuid4()) # Aseguramos un ID único
            
            lista_actual.extend(new_items)
            props_actuales[section_name] = lista_actual

        # --- PREPARAR Y ENVIAR EL PATCH ---
        headers = {
            'Authorization': f'Bearer {token}',
            'Accept': 'application/vnd.bentley.itwin-platform.v2+json',
            'Content-Type': 'application/json'
        }
        
        cambios = {
            'properties': props_actuales
        }
        
        # --- CAMBIO 1: Usar SYNCHRO_CONFIG en lugar de BASE_URL ---
        url = f"{SYNCHRO_CONFIG['forms_url']}/{form_id}"
        
        # --- CAMBIO 2: Usar requests.patch (con 's') ---
        response = requests.patch(url, headers=headers, json=cambios)
        
        if response.status_code == 200:
            return jsonify(response.json())
        else:
            print(f"Error al actualizar Synchro ({response.status_code}): {response.text}")
            return jsonify({'error': 'Error al actualizar Synchro', 'details': response.text}), response.status_code

    except Exception as e:
        print(f"Excepción en update_synchro_data: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Error interno del servidor: {str(e)}'}), 500


# ── Vista de proyecto: GET ──────────────────────────────────────
@app.route('/proyecto/<int:project_id>')
def vista_proyecto(project_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))

    empresa_id = session.get('empresa_id')
    user_id    = session['user_id']

    proyecto         = None
    formularios_proy = []
    color_primario   = '#FFAF33'
    logo_actual      = None

    try:
        with db_connection() as (conn, cursor):
            # Cargar el proyecto
            cursor.execute("""
                SELECT id, nombre_proyecto, estado
                FROM proyectos
                WHERE id = %s AND empresa_id = %s
            """, (project_id, empresa_id))
            row = cursor.fetchone()
            if not row:
                return redirect(url_for('registros'))

            proyecto = {
                'id':     row[0],
                'nombre': row[1],
                'estado': row[2] or 'En Curso'
            }

            # Formularios disponibles del proyecto + estado de activación para este usuario
            cursor.execute("""
                SELECT 
                    f.id,
                    f.nombre,
                    f.descripcion,
                    CASE WHEN pfa.id IS NOT NULL THEN TRUE ELSE FALSE END AS activo
                FROM proyecto_formularios pf
                INNER JOIN formularios f ON f.id = pf.formulario_id
                LEFT JOIN proyecto_formularios_activos pfa 
                    ON pfa.formulario_id = pf.formulario_id 
                    AND pfa.proyecto_id = pf.proyecto_id
                    AND pfa.user_id = %s
                WHERE pf.proyecto_id = %s
                ORDER BY f.nombre ASC
            """, (user_id, project_id))
            formularios_proy = [
                {'id': r[0], 'nombre': r[1], 'descripcion': r[2] or '', 'activo': r[3]}
                for r in cursor.fetchall()
            ]

            # Colores y logo
            cursor.execute("""
                SELECT logo_url, color_primario
                FROM empresas WHERE id = %s
            """, (empresa_id,))
            emp = cursor.fetchone()
            if emp:
                logo_actual    = emp[0]
                color_primario = emp[1] or '#FFAF33'

    except Exception as e:
        print(f"Error en vista_proyecto: {e}")
        return redirect(url_for('registros'))

    return render_template('vistaProyecto.html',
                           proyecto=proyecto,
                           formularios_proy=formularios_proy,
                           logo_actual=logo_actual,
                           color_primario=color_primario,
                           empresa_id=empresa_id)


# ── API: registros del proyecto ─────────────────────────────────
@app.route('/api/proyecto/<int:project_id>/registros')
def api_registros_proyecto(project_id):
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    try:
        formulario_id = request.args.get('formulario_id', type=int)

        with db_connection() as (conn, cursor):
            base_query = """
                SELECT 
                    rf.id,
                    rf.formulario_id,
                    f.nombre AS formulario_nombre,
                    rf.respuestas,
                    rf.created_at,
                    u.name,
                    u.apellido
                FROM respuestas_formulario rf
                INNER JOIN formularios f ON f.id = rf.formulario_id
                LEFT JOIN usuario u ON u.user_id = rf.user_id
                WHERE rf.id_proyecto = %s
            """
            params = [project_id]
            if formulario_id:
                base_query += " AND rf.formulario_id = %s"
                params.append(formulario_id)

            base_query += " ORDER BY rf.created_at DESC LIMIT 100"

            cursor.execute(base_query, params)
            registros = []
            for r in cursor.fetchall():
                # Generar preview del registro
                respuestas = r[3] or {}
                preview    = ''
                if isinstance(respuestas, dict):
                    valores = [
                        str(v) for k, v in respuestas.items()
                        if k != '__repeticiones' and v and isinstance(v, (str, int, float))
                    ]
                    for lista in (respuestas.get('__repeticiones') or {}).values():
                        for rep in (lista or []):
                            if isinstance(rep, dict):
                                valores += [
                                    str(v) for v in rep.values()
                                    if v and isinstance(v, (str, int, float))
                                ]
                    preview = ' · '.join(valores[:3])[:150]

                registros.append({
                    'id':                r[0],
                    'formulario_id':     r[1],
                    'formulario_nombre': r[2],
                    'preview':           preview,
                    'created_at':        (
                        r[4].replace(tzinfo=timezone.utc)
                            .astimezone(pytz.timezone('America/Bogota'))
                            .isoformat()
                    ) if r[4] else None,
                    'autor':             f"{r[5] or ''} {r[6] or ''}".strip() or 'Usuario'
                })

            return jsonify({'registros': registros})

    except Exception as e:
        print(f"Error en api_registros_proyecto: {e}")
        return jsonify({'error': str(e)}), 500


# ── API: activar/desactivar formulario ──────────────────────────
@app.route('/api/proyecto/<int:project_id>/formulario/<int:formulario_id>/toggle', methods=['POST'])
def toggle_formulario_activo(project_id, formulario_id):
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    try:
        data       = request.get_json() or {}
        activar    = data.get('activar', True)
        empresa_id = session.get('empresa_id')
        user_id    = session['user_id']

        with db_connection() as (conn, cursor):
            if activar:
                cursor.execute("""
                    INSERT INTO proyecto_formularios_activos 
                        (proyecto_id, formulario_id, user_id, empresa_id)
                    VALUES (%s, %s, %s, %s)
                    ON CONFLICT (proyecto_id, formulario_id, user_id) DO NOTHING
                """, (project_id, formulario_id, user_id, empresa_id))
            else:
                cursor.execute("""
                    DELETE FROM proyecto_formularios_activos
                    WHERE proyecto_id = %s 
                      AND formulario_id = %s 
                      AND user_id = %s
                """, (project_id, formulario_id, user_id))

            return jsonify({'success': True, 'activo': activar})

    except Exception as e:
        print(f"Error en toggle_formulario_activo: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/campos-globales/<int:campo_id>', methods=['PUT'])
def actualizar_campo_global(campo_id):
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        data          = request.get_json()
        nombre        = data.get('nombre', '').strip()
        tipo          = data.get('tipo')
        objeto        = data.get('objeto', 'formulario')
        opciones      = data.get('opciones', [])
        configuracion = data.get('configuracion', {})

        if not nombre or not tipo:
            return jsonify({'error': 'Nombre y tipo son obligatorios'}), 400

        with db_connection() as (conn, cursor):
            cursor.execute("""
                UPDATE campos_globales
                SET nombre = %s, tipo = %s, objeto = %s, 
                    opciones = %s, configuracion = %s
                WHERE id = %s AND empresa_id = %s AND es_sistema = FALSE
            """, (
                nombre, tipo, objeto,
                json.dumps(opciones), json.dumps(configuracion),
                campo_id, session.get('empresa_id')
            ))
            if cursor.rowcount == 0:
                return jsonify({'error': 'Campo no encontrado o no editable'}), 404
            return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/get-synchro-project-data')
def get_synchro_project_data():
    return jsonify({
        'codigo_proyecto':    'CO-CARR',
        'contrato':           '4500042183',
        'contratista':        'J.E. JAIMES INGENIEROS S.A.',
        'form_definition_id': SYNCHRO_FORM_DEFINITION_ID
    })

@app.route('/get-form-definitions')
def get_form_definitions():
    token = obtener_token()
    if not token:
        return jsonify({'error': 'No se pudo obtener token'}), 500
    
    headers = {
        'Authorization': f'Bearer {token}',
        'Accept': 'application/vnd.bentley.itwin-platform.v2+json'
    }
    
    response = requests.get(
        'https://api.bentley.com/forms',
        headers=headers,
        params={
            'iTwinId': SYNCHRO_CONFIG['itwin_id'],
            '$top': 50
        }
    )
    
    return jsonify({
        'status': response.status_code,
        'body': response.json()
    })


@app.route('/cambiar-password', methods=['GET'])
def cambiar_password_page():
    if 'user_id' not in session:
        return redirect(url_for('principalscreen'))
    return render_template('cambiarPassword.html')

@app.route('/cambiar-password', methods=['POST'])
def cambiar_password():
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    data             = request.get_json()
    password_nuevo   = data.get('password_nuevo')
    password_confirm = data.get('password_confirm')

    if password_nuevo != password_confirm:
        return jsonify({'error': 'Las contraseñas no coinciden'}), 400

    if len(password_nuevo) < 8:
        return jsonify({'error': 'La contraseña debe tener al menos 8 caracteres'}), 400

    try:
        with db_connection() as (conn, cursor):
            cursor.execute("""
                UPDATE usuario 
                SET password = %s, estado = 'activo'
                WHERE user_id = %s
            """, (generate_password_hash(password_nuevo), session['user_id']))

        return jsonify({'success': True})

    except Exception as e:
        print(f"Error cambiando contraseña: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/get-form-detail')
def get_form_detail():
    token = obtener_token()
    headers = {
        'Authorization': f'Bearer {token}',
        'Accept': 'application/vnd.bentley.itwin-platform.v2+json'
    }
    response = requests.get(
        f"https://api.bentley.com/forms/e4bQKVghekuuA8Y6dmHKWFDXLUqEPIpFt_QjKefA5yk",
        headers=headers
    )
    forms = response.json().get('forms', {}).get('formDataInstances', [])
    
    # Buscar uno que sea del tipo 2.02
    for form in forms:
        if '2.02' in form.get('number', '') or 'Calidad' in form.get('type', ''):
            # Obtener detalle completo de ese formulario
            detail = requests.get(
                f"https://api.bentley.com/forms/{form['id']}",
                headers=headers
            )
            return jsonify({'status': detail.status_code, 'body': detail.json()})
    
    return jsonify(response.json())

@app.route('/create-synchro-form', methods=['POST'])
def create_synchro_form():
    try:
        data               = request.json
        form_definition_id = data.get('form_definition_id') or SYNCHRO_FORM_DEFINITION_ID
        new_properties     = data.get('properties', {})
        token = obtener_token()   # reutiliza la función que ya existe
        if not token:
            return jsonify({'error': 'No se pudo obtener token'}), 500
        headers = {
            'Authorization': f'Bearer {token}',
            'Accept':        'application/vnd.bentley.itwin-platform.v2+json',
            'Content-Type':  'application/json',
            'Prefer':        'return=representation'
        }
        body = {
            'formId': form_definition_id,
            'properties': new_properties
        }
        response = requests.post(SYNCHRO_CONFIG['forms_url'], headers=headers, json=body, timeout=15)
        if response.status_code in (200, 201):
            created = response.json().get('form', response.json())
            return jsonify({'success': True, 'form_id': created.get('id'), 'number': created.get('number')}), 201
        return jsonify({'error': response.text}), response.status_code
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# Obtener campos globales de la empresa
@app.route('/api/campos-globales', methods=['GET'])
def obtener_campos_globales():
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        objeto = request.args.get('objeto')
        with db_connection() as (conn, cursor):
            if objeto:
                cursor.execute("""
                    SELECT id, nombre, tipo, objeto, opciones, configuracion, es_sistema
                    FROM campos_globales
                    WHERE empresa_id = %s AND objeto = %s
                    ORDER BY es_sistema DESC, created_at DESC
                """, (session.get('empresa_id'), objeto))
            else:
                cursor.execute("""
                    SELECT id, nombre, tipo, objeto, opciones, configuracion, es_sistema
                    FROM campos_globales
                    WHERE empresa_id = %s
                    ORDER BY es_sistema DESC, created_at DESC
                """, (session.get('empresa_id'),))
            rows = cursor.fetchall()
            campos = [
                {
                    'id':            r[0],
                    'nombre':        r[1],
                    'tipo':          r[2],
                    'objeto':        r[3],
                    'opciones':      r[4] or [],
                    'configuracion': r[5] or {},
                    'es_sistema':    r[6] or False
                }
                for r in rows
            ]
            return jsonify({'campos': campos})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# Crear campo global
@app.route('/api/campos-globales', methods=['POST'])
def crear_campo_global():
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        data          = request.get_json()
        nombre        = data.get('nombre', '').strip()
        tipo          = data.get('tipo')
        objeto        = data.get('objeto', 'formulario')  # ← nuevo
        opciones      = data.get('opciones', [])
        configuracion = data.get('configuracion', {})

        if not nombre or not tipo:
            return jsonify({'error': 'Nombre y tipo son obligatorios'}), 400

        if objeto not in ('formulario', 'proyecto'):
            return jsonify({'error': 'Objeto inválido'}), 400

        with db_connection() as (conn, cursor):
            cursor.execute("""
                INSERT INTO campos_globales
                    (empresa_id, nombre, tipo, objeto, opciones, configuracion)
                VALUES (%s, %s, %s, %s, %s, %s)
                RETURNING id
            """, (
                session.get('empresa_id'), nombre, tipo, objeto,
                json.dumps(opciones), json.dumps(configuracion)
            ))
            nuevo_id = cursor.fetchone()[0]
            return jsonify({'success': True, 'id': nuevo_id})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# Eliminar campo global
@app.route('/api/campos-globales/<int:campo_id>', methods=['DELETE'])
def eliminar_campo_global(campo_id):
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        with db_connection() as (conn, cursor):
            cursor.execute("""
                DELETE FROM campos_globales
                WHERE id = %s AND empresa_id = %s
            """, (campo_id, session.get('empresa_id')))
            return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# Obtener formularios de la empresa
@app.route('/api/formularios', methods=['GET'])
def get_formularios():
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        with db_connection() as (conn, cursor):
            cursor.execute("""
                SELECT id, nombre, descripcion, campos, created_at
                FROM formularios
                WHERE empresa_id = %s
                ORDER BY created_at DESC
            """, (session.get('empresa_id'),))
            formularios = []
            for row in cursor.fetchall():
                formularios.append({
                    'id':          row[0],
                    'nombre':      row[1],
                    'descripcion': row[2] or '',
                    'campos':      row[3] or [],
                    'created_at':  row[4].strftime('%d/%m/%Y') if row[4] else ''
                })
            return jsonify({'formularios': formularios})
    except Exception as e:
        print(f"[ERROR] get_formularios: {e}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ── Formulario Dinámico: GET ────────────────────────────────────
@app.route('/formulario-dinamico')
def formulario_dinamico():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    project_id    = request.args.get('project_id')
    formulario_id = request.args.get('formulario_id')
    registro_id   = request.args.get('registro_id')  # ← nuevo

    if not project_id or not formulario_id:
        return redirect(url_for('registros'))

    formulario     = None
    campos         = []
    secciones      = []
    logo_actual    = None
    color_primario = '#FFAF33'
    registro       = None
    puede_editar   = True

    try:
        with db_connection() as (conn, cursor):
            # Obtener formulario
            cursor.execute("""
                SELECT id, nombre, descripcion, campos
                FROM formularios
                WHERE id = %s AND empresa_id = %s
            """, (formulario_id, session.get('empresa_id')))
            row = cursor.fetchone()

            if not row:
                return redirect(url_for('registros'))

            formulario = {
                'id':          row[0],
                'nombre':      row[1],
                'descripcion': row[2] or '',
                'campos_config': row[3] or []
            }

            # Si es edición, cargar el registro
            if registro_id:
                cursor.execute("""
                    SELECT rf.id, rf.respuestas, rf.user_id, u.rol
                    FROM respuestas_formulario rf
                    LEFT JOIN usuario u ON u.user_id = %s
                    WHERE rf.id = %s
                """, (session['user_id'], registro_id))
                reg = cursor.fetchone()
                if reg:
                    registro = {
                        'id':         reg[0],
                        'respuestas': reg[1] or {},
                        'autor_id':   reg[2]
                    }
                    # Verificar permisos: autor o admin
                    rol_usuario = reg[3] or ''
                    puede_editar = (
                        reg[2] == session['user_id'] or
                        rol_usuario.lower() in ('admin', 'administrador')
                    )

            # Obtener campos globales
            def es_grupo(item):
                return isinstance(item, dict) and item.get('tipo') == 'grupo'

            campo_ids = [
                (item['id'] if isinstance(item, dict) else item)
                for item in formulario['campos_config']
                if not es_grupo(item)
            ]

            campos_db = {}
            if campo_ids:
                cursor.execute("""
                    SELECT id, nombre, tipo, opciones, configuracion
                    FROM campos_globales
                    WHERE id = ANY(%s) AND empresa_id = %s
                """, (campo_ids, session.get('empresa_id')))

                campos_db = {r[0]: {
                    'id': r[0], 'nombre': r[1], 'tipo': r[2],
                    'opciones': r[3] or [], 'configuracion': r[4] or {}
                } for r in cursor.fetchall()}

            for item in formulario['campos_config']:
                if es_grupo(item):
                    campos.append({
                        'tipo':      'grupo',
                        'gid':       item.get('gid') or '',
                        'nombre':    item.get('nombre', 'Grupo'),
                        'repetible': item.get('repetible', False)
                    })
                    continue
                cid       = item['id'] if isinstance(item, dict) else item
                requerido = item.get('requerido', False) if isinstance(item, dict) else False
                if cid in campos_db:
                    campo = campos_db[cid].copy()
                    campo['requerido'] = requerido
                    # Pre-cargar valor si es edición
                    if registro:
                        campo['valor'] = registro['respuestas'].get(str(cid)) or registro['respuestas'].get(cid) or ''
                    else:
                        campo['valor'] = ''
                    campos.append(campo)

            # ── Agrupar en secciones (para el render con grupos) ──
            reps_guardadas = {}
            if registro:
                reps_guardadas = (registro['respuestas'] or {}).get('__repeticiones') or {}

            actual = None
            for c in campos:
                if c.get('tipo') == 'grupo':
                    actual = {
                        'tipo':      'grupo',
                        'gid':       c['gid'],
                        'nombre':    c['nombre'],
                        'repetible': c['repetible'],
                        'campos':    []
                    }
                    secciones.append(actual)
                else:
                    if actual is None:
                        actual = {'tipo': 'sueltos', 'gid': '', 'nombre': '',
                                  'repetible': False, 'campos': []}
                        secciones.append(actual)
                    actual['campos'].append(c)

            for sec in secciones:
                if sec['tipo'] != 'grupo':
                    sec['repeticiones'] = []
                    continue
                guardadas = reps_guardadas.get(sec['gid']) or []
                sec['repeticiones'] = guardadas if guardadas else [{}]
                for c in sec['campos']:
                    c['valor'] = ''

            # Colores y logo
            cursor.execute("""
                SELECT logo_url, color_primario
                FROM empresas WHERE id = %s
            """, (session.get('empresa_id'),))
            emp = cursor.fetchone()
            if emp:
                logo_actual    = emp[0]
                color_primario = emp[1] or '#FFAF33'

    except Exception as e:
        print(f"Error en formulario-dinamico GET: {e}")
        return redirect(url_for('registros'))

    return render_template('formDinamico.html',
                           project_id=project_id,
                           formulario=formulario,
                           campos=campos,
                           secciones=secciones,
                           registro_id=registro_id,
                           puede_editar=puede_editar,
                           logo_actual=logo_actual,
                           color_primario=color_primario)


# ── Formulario Dinámico: POST ───────────────────────────────────
@app.route('/api/respuestas-formulario', methods=['POST'])
def guardar_respuesta_formulario(_reintento=0):
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    try:
        data          = request.get_json()
        formulario_id = data.get('formulario_id')
        project_id    = data.get('project_id')
        respuestas    = data.get('respuestas', {})

        if not formulario_id or not project_id:
            return jsonify({'error': 'Faltan datos obligatorios'}), 400

        with db_connection() as (conn, cursor):
            cursor.execute("""
                INSERT INTO respuestas_formulario
                    (formulario_id, id_proyecto, user_id, respuestas)
                VALUES (%s, %s, %s, %s)
                RETURNING id
            """, (
                formulario_id,
                project_id,
                session['user_id'],
                json.dumps(respuestas)
            ))
            nuevo_id = cursor.fetchone()[0]

            # Enlazar las transcripciones de esta sesión con el registro final.
            # Permite comparar lo que extrajo el modelo contra lo que quedó
            # guardado, es decir, medir las correcciones manuales del usuario.
            sesion_uuid = data.get('sesion_uuid')
            if sesion_uuid:
                try:
                    cursor.execute("""
                        UPDATE transcripciones_log
                        SET respuesta_id = %s
                        WHERE sesion_uuid = %s AND respuesta_id IS NULL
                    """, (nuevo_id, sesion_uuid))
                except Exception as e:
                    print(f"[LOG-TRANS] No se pudo enlazar la sesión: {e}")

            return jsonify({'success': True, 'id': nuevo_id})

    except (psycopg2.OperationalError, psycopg2.InterfaceError) as e:
        print(f"Error de conexión al guardar respuesta: {e} (intento {_reintento + 1})")
        if _reintento < 2:
            return guardar_respuesta_formulario(_reintento=_reintento + 1)
        return jsonify({'error': 'No se pudo conectar. Intenta guardar de nuevo.'}), 503

    except Exception as e:
        print(f"Error al guardar respuesta: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/respuestas-formulario/<int:registro_id>', methods=['PUT'])
def actualizar_respuesta_formulario(registro_id):
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    try:
        data       = request.get_json()
        respuestas = data.get('respuestas', {})
        user_id    = session['user_id']

        with db_connection() as (conn, cursor):
            # Verificar permisos: autor o admin
            cursor.execute("""
                SELECT rf.user_id, u.rol
                FROM respuestas_formulario rf
                LEFT JOIN usuario u ON u.user_id = %s
                WHERE rf.id = %s
            """, (user_id, registro_id))
            row = cursor.fetchone()
            if not row:
                return jsonify({'error': 'Registro no encontrado'}), 404

            autor_id  = row[0]
            rol       = (row[1] or '').lower()
            es_admin  = rol in ('admin', 'administrador')

            if autor_id != user_id and not es_admin:
                return jsonify({'error': 'No tienes permiso para editar este registro'}), 403

            # Actualizar
            cursor.execute("""
                UPDATE respuestas_formulario
                SET respuestas = %s,
                    updated_by = %s,
                    updated_at = NOW()
                WHERE id = %s
            """, (json.dumps(respuestas), user_id, registro_id))

            return jsonify({'success': True})

    except Exception as e:
        print(f"Error al actualizar respuesta: {e}")
        return jsonify({'error': str(e)}), 500

# Crear formulario
@app.route('/api/formularios', methods=['POST'])
def crear_formulario():
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        data        = request.get_json()
        nombre      = data.get('nombre', '').strip()
        descripcion = data.get('descripcion', '').strip()
        campos      = data.get('campos', [])

        if not nombre:
            return jsonify({'error': 'El nombre es obligatorio'}), 400

        with db_connection() as (conn, cursor):
            cursor.execute("""
                INSERT INTO formularios (empresa_id, nombre, descripcion, campos)
                VALUES (%s, %s, %s, %s)
                RETURNING id
            """, (
                session.get('empresa_id'), nombre, descripcion,
                json.dumps(campos)
            ))
            nuevo_id = cursor.fetchone()[0]
            return jsonify({'success': True, 'id': nuevo_id})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# Actualizar formulario
@app.route('/api/formularios/<int:form_id>', methods=['PUT'])
def actualizar_formulario(form_id):
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        data        = request.get_json()
        nombre      = data.get('nombre', '').strip()
        descripcion = data.get('descripcion', '').strip()
        campos      = data.get('campos', [])

        if not nombre:
            return jsonify({'error': 'El nombre es obligatorio'}), 400

        with db_connection() as (conn, cursor):
            cursor.execute("""
                UPDATE formularios
                SET nombre = %s, descripcion = %s, campos = %s
                WHERE id = %s AND empresa_id = %s
            """, (
                nombre, descripcion, json.dumps(campos),
                form_id, session.get('empresa_id')
            ))
            return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/informeDiario')
def informe_diario():
    if 'user_id' not in session:
        return redirect(url_for('principalscreen'))

    project_id = request.args.get('project_id')
    project    = None
    color_primario   = '#FFAF33'
    color_secundario = '#E3E3E3'
    logo_actual      = None

    try:
        with db_connection() as (conn, cursor):
            if project_id:
                cursor.execute("""
                    SELECT nombre_proyecto, cliente, contratista,
                           orden_de_trabajo, ubicacion
                    FROM proyectos WHERE id = %s
                """, (project_id,))
                row = cursor.fetchone()
                if row:
                    project = {
                        'nombre_proyecto':  row[0],
                        'cliente':          row[1],
                        'contratista':      row[2],
                        'orden_de_trabajo': row[3],
                        'ubicacion':        row[4]
                    }
            cursor.execute("""
                SELECT color_primario, color_secundario, logo_url
                FROM empresas WHERE id = %s
            """, (session.get('empresa_id'),))
            emp = cursor.fetchone()
            if emp:
                color_primario   = emp[0] or '#FFAF33'
                color_secundario = emp[1] or '#E3E3E3'
                logo_actual      = emp[2]
    except Exception as e:
        print(f"Error en informe_diario: {e}")

    return render_template('informeDiario.html',
                           project=project,
                           color_primario=color_primario,
                           color_secundario=color_secundario,
                           logo_actual=logo_actual)

# Eliminar formulario
@app.route('/api/formularios/<int:form_id>', methods=['DELETE'])
def eliminar_formulario(form_id):
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        with db_connection() as (conn, cursor):
            cursor.execute("""
                DELETE FROM formularios
                WHERE id = %s AND empresa_id = %s
            """, (form_id, session.get('empresa_id')))
            return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/formulario')
def indexFormulario():
    """Muestra el formulario con datos pre-cargados"""
    proyecto = {
        'codigo': '10111',
        'contratista': 'ABCD',
        'contrato': '001'
    }
    return render_template('indexFormulario.html', proyecto=proyecto)

@app.route('/registros')
def registros():
    print(f"empresa_id en sesión: {session.get('empresa_id')}")
    if 'user_id' not in session:
        return redirect(url_for('principalscreen'))
    
    # Obtener proyectos de PostgreSQL
    db_projects = get_user_projects(session['user_id'])
    
    # Obtener proyectos de Azure Blob (si aún los necesitas)
    #blob_projects = get_projects_from_blob()  # Tu función existente
    
    # Combinar proyectos (o usar solo los de PostgreSQL)
    conn, cursor = get_db_connection()
    cursor.execute("""
        SELECT logo_url, color_primario, color_secundario
        FROM empresas
        WHERE id = %s
    """, (session.get('empresa_id'),))
    empresa_row = cursor.fetchone()
    logo_actual      = empresa_row[0] if empresa_row else None
    color_primario   = empresa_row[1] if empresa_row else '#FFAF33'
    color_secundario = empresa_row[2] if empresa_row else '#E3E3E3'
    cursor.close()
    connection_pool.putconn(conn)

    return render_template('registros.html',
        db_projects=db_projects,
        logo_actual=logo_actual,
        color_primario=color_primario,
        color_secundario=color_secundario
    )

# Ruta para la vista "history"
@app.route('/history')
def history():
    if 'user_id' not in session:
        return redirect(url_for('principalscreen'))
    
    # Obtenemos los proyectos de PostgreSQL
    db_projects = get_user_projects(session['user_id'])
    
    # IMPORTANTE: get_user_projects ya devuelve 'name', 
    # pero asegúrate de que el HTML lo use correctamente.
    return render_template('history.html', db_projects=db_projects)

@app.route('/configuracion')
def configuracion():
    if 'user_id' not in session:
        return redirect(url_for('principalscreen'))
    if session.get('user_rol') != 'admin':
        return redirect(url_for('registros'))
    conn = None
    try:
        conn, cursor = get_db_connection()
        
        # Usuarios del mismo tenant/organización
        cursor.execute("""
            SELECT user_id, name, apellido, email, rol, estado
            FROM usuario
            WHERE empresa_id = %s
            ORDER BY name ASC
        """, (session.get('empresa_id'),))
        
        miembros = []
        for row in cursor.fetchall():
            miembros.append({
                'user_id': row[0], 'nombre': row[1], 'apellido': row[2] or '',
                'email': row[3], 'rol': row[4] or 'Sin rol',
                'estado': row[5] or 'pendiente', 'foto': None
            })
        
        cursor.execute("""
            SELECT url FROM empresa_logos
            WHERE empresa_id = %s
            ORDER BY created_at DESC
            LIMIT 1
        """, (session.get('empresa_id'),))
        logo_row = cursor.fetchone()
        logo_actual = logo_row[0] if logo_row else None

        cursor.execute("""
            SELECT logo_url, color_primario, color_secundario
            FROM empresas
            WHERE id = %s
        """, (session.get('empresa_id'),))
        empresa_row = cursor.fetchone()
        logo_actual      = empresa_row[0] if empresa_row else None
        color_primario   = empresa_row[1] if empresa_row else '#FBAF33'
        color_secundario = empresa_row[2] if empresa_row else '#E3E3E3'

        return render_template('configuracion.html', 
            miembros=miembros,
            logo_actual=logo_actual,
            color_primario=color_primario,
            color_secundario=color_secundario
        )
        #return render_template('configuracion.html', miembros=miembros)
    except Exception as e:
        print(f"Error en usuario: {e}")
        return render_template('configuracion.html', miembros=[])
    finally:
        if conn:
            cursor.close()
            connection_pool.putconn(conn)

@app.route('/inventario')
def inventario():
    return render_template('inventario.html')

# En tu archivo app.py

@app.route('/historialRegistro/<int:id_proyecto>')
def historialregistro(id_proyecto):
    if 'user_id' not in session:
        return redirect(url_for('principalscreen'))

    try:
        with db_connection() as (conn, cursor):

            # Info del proyecto
            cursor.execute('SELECT nombre_proyecto, cliente FROM proyectos WHERE id = %s', (id_proyecto,))
            proyecto_info = cursor.fetchone()
            if not proyecto_info:
                return redirect(url_for('history'))

            # Paginación
            page     = request.args.get('page', 1, type=int)
            per_page = 10
            offset   = (page - 1) * per_page

            print(f"PAGE: {page}, PER_PAGE: {per_page}, OFFSET: {offset}, TOTAL: ", end='')

            # Total de registros
            cursor.execute("""
                SELECT COUNT(*) FROM contactos WHERE id_proyecto = %s
            """, (id_proyecto,))
            total     = cursor.fetchone()[0]
            print(total)
            paginas   = (total + per_page - 1) // per_page

            # Consulta paginada
            cursor.execute("""
                SELECT c.id, c.nombre, c.empresa, c.cargo,
                       c.telefono, c.email, c.ciudad, c.notas,
                       c.created_at,
                       u.name, u.apellido, u.cargo as user_cargo
                FROM contactos c
                LEFT JOIN usuario u ON u.user_id = c.user_id
                WHERE c.id_proyecto = %s
                ORDER BY c.created_at DESC
                LIMIT %s OFFSET %s
            """, (id_proyecto, per_page, offset))

            registros_rows = cursor.fetchall()
            ids_contactos  = [r[0] for r in registros_rows]

            # Una sola consulta para todas las fotos
            fotos_por_contacto = {}
            if ids_contactos:
                cursor.execute("""
                    SELECT contacto_id, imagen_url, descripcion
                    FROM contacto_imagenes
                    WHERE contacto_id = ANY(%s)
                """, (ids_contactos,))
                for f in cursor.fetchall():
                    cid, url, desc = f
                    if cid not in fotos_por_contacto:
                        fotos_por_contacto[cid] = []
                    if url:
                        fotos_por_contacto[cid].append({'url': url, 'base64': None, 'desc': desc or ''})

            reportes_completos = []
            colombia_tz = pytz.timezone('America/Bogota')

            for r_row in registros_rows:
                id_c, nombre, empresa, cargo, telefono, email, ciudad, notas, created_at, u_name, u_apellido, u_cargo = r_row

                u_name     = u_name or ''
                u_apellido = u_apellido or ''
                iniciales  = (u_name[0] + u_apellido[0]).upper() if u_name and u_apellido else '??'

                if created_at:
                    created_at_col = created_at.replace(tzinfo=timezone.utc).astimezone(colombia_tz)
                    hora      = created_at_col.strftime('%I:%M %p')
                    fecha_str = created_at_col.strftime('%d/%m/%Y')
                else:
                    hora      = None
                    fecha_str = 'S/F'

                reportes_completos.append({
                    'id_registro':       id_c,
                    'actividad':         nombre,
                    'descripcion':       f"{empresa or ''} · {cargo or ''}".strip(' ·'),
                    'estado':            ciudad or '',
                    'avance':            None,
                    'fecha':             fecha_str,
                    'hora':              hora,
                    'usuario_nombre':    f"{u_name} {u_apellido}".strip() or 'Sin asignar',
                    'usuario_cargo':     u_cargo or '',
                    'usuario_iniciales': iniciales,
                    'fotos':             fotos_por_contacto.get(id_c, []),
                    'telefono':          telefono or '',
                    'email':             email or '',
                    'notas':             notas or ''
                })

            # Colores de empresa
            cursor.execute("""
                SELECT color_primario, color_secundario
                FROM empresas WHERE id = %s
            """, (session.get('empresa_id'),))
            empresa_row      = cursor.fetchone()
            color_primario   = empresa_row[0] if empresa_row else '#FFAF33'
            color_secundario = empresa_row[1] if empresa_row else '#E3E3E3'

            return render_template('historialRegistro.html',
                proyecto=proyecto_info,
                reportes=reportes_completos,
                id_proyecto=id_proyecto,
                color_primario=color_primario,
                color_secundario=color_secundario,
                page=page,
                paginas=paginas,
                total=total
            )
    except Exception as e:
        print(f"Error en historialregistro: {e}")
        return redirect(url_for('history'))

@app.route('/guardar_contacto', methods=['POST'])
def guardar_contacto():
    if 'user_id' not in session:
        return jsonify({"error": "No autorizado"}), 401
    try:
        data       = request.json
        empresa_id = session.get('empresa_id')

        with db_connection() as (conn, cursor):
            cursor.execute("""
                INSERT INTO contactos 
                    (empresa_id, user_id, id_proyecto, nombre, empresa, cargo, 
                    telefono, email, ciudad, notas)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            """, (
                empresa_id, session['user_id'],
                data.get('id_proyecto'),
                data.get('nombre'), data.get('empresa'), data.get('cargo'),
                data.get('telefono'), data.get('email'),
                data.get('ciudad'), data.get('notas')
            ))

            contacto_id = cursor.fetchone()[0]

            for img in data.get('imagenes', []):
                cursor.execute("""
                    INSERT INTO contacto_imagenes 
                        (contacto_id, empresa_id, imagen_url, descripcion)
                    VALUES (%s, %s, %s, %s)
                """, (contacto_id, empresa_id, img.get('url'), img.get('descripcion')))

        return jsonify({"status": "success", "message": "Contacto guardado"}), 201

    except Exception as e:
        print(f"Error guardando contacto: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/resolver-ambiguedad', methods=['POST'])
def resolver_ambiguedad():
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    try:
        data             = request.get_json()
        pregunta         = data.get('pregunta', '')
        respuesta        = data.get('respuesta', '')
        valor            = data.get('valor', '')
        campos_posibles  = data.get('campos_posibles', [])

        campos_desc = ', '.join([f'"{c["nombre"]}" (ID: {c["id"]})' for c in campos_posibles])

        prompt = f"""El sistema le preguntó al usuario: "{pregunta}"
El usuario respondió: "{respuesta}"

El valor en cuestión es: "{valor}"
Los campos posibles son: {campos_desc}

Determina a cuál campo pertenece el valor según la respuesta del usuario.
Devuelve SOLO un JSON:
{{"campo_id": "el ID del campo elegido", "valor": "{valor}"}}

Si no puedes determinar el campo, devuelve:
{{"campo_id": "", "valor": ""}}"""

        response = openai_client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "user", "content": prompt}
            ],
            temperature=0.1,
            max_tokens=200
        )

        respuesta_texto = response.choices[0].message.content.strip()
        respuesta_texto = respuesta_texto.replace('```json', '').replace('```', '').strip()
        resultado = json.loads(respuesta_texto)

        print(f"[AMBIGUEDAD] Valor '{valor}' → Campo: {resultado.get('campo_id', 'ninguno')}")

        return jsonify({
            'success': True,
            'campo_id': resultado.get('campo_id', ''),
            'valor': resultado.get('valor', valor)
        })

    except Exception as e:
        print(f"[AMBIGUEDAD] Error: {e}")
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/exportar-formulario/<int:project_id>/<int:formulario_id>')
def exportar_formulario(project_id, formulario_id):
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    try:
        fecha_desde = request.args.get('desde', '')
        fecha_hasta = request.args.get('hasta', '')

        with db_connection() as (conn, cursor):
            # Obtener nombre del proyecto
            cursor.execute("SELECT nombre_proyecto FROM proyectos WHERE id = %s", (project_id,))
            proyecto_row = cursor.fetchone()
            nombre_proyecto = proyecto_row[0] if proyecto_row else 'Proyecto'

            # Obtener nombre del formulario y sus campos
            cursor.execute("""
                SELECT f.nombre, f.campos
                FROM formularios f
                WHERE f.id = %s AND f.empresa_id = %s
            """, (formulario_id, session.get('empresa_id')))
            form_row = cursor.fetchone()
            if not form_row:
                return jsonify({'error': 'Formulario no encontrado'}), 404

            nombre_formulario = form_row[0]
            campos_config = form_row[1] or []

            def es_grupo(item):
                return isinstance(item, dict) and item.get('tipo') == 'grupo'

            # Obtener IDs de campos (excluyendo los contenedores de grupo,
            # que no tienen 'id' propio — solo agrupan otros campos)
            campo_ids = [
                (item['id'] if isinstance(item, dict) else item)
                for item in campos_config
                if not es_grupo(item)
            ]

            # Guardamos también los grupos, para poder ubicar sus 'gid' por nombre
            grupos_config = [item for item in campos_config if es_grupo(item)]

            # Obtener nombres de campos
            campos_map = {}
            if campo_ids:
                cursor.execute("""
                    SELECT id, nombre FROM campos_globales WHERE id = ANY(%s)
                """, (campo_ids,))
                for r in cursor.fetchall():
                    campos_map[str(r[0])] = r[1]

            # Obtener registros con filtro de fechas
            query = """
                SELECT rf.respuestas, rf.created_at
                FROM respuestas_formulario rf
                WHERE rf.id_proyecto = %s AND rf.formulario_id = %s
            """
            params = [project_id, formulario_id]

            if fecha_desde:
                query += " AND rf.created_at::date >= %s"
                params.append(fecha_desde)
            if fecha_hasta:
                query += " AND rf.created_at::date <= %s"
                params.append(fecha_hasta)

            query += " ORDER BY rf.created_at ASC"
            cursor.execute(query, params)
            registros = cursor.fetchall()

        # ── Mapeo de campos por nombre → campo_id ──
        nombre_a_id = {v: k for k, v in campos_map.items()}

        # ── Ubicar los gid de los grupos "Paro Programado" y "Paro No Programado" por nombre ──
        def _norm(s):
            return (s or '').strip().lower()

        def _gid_de_grupo(condicion):
            return next(
                (g.get('gid') for g in grupos_config if condicion(_norm(g.get('nombre', '')))),
                None
            )

        gid_pp  = _gid_de_grupo(lambda n: 'programado' in n and 'no programado' not in n)
        gid_pnp = _gid_de_grupo(lambda n: 'no programado' in n)

        # Meses en español
        meses_es = {
            1: 'ENERO', 2: 'FEBRERO', 3: 'MARZO', 4: 'ABRIL',
            5: 'MAYO', 6: 'JUNIO', 7: 'JULIO', 8: 'AGOSTO',
            9: 'SEPTIEMBRE', 10: 'OCTUBRE', 11: 'NOVIEMBRE', 12: 'DICIEMBRE'
        }

        # Definir columnas del Excel
        columnas_excel = [
            'Mes', 'Día', 'Turno', 'Línea', 'Código', 'Nombre Producto',
            'Cant. Aceptada', 'Horas Maq.', 'H.PP', 'Codigo PP',
            'Causa PP', 'H.PNP', 'Codigo PNP', 'Causa PNP'
        ]

        # ── Mapeo de columnas → (nombre real del campo, qué parte leer) ──
        # 'valor'  = el valor tal cual quedó guardado en ese campo
        # 'codigo' = el código asociado (dataset.asociado) que viaja junto a ese campo
        #
        # Nota: "Código de producto" y "Código Paro (No) Programado" son selects
        # que traen dos datos en un solo campo: el texto visible (nombre/causa)
        # y el código asociado — por eso una misma clave de campo alimenta 2 columnas.

        columnas_raiz = {
            'Turno':           ('N° de Turno', 'valor'),
            'Línea':           ('Línea de Producción', 'valor'),
            'Código':          ('Código de producto', 'valor'),
            'Nombre Producto': ('Código de producto', 'codigo'),
            'Cant. Aceptada':  ('Cantidad Aceptada', 'valor'),
            'Horas Maq.':      ('Tiempo de Máquina (Hora M)', 'valor'),
        }

        columnas_grupo_pp = {
            'H.PP':      ('Horas Paro Programado', 'valor'),
            'Causa PP':  ('Código Paro Programado', 'codigo'),
            'Codigo PP': ('Código Paro Programado', 'valor'),
        }

        columnas_grupo_pnp = {
            'H.PNP':      ('Horas Paro No Programado', 'valor'),
            'Causa PNP':  ('Código Paro No Programado', 'codigo'),
            'Codigo PNP': ('Código Paro No Programado', 'valor'),
        }

        def _leer(fuente, campo_nombre, parte):
            """Lee 'valor' o 'codigo' de un campo, dentro de resp o de un bloque de grupo."""
            campo_id = nombre_a_id.get(campo_nombre, '')
            if not campo_id:
                return ''
            clave = campo_id + '_codigo' if parte == 'codigo' else campo_id
            return fuente.get(clave, '')

        def _fecha_produccion(fuente):
            """Lee el campo 'Fecha de producción' y lo convierte a date.
            Devuelve None si está vacío o no es parseable."""
            campo_id = nombre_a_id.get('Fecha de producción', '')
            if not campo_id:
                return None
            raw = str(fuente.get(campo_id, '') or '').strip()
            if not raw:
                return None
            for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
                try:
                    return datetime.strptime(raw[:10], fmt).date()
                except ValueError:
                    continue
            return None
        
        # Columnas que deben quedar como número en el Excel.
        # Los códigos de paro son numéricos sin ceros a la izquierda
        # (2, 34, 90), así que se convierten a entero sin riesgo.
        COLUMNAS_ENTERAS   = {'Turno', 'Cant. Aceptada', 'Codigo PP', 'Codigo PNP'}
        COLUMNAS_DECIMALES = {'Horas Maq.', 'H.PP', 'H.PNP'}

        def _a_numero(valor, entero=False):
            """Convierte un string numérico a int/float para que Excel lo
            reconozca como número. Devuelve None si está vacío o no es
            convertible, para que la celda quede vacía de verdad (una
            cadena '' haría que la columna se lea como texto)."""
            if valor is None or valor == '':
                return None
            if isinstance(valor, (int, float)):
                return int(valor) if entero else valor
            try:
                num = float(str(valor).strip().replace(',', '.'))
            except ValueError:
                return None
            return int(round(num)) if entero else num

        # ── Generar Excel ──
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Reporte'

        # Estilos
        header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='333333', end_color='333333', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_font = Font(name='Arial', size=10)
        cell_align = Alignment(vertical='center')
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        # Título
        ws.merge_cells('A1:N1')
        title_cell = ws['A1']
        title_cell.value = f'{nombre_formulario} — {nombre_proyecto}'
        title_cell.font = Font(name='Arial', bold=True, size=13)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Headers
        for col_idx, col_name in enumerate(columnas_excel, 1):
            cell = ws.cell(row=2, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Datos (una o varias filas por registro, según cuántos bloques de PP/PNP tenga)
        fila_actual = 3
        for respuestas, created_at in registros:
            resp = respuestas if isinstance(respuestas, dict) else {}
            repeticiones = resp.get('__repeticiones') or {}

            # Mes y Día salen del campo del formulario, no de created_at
            fecha_prod = _fecha_produccion(resp)

            bloques_pp  = repeticiones.get(gid_pp, [])  if gid_pp  else []
            bloques_pnp = repeticiones.get(gid_pnp, []) if gid_pnp else []
            num_filas   = max(len(bloques_pp), len(bloques_pnp), 1)

            for i in range(num_filas):
                bloque_pp  = bloques_pp[i]  if i < len(bloques_pp)  else {}
                bloque_pnp = bloques_pnp[i] if i < len(bloques_pnp) else {}

                for col_idx, col_name in enumerate(columnas_excel, 1):
                    valor = ''

                    if col_name == 'Mes':
                        valor = meses_es.get(fecha_prod.month, '') if fecha_prod else ''
                    elif col_name == 'Día':
                        valor = fecha_prod.day if fecha_prod else ''
                    elif col_name in columnas_grupo_pp:
                        campo_nombre, parte = columnas_grupo_pp[col_name]
                        valor = _leer(bloque_pp, campo_nombre, parte)
                    elif col_name in columnas_grupo_pnp:
                        campo_nombre, parte = columnas_grupo_pnp[col_name]
                        valor = _leer(bloque_pnp, campo_nombre, parte)
                    elif col_name in columnas_raiz:
                        campo_nombre, parte = columnas_raiz[col_name]
                        valor = _leer(resp, campo_nombre, parte)

                    # Tipado numérico para tablas dinámicas
                    if col_name in COLUMNAS_ENTERAS:
                        valor = _a_numero(valor, entero=True)
                    elif col_name in COLUMNAS_DECIMALES:
                        valor = _a_numero(valor)
                    elif valor == '':
                        valor = None

                    cell = ws.cell(row=fila_actual, column=col_idx, value=valor)
                    cell.font = cell_font
                    cell.alignment = cell_align
                    cell.border = thin_border

                    # Formato visual de los números
                    if col_name in COLUMNAS_DECIMALES:
                        cell.number_format = '0.00'
                    elif col_name in COLUMNAS_ENTERAS or col_name == 'Día':
                        cell.number_format = '0'

                    if col_name in {'Codigo PP', 'Codigo PNP'}:
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                fila_actual += 1

        # Ajustar anchos de columna
        anchos = {
            'A': 12, 'B': 8, 'C': 8, 'D': 14, 'E': 12, 'F': 45,
            'G': 16, 'H': 12, 'I': 10, 'J': 12,
            'K': 30, 'L': 10, 'M': 12, 'N': 30
        }
        for col, ancho in anchos.items():
            ws.column_dimensions[col].width = ancho

        # Filtro automático
        ws.auto_filter.ref = f'A2:N{fila_actual - 1}'

        # Congelar encabezados
        ws.freeze_panes = 'A3'

        # Generar archivo
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f'Reporte_{nombre_formulario}_{nombre_proyecto}.xlsx'
        filename = filename.replace(' ', '_')

        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )

    except Exception as e:
        print(f"Error al exportar: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/text-to-speech', methods=['POST'])
def text_to_speech():
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    try:
        data  = request.get_json()
        texto = data.get('texto', '')

        if not texto:
            return jsonify({'error': 'No hay texto'}), 400

        response = openai_client.audio.speech.create(
            model="gpt-4o-mini-tts",
            voice="alloy",
            input=texto,
            instructions="Habla rápido en español latinoamericano neutro. Acento colombiano profesional. Sin pausas innecesarias. Ritmo ágil como si tuvieras prisa pero pronunciando claro."
        )

        # Devolver el audio como MP3
        audio_bytes = response.content
        return Response(audio_bytes, mimetype='audio/mpeg')

    except Exception as e:
        print(f"[TTS] Error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/detalleContacto/<int:id_contacto>')
def detalleContacto(id_contacto):
    if 'user_id' not in session:
        return redirect(url_for('principalscreen'))
 
    conn = None
    try:
        #conn, cursor = get_db_connection()
        with db_connection() as (conn, cursor):

            cursor.execute("""
                SELECT 
                    c.id, c.nombre, c.empresa, c.cargo,
                    c.telefono, c.email, c.ciudad, c.notas,
                    c.created_at,
                    u.name, u.apellido, u.cargo as user_cargo
                FROM contactos c
                LEFT JOIN usuario u ON u.user_id = c.user_id
                WHERE c.id = %s
            """, (id_contacto,))
    
            row = cursor.fetchone()
            if not row:
                return redirect(url_for('registros'))
    
            created_at = row[8]
            #temporal
            print(f"created_at raw: {created_at}")
            print(f"created_at tzinfo: {created_at.tzinfo if created_at else 'None'}")
            nombre     = row[9] or ''
            apellido   = row[10] or ''
            iniciales  = (nombre[0] + apellido[0]).upper() if nombre and apellido else '??'

            # Convertir UTC a Colombia
            if created_at:
                colombia_tz    = pytz.timezone('America/Bogota')
                created_at_col = created_at.replace(tzinfo=timezone.utc).astimezone(colombia_tz)
                fecha_str      = created_at_col.strftime('%d %b %Y')
                created_texto  = created_at_col.strftime('%d %b %Y - %H:%M')
            else:
                fecha_str     = ''
                created_texto = ''
    
            contacto = {
                'id':                id_contacto,
                'nombre':            row[1],
                'empresa':           row[2] or '',
                'cargo':             row[3] or '',
                'telefono':          row[4] or '',
                'email':             row[5] or '',
                'ciudad':            row[6] or '',
                'notas':             row[7] or '',
                'fecha':            fecha_str,
                'created_at_texto': created_texto,
                #'fecha':             created_at.strftime('%d %b %Y') if created_at else '',
                #'created_at_texto':  created_at.strftime('%d %b %Y - %H:%M') if created_at else '',
                'usuario_nombre':    f"{nombre} {apellido}".strip() or 'Sin asignar',
                'usuario_cargo':     row[11] or '',
                'usuario_iniciales': iniciales,
                'imagenes':          []
            }
    
            cursor.execute("""
                SELECT imagen_url, descripcion
                FROM contacto_imagenes
                WHERE contacto_id = %s
            """, (id_contacto,))
    
            for img_row in cursor.fetchall():
                if img_row[0]:
                    contacto['imagenes'].append({
                        'url':  img_row[0],
                        'desc': img_row[1] or ''
                    })
    
            cursor.execute("""
                SELECT color_primario, color_secundario
                FROM empresas WHERE id = %s
            """, (session.get('empresa_id'),))
            empresa_row      = cursor.fetchone()
            color_primario   = empresa_row[0] if empresa_row else '#FFAF33'
            color_secundario = empresa_row[1] if empresa_row else '#E3E3E3'

            return render_template('detalleContacto.html',
                                contacto=contacto,
                                color_primario=color_primario,
                                color_secundario=color_secundario)
 
    except Exception as e:
        print(f"Error en detalleContacto: {e}")
        return redirect(url_for('registros'))


# ── Tipos de Proyecto ──────────────────────────────────────────

@app.route('/api/tipos-proyecto', methods=['GET'])
def obtener_tipos_proyecto():
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    try:
        empresa_id = session.get('empresa_id')
        print(f"[TIPOS PROYECTO] empresa_id={empresa_id}")

        with db_connection() as (conn, cursor):

            print("[TIPOS PROYECTO] Ejecutando SELECT")

            cursor.execute("""
                SELECT id, nombre, descripcion, campos, created_at
                FROM tipos_proyecto
                WHERE empresa_id = %s
                ORDER BY created_at DESC
            """, (empresa_id,))

            rows = cursor.fetchall()

            print(f"[TIPOS PROYECTO] Registros encontrados: {len(rows)}")

            tipos = []

            for r in rows:
                print(f"[TIPOS PROYECTO] Procesando ID={r[0]} created_at={r[4]}")

                tipos.append({
                    'id': r[0],
                    'nombre': r[1],
                    'descripcion': r[2],
                    'campos': r[3] or [],
                    'created_at': r[4].strftime('%d/%m/%Y') if r[4] else ''
                })

            return jsonify({'tipos': tipos})

    except Exception as e:
        print(f"[ERROR] obtener_tipos_proyecto: {e}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/tipos-proyecto', methods=['POST'])
def crear_tipo_proyecto():
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        data        = request.get_json()
        nombre      = data.get('nombre', '').strip()
        descripcion = data.get('descripcion', '')
        campos      = data.get('campos', [])

        if not nombre:
            return jsonify({'error': 'El nombre es obligatorio'}), 400

        with db_connection() as (conn, cursor):
            cursor.execute("""
                INSERT INTO tipos_proyecto (empresa_id, nombre, descripcion, campos)
                VALUES (%s, %s, %s, %s)
                RETURNING id
            """, (session.get('empresa_id'), nombre, descripcion, json.dumps(campos)))
            nuevo_id = cursor.fetchone()[0]
            return jsonify({'success': True, 'id': nuevo_id})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/tipos-proyecto/<int:tipo_id>', methods=['PUT'])
def actualizar_tipo_proyecto(tipo_id):
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        data        = request.get_json()
        nombre      = data.get('nombre', '').strip()
        descripcion = data.get('descripcion', '')
        campos      = data.get('campos', [])

        with db_connection() as (conn, cursor):
            cursor.execute("""
                UPDATE tipos_proyecto
                SET nombre = %s, descripcion = %s, campos = %s
                WHERE id = %s AND empresa_id = %s
            """, (nombre, descripcion, json.dumps(campos), tipo_id, session.get('empresa_id')))
            return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/tipos-proyecto/<int:tipo_id>', methods=['DELETE'])
def eliminar_tipo_proyecto(tipo_id):
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        with db_connection() as (conn, cursor):
            cursor.execute("""
                DELETE FROM tipos_proyecto
                WHERE id = %s AND empresa_id = %s
            """, (tipo_id, session.get('empresa_id')))
            return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/formContacto')
def form_contacto():
    if 'user_id' not in session:
        return redirect(url_for('principalscreen'))

    color_primario   = '#FFAF33'
    color_secundario = '#E3E3E3'
    try:
        with db_connection() as (conn, cursor):
            cursor.execute("""
                SELECT color_primario, color_secundario, logo_url
                FROM empresas WHERE id = %s
            """, (session.get('empresa_id'),))
            row = cursor.fetchone()
            if row:
                color_primario   = row[0] or '#FFAF33'
                color_secundario = row[1] or '#E3E3E3'
                logo_actual      = row[2] or None
    except Exception as e:
        print(f"Error cargando colores formContacto: {e}")

    return render_template('formContacto.html',
                       color_primario=color_primario,
                       color_secundario=color_secundario,
                       logo_actual=logo_actual)


@app.route('/detalleRegistro/<int:id_registro>')
def detalleRegistro(id_registro):
    if 'user_id' not in session:
        return redirect(url_for('principalscreen'))
 
    conn = None
    try:
        conn, cursor = get_db_connection()
 
        # Traer el registro con datos del usuario
        cursor.execute("""
            SELECT 
                r.id,
                r.actividad,
                r.descripcion_actividad,
                r.estado,
                r.porcentaje_avance,
                r.fecha,
                r.created_at,
                u.name,
                u.apellido,
                u.cargo
            FROM registros r
            LEFT JOIN usuario u ON u.user_id = r.user_id
            WHERE r.id = %s
        """, (id_registro,))
 
        row = cursor.fetchone()
 
        if not row:
            return redirect(url_for('registros'))
 
        # Extraer hora de created_at
        created_at = row[6]
        #hora = created_at.strftime('%I:%M %p') if created_at else None
        
        if created_at:
            colombia_tz  = pytz.timezone('America/Bogota')
            created_at_col = created_at.replace(tzinfo=timezone.utc).astimezone(colombia_tz)
            hora = created_at_col.strftime('%I:%M %p')
        else:
            hora = None
        created_at_texto = created_at.strftime('%d %b %Y - %H:%M') if created_at else None
 
        # Iniciales del usuario
        nombre  = row[7] or ''
        apellido = row[8] or ''
        iniciales = (nombre[0] + apellido[0]).upper() if nombre and apellido else '??'
 
        registro = {
            'id_registro':      row[0],
            'actividad':        row[1],
            'descripcion':      row[2],
            'estado':           row[3] or 'Sin estado',
            'avance':           row[4] or 0,
            'fecha':            row[5].strftime('%d %b %Y') if row[5] else '',
            'hora':             hora,
            'created_at_texto': created_at_texto,
            'usuario_nombre':   f"{nombre} {apellido}".strip() or 'Sin asignar',
            'usuario_cargo':    row[9] or '',
            'usuario_iniciales': iniciales,
            'fotos':            []
        }
 
        # Traer fotos del registro
        cursor.execute("""
            SELECT imagen_url, description
            FROM fotos_registro
            WHERE id_registro = %s
        """, (id_registro,))

        for foto_row in cursor.fetchall():
            url  = foto_row[0]
            desc = foto_row[1] or ''
            if url:
                registro['fotos'].append({'url': url, 'desc': desc})

        return render_template('detalleRegistro.html', registro=registro)
 
    except Exception as e:
        print(f"Error en detalleRegistro: {e}")
        return redirect(url_for('registros'))
    finally:
        if conn:
            cursor.close()
            connection_pool.putconn(conn)

@app.route('/disciplinerecords')
def disciplinerecords():
    return render_template('disciplinerecords.html')

@app.route('/projectdetails')
def projectdetails():
    return render_template('projectdetails.html')

import json

@app.route('/guardar_reporte_terranovus', methods=['POST'])
def guardar_reporte_terranovus():
    data = request.json
    conn = None
    try:
        conn, cursor = get_db_connection()

        # Extraer el ID del proyecto y del usuario
        id_proyecto = data.get('id_proyecto')
        user_id = session.get('user_id')

        # Recorrer cada actividad enviada desde el frontend
        for nota in data.get('notas', []):
            # 1. Insertar en la tabla maestra de registros
            cursor.execute("""
                INSERT INTO registros(
                    id_proyecto, actividad, descripcion_actividad, 
                    estado, porcentaje_avance, user_id, empresa_id
                ) VALUES (%s, %s, %s, %s, %s, %s, %s) 
                RETURNING id
            """, (
                id_proyecto, 
                nota.get('titulo'), # Campo 'Actividad' en tu SQL
                nota.get('texto'),  # Campo 'descripcion_actividad'
                nota.get('estado'), 
                nota.get('avance'), 
                user_id,
                session.get('empresa_id')
            ))
            
            id_registro = cursor.fetchone()[0]

            # 2. Insertar las fotos asociadas a ESTA actividad específica
            for foto_obj in nota.get('fotos_detalle', []):
                cursor.execute("""
                    INSERT INTO fotos_registro (
                        id_registro, imagen_url, description, empresa_id
                    ) VALUES (%s, %s, %s, %s)
                """, (
                    id_registro,
                    foto_obj.get('imagen_url'),
                    foto_obj.get('description'),
                    session.get('empresa_id')
                ))

        conn.commit()
        return jsonify({"status": "success", "message": "Reporte guardado correctamente"}), 201

    except Exception as e:
        if conn: conn.rollback()
        print(f"Error al guardar: {e}")
        return jsonify({"status": "error", "error": str(e)}), 500
    finally:
        if conn:
            cursor.close()
            connection_pool.putconn(conn)

@app.route('/add_project', methods=['GET', 'POST'])
def add_project():
    if 'user_id' not in session:
        return jsonify({"error": "No autorizado"}), 401

    if request.method == 'POST':
        try:
            data       = request.json
            empresa_id = session.get('empresa_id', 1)

            with db_connection() as (conn, cursor):
                cursor.execute("""
                    INSERT INTO proyectos (
                        nombre_proyecto, fecha_inicio, fecha_fin,
                        tipo_proyecto_id, datos_tipo,
                        user_id, empresa_id
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s)
                    RETURNING id
                """, (
                    data.get('project-name'),
                    data.get('start-date'),
                    data.get('end-date'),
                    data.get('tipo_proyecto_id'),
                    json.dumps(data.get('datos_tipo', {})),
                    session['user_id'],
                    empresa_id
                ))

                nuevo_id = cursor.fetchone()[0]

                # Miembros del equipo
                miembros = data.get('miembros', [])
                if not miembros:
                    miembros = [session['user_id']]

                for uid in miembros:
                    cursor.execute("""
                        INSERT INTO proyecto_usuarios (id_proyecto, user_id, empresa_id)
                        VALUES (%s, %s, %s)
                    """, (nuevo_id, uid, empresa_id))

                # Formularios asociados
                formularios_ids = data.get('formularios', [])
                for fid in formularios_ids:
                    cursor.execute("""
                        INSERT INTO proyecto_formularios (proyecto_id, formulario_id, empresa_id)
                        VALUES (%s, %s, %s)
                    """, (nuevo_id, fid, empresa_id))

            return jsonify({"status": "success", "message": "Proyecto registrado exitosamente"}), 201

        except Exception as e:
            print(f"Error en BD: {str(e)}")
            return jsonify({"status": "error", "error": str(e)}), 500

    # GET
    usuarios         = []
    color_primario   = '#FFAF33'
    color_secundario = '#E3E3E3'
    try:
        with db_connection() as (conn, cursor):
            cursor.execute("""
                SELECT user_id, name, apellido, cargo 
                FROM usuario 
                WHERE estado = 'activo'
                AND empresa_id = %s
                ORDER BY name ASC
            """, (session.get('empresa_id'),))
            for row in cursor.fetchall():
                usuarios.append({
                    'user_id':  row[0],
                    'name':     row[1],
                    'apellido': row[2],
                    'cargo':    row[3] or 'Sin cargo'
                })

            cursor.execute("""
                SELECT color_primario, color_secundario
                FROM empresas WHERE id = %s
            """, (session.get('empresa_id'),))
            empresa_row = cursor.fetchone()
            if empresa_row:
                color_primario   = empresa_row[0] or '#FFAF33'
                color_secundario = empresa_row[1] or '#E3E3E3'
                # Cargar tipos de proyecto
                cursor.execute("""
                    SELECT id, nombre, descripcion, campos
                    FROM tipos_proyecto
                    WHERE empresa_id = %s
                    ORDER BY nombre ASC
                """, (session.get('empresa_id'),))
                tipos_proyecto = [
                    {'id': r[0], 'nombre': r[1], 'descripcion': r[2], 'campos': r[3] or []}
                    for r in cursor.fetchall()
                ]

                # Cargar formularios disponibles
                cursor.execute("""
                    SELECT id, nombre, descripcion
                    FROM formularios
                    WHERE empresa_id = %s
                    ORDER BY nombre ASC
                """, (session.get('empresa_id'),))
                formularios = [
                    {'id': r[0], 'nombre': r[1], 'descripcion': r[2] or ''}
                    for r in cursor.fetchall()
                ]

                # Cargar campos globales de proyecto para resolver los tipos
                cursor.execute("""
                    SELECT id, nombre, tipo, opciones, configuracion, es_sistema
                    FROM campos_globales
                    WHERE empresa_id = %s AND objeto = 'proyecto'
                    ORDER BY es_sistema DESC, created_at DESC
                """, (session.get('empresa_id'),))
                campos_proyecto = [
                    {'id': r[0], 'nombre': r[1], 'tipo': r[2], 'opciones': r[3] or [], 'configuracion': r[4] or {}, 'es_sistema': r[5] or False}
                    for r in cursor.fetchall()
                ]
    except Exception as e:
        print(f"Error en add_project GET: {e}")

    return render_template('addproject.html', usuarios=usuarios,
                        color_primario=color_primario,
                        color_secundario=color_secundario,
                        tipos_proyecto=tipos_proyecto,
                        formularios=formularios,
                        campos_proyecto=campos_proyecto)


#@app.route('/generar-hash')
#def generar_hash():
    #from werkzeug.security import generate_password_hash
    #return generate_password_hash('Bitacora2026*')


@app.route('/edit_project/<int:project_id>', methods=['GET', 'PUT', 'DELETE'])
def edit_project(project_id):
    if 'user_id' not in session:
        return jsonify({"error": "No autorizado"}), 401

    empresa_id = session.get('empresa_id')

    # ── PUT: actualizar proyecto ──
    if request.method == 'PUT':
        try:
            data = request.json
            with db_connection() as (conn, cursor):
                # Verificar que el proyecto pertenece a la empresa
                cursor.execute("""
                    SELECT id FROM proyectos
                    WHERE id = %s AND empresa_id = %s
                """, (project_id, empresa_id))
                if not cursor.fetchone():
                    return jsonify({"error": "Proyecto no encontrado"}), 404

                # Actualizar proyecto
                cursor.execute("""
                    UPDATE proyectos
                    SET nombre_proyecto = %s,
                        fecha_inicio    = %s,
                        fecha_fin       = %s,
                        datos_tipo      = %s
                    WHERE id = %s AND empresa_id = %s
                """, (
                    data.get('project-name'),
                    data.get('start-date'),
                    data.get('end-date'),
                    json.dumps(data.get('datos_tipo', {})),
                    project_id, empresa_id
                ))

                # Reemplazar miembros: borrar e insertar
                cursor.execute("""
                    DELETE FROM proyecto_usuarios
                    WHERE id_proyecto = %s
                """, (project_id,))

                miembros = data.get('miembros', [])
                if not miembros:
                    miembros = [session['user_id']]

                for uid in miembros:
                    cursor.execute("""
                        INSERT INTO proyecto_usuarios (id_proyecto, user_id, empresa_id)
                        VALUES (%s, %s, %s)
                    """, (project_id, uid, empresa_id))

                # Reemplazar formularios: borrar e insertar
                cursor.execute("""
                    DELETE FROM proyecto_formularios
                    WHERE proyecto_id = %s
                """, (project_id,))

                for fid in data.get('formularios', []):
                    cursor.execute("""
                        INSERT INTO proyecto_formularios (proyecto_id, formulario_id, empresa_id)
                        VALUES (%s, %s, %s)
                    """, (project_id, fid, empresa_id))

            return jsonify({"status": "success", "message": "Proyecto actualizado"}), 200

        except Exception as e:
            print(f"Error al actualizar proyecto: {e}")
            return jsonify({"status": "error", "error": str(e)}), 500

    # ── DELETE: eliminar proyecto ──
    if request.method == 'DELETE':
        try:
            with db_connection() as (conn, cursor):
                cursor.execute("""
                    DELETE FROM proyectos
                    WHERE id = %s AND empresa_id = %s
                """, (project_id, empresa_id))
            return jsonify({"status": "success"}), 200
        except Exception as e:
            return jsonify({"status": "error", "error": str(e)}), 500

    # ── GET: cargar página de edición ──
    proyecto         = None
    usuarios         = []
    miembros_actuales = []
    formularios_actuales = []
    color_primario   = '#FFAF33'
    color_secundario = '#E3E3E3'
    tipos_proyecto   = []
    formularios      = []
    campos_proyecto  = []

    try:
        with db_connection() as (conn, cursor):
            # Cargar el proyecto
            cursor.execute("""
                SELECT id, nombre_proyecto, fecha_inicio, fecha_fin,
                       tipo_proyecto_id, datos_tipo
                FROM proyectos
                WHERE id = %s AND empresa_id = %s
            """, (project_id, empresa_id))
            row = cursor.fetchone()
            if not row:
                return redirect(url_for('registros'))

            proyecto = {
                'id':               row[0],
                'nombre':           row[1],
                'fecha_inicio':     row[2].strftime('%Y-%m-%d') if row[2] else '',
                'fecha_fin':        row[3].strftime('%Y-%m-%d') if row[3] else '',
                'tipo_proyecto_id': row[4],
                'datos_tipo':       row[5] or {}
            }

            # Miembros actuales
            cursor.execute("""
                SELECT user_id FROM proyecto_usuarios
                WHERE id_proyecto = %s
            """, (project_id,))
            miembros_actuales = [r[0] for r in cursor.fetchall()]

            # Formularios actuales
            cursor.execute("""
                SELECT formulario_id FROM proyecto_formularios
                WHERE proyecto_id = %s
            """, (project_id,))
            formularios_actuales = [r[0] for r in cursor.fetchall()]

            # Usuarios disponibles
            cursor.execute("""
                SELECT user_id, name, apellido, cargo
                FROM usuario
                WHERE estado = 'activo' AND empresa_id = %s
                ORDER BY name ASC
            """, (empresa_id,))
            for r in cursor.fetchall():
                usuarios.append({
                    'user_id':  r[0],
                    'name':     r[1],
                    'apellido': r[2],
                    'cargo':    r[3] or 'Sin cargo'
                })

            # Colores
            cursor.execute("""
                SELECT color_primario, color_secundario
                FROM empresas WHERE id = %s
            """, (empresa_id,))
            emp = cursor.fetchone()
            if emp:
                color_primario   = emp[0] or '#FFAF33'
                color_secundario = emp[1] or '#E3E3E3'

            # Tipos de proyecto (para resolver el tipo actual)
            cursor.execute("""
                SELECT id, nombre, descripcion, campos
                FROM tipos_proyecto
                WHERE empresa_id = %s
            """, (empresa_id,))
            tipos_proyecto = [
                {'id': r[0], 'nombre': r[1], 'descripcion': r[2], 'campos': r[3] or []}
                for r in cursor.fetchall()
            ]

            # Formularios disponibles
            cursor.execute("""
                SELECT id, nombre, descripcion
                FROM formularios
                WHERE empresa_id = %s
                ORDER BY nombre ASC
            """, (empresa_id,))
            formularios = [
                {'id': r[0], 'nombre': r[1], 'descripcion': r[2] or ''}
                for r in cursor.fetchall()
            ]

            # Campos globales de proyecto
            cursor.execute("""
                SELECT id, nombre, tipo, opciones, configuracion, es_sistema
                FROM campos_globales
                WHERE empresa_id = %s AND objeto = 'proyecto'
                ORDER BY es_sistema DESC, created_at DESC
            """, (empresa_id,))
            campos_proyecto = [
                {'id': r[0], 'nombre': r[1], 'tipo': r[2], 'opciones': r[3] or [],
                 'configuracion': r[4] or {}, 'es_sistema': r[5] or False}
                for r in cursor.fetchall()
            ]

    except Exception as e:
        print(f"Error en edit_project GET: {e}")
        return redirect(url_for('registros'))

    return render_template('editproject.html',
                           proyecto=proyecto,
                           usuarios=usuarios,
                           miembros_actuales=miembros_actuales,
                           formularios_actuales=formularios_actuales,
                           tipos_proyecto=tipos_proyecto,
                           formularios=formularios,
                           campos_proyecto=campos_proyecto,
                           color_primario=color_primario,
                           color_secundario=color_secundario)

@app.route('/ask', methods=['POST'])
def ask_question_route():
    data = request.json
    question = data.get('question', '')
    if not question:
        return jsonify({'error': 'No question provided'}), 400

    success = synthesize_speech(question)
    if success:
        return jsonify({'response': ''}), 200
    else:
        return jsonify({'error': 'Error al sintetizar la pregunta.'}), 500


@app.route('/guardar-inspeccion', methods=['POST'])
def guardar_inspeccion():
    try:
        data = request.json
        project_id = data.get('project_id')
        items = data.get('items', [])

        if not project_id or not items:
            return jsonify({'success': False, 'error': 'Datos incompletos'}), 400

        conn, cursor = get_db_connection()

        for item in items:
            cursor.execute("""
                INSERT INTO reporte_fiscalizacion (
                    id_proyecto, 
                    edificacion_zona, 
                    item_numero, 
                    area_inspeccionada, 
                    especificacion_tecnica, 
                    condicion_observada, 
                    cumple, 
                    observaciones, 
                    acciones_correctivas
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                project_id,
                item['edificacion_zona'],
                item['item_numero'],
                item['area_inspeccionada'],
                item['especificacion_tecnica'],
                item['condicion_observada'],
                item['cumple'],
                item['observaciones'],
                item['acciones_correctivas']
            ))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify({'success': True, 'mensaje': 'Inspección guardada correctamente'})
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/guardar-registro', methods=['POST'])
def guardar_registro():
    conn = None
    try:
        data = request.get_json()
        project_id = data.get('project_id')
        items = data.get('items', [])
        # Nota: 'fotos' y 'videos' ahora deberían venir dentro de cada objeto en 'items'
        
        if not project_id or not items:
            return jsonify({"error": "Faltan datos requeridos."}), 400

        conn, cursor = get_db_connection()

        # 1. Bucle principal para guardar cada ítem
        for item in items:
            cursor.execute("""
                INSERT INTO reporte_fiscalizacion (
                    id_proyecto, edificacion_zona, item_numero, area_inspeccionada, 
                    especificacion_tecnica, condicion_observada, cumple, 
                    observaciones, acciones_correctivas
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id_reporte
            """, (
                project_id, item.get('edificacion_zona'), item.get('item_numero'),
                item.get('area_inspeccionada'), item.get('especificacion_tecnica'),
                item.get('condicion_observada'), item.get('cumple'),
                item.get('observaciones'), item.get('acciones_correctivas')
            ))
            
            # Capturamos el ID específico de ESTE ítem recién insertado
            id_item_actual = cursor.fetchone()[0]

            # 2. GUARDAR FOTOS ESPECÍFICAS DE ESTE ÍTEM (NUEVA UBICACIÓN)
            # El frontend ahora debe enviar las fotos dentro de cada item
            fotos_item = item.get('fotos', []) 
            for foto_obj in fotos_item:
                cursor.execute("""
                    INSERT INTO fotos_registro (id_registro, imagen_base64, description) 
                    VALUES (%s, %s, %s)
                """, (id_item_actual, foto_obj.get('file_data'), foto_obj.get('description')))

            # 3. GUARDAR VIDEOS ESPECÍFICOS DE ESTE ÍTEM
            videos_item = item.get('videos', [])
            for video_obj in videos_item:
                cursor.execute("""
                    INSERT INTO videos_registro (id_registro, video_base64, description) 
                    VALUES (%s, %s, %s)
                """, (id_item_actual, video_obj.get('file_data'), video_obj.get('description')))

        conn.commit()
        return jsonify({"mensaje": "¡Reporte guardado!"}), 200

    except Exception as e:
        if conn: conn.rollback()
        print(f"Error: {str(e)}")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            cursor.close()
            connection_pool.putconn(conn)

@app.route('/eliminar-usuario/<int:user_id>', methods=['DELETE'])
def eliminar_usuario(user_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'No autenticado'}), 401

    # Evitar que el admin se elimine a sí mismo
    if user_id == session['user_id']:
        return jsonify({'success': False, 'error': 'No puedes eliminarte a ti mismo'})

    conn = None
    try:
        conn, cursor = get_db_connection()

        # Verificar que el usuario pertenece a la misma empresa
        cursor.execute("""
            SELECT user_id FROM usuario
            WHERE user_id = %s
            AND empresa_id = %s
        """, (user_id, session.get('empresa_id')))

        if not cursor.fetchone():
            return jsonify({'success': False, 'error': 'Usuario no encontrado en tu organización'})

        cursor.execute("DELETE FROM usuario WHERE user_id = %s", (user_id,))
        conn.commit()
        return jsonify({'success': True})

    except Exception as e:
        print(f"Error eliminando usuario: {e}")
        return jsonify({'success': False, 'error': str(e)})
    finally:
        if conn:
            cursor.close()
            connection_pool.putconn(conn)


@app.route('/subir-logo', methods=['POST'])
def subir_logo():
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    data = request.get_json()
    file_data = data.get('imagen')
    empresa_id = session.get('empresa_id')

    try:
        if ',' in file_data:
            header, b64 = file_data.split(',', 1)
            ext = 'png' if 'png' in header else 'jpg'
        else:
            b64, ext = file_data, 'jpg'

        imagen_bytes = base64.b64decode(b64)
        nombre_archivo = f"{uuid.uuid4()}.{ext}"
        ruta = f"logos/{empresa_id}/{nombre_archivo}"

        supabase_client.storage.from_('fotos-bitacora').upload(
            ruta,
            imagen_bytes,
            {"content-type": f"image/{ext}"}
        )

        url_publica = f"{SUPABASE_URL}/storage/v1/object/public/fotos-bitacora/{ruta}"

        # Guardar en BD
        with db_connection() as (conn, cursor):
            cursor.execute("""
                INSERT INTO empresa_logos (empresa_id, url, creado_por)
                VALUES (%s, %s, %s)
            """, (empresa_id, url_publica, session['user_id']))

        return jsonify({'url': url_publica}), 200

    except Exception as e:
        print(f"Error subiendo logo: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/logos-empresa', methods=['GET'])
def logos_empresa():
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        with db_connection() as (conn, cursor):
            cursor.execute("""
                SELECT id, url, created_at
                FROM empresa_logos
                WHERE empresa_id = %s
                ORDER BY created_at DESC
            """, (session.get('empresa_id'),))
            logos = [{'id': r[0], 'url': r[1], 'fecha': r[2].strftime('%d/%m/%Y') if r[2] else ''} 
                     for r in cursor.fetchall()]
        return jsonify({'logos': logos}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500 


@app.route('/eliminar-proyecto', methods=['POST'])
def eliminar_proyecto():
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    data = request.get_json()
    proyecto_id = data.get('id_proyecto')

    if not proyecto_id:
        return jsonify({'error': 'Falta el ID del proyecto'}), 400

    try:
        conn, cursor = get_db_connection()

        # Asegurarse de que el proyecto pertenece al usuario
        cursor.execute("""
            DELETE FROM proyectos
            WHERE id_proyecto = %s AND user_id = %s
        """, (proyecto_id, session['user_id']))
        conn.commit()

        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            cursor.close()
            connection_pool.putconn(conn)

@app.route('/api/analitica/pareto')
def analitica_pareto():
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    project_id    = request.args.get('project_id', '')
    formulario_id = request.args.get('formulario_id', '')
    linea         = request.args.get('linea', '')
    desde         = request.args.get('desde', '')
    hasta         = request.args.get('hasta', '')

    if not formulario_id:
        return jsonify({'error': 'Debes seleccionar un formulario'}), 400

    try:
        with db_connection() as (conn, cursor):
            cursor.execute("""
                SELECT campos FROM formularios
                WHERE id = %s AND empresa_id = %s
            """, (formulario_id, session.get('empresa_id')))
            form_row = cursor.fetchone()
            if not form_row:
                return jsonify({'error': 'Formulario no encontrado'}), 404

            campos_config = form_row[0] or []

            def es_grupo(item):
                return isinstance(item, dict) and item.get('tipo') == 'grupo'

            grupos_config = [item for item in campos_config if es_grupo(item)]
            campo_ids = [
                (item['id'] if isinstance(item, dict) else item)
                for item in campos_config if not es_grupo(item)
            ]

            campos_map = {}
            if campo_ids:
                cursor.execute("SELECT id, nombre FROM campos_globales WHERE id = ANY(%s)", (campo_ids,))
                for r in cursor.fetchall():
                    campos_map[str(r[0])] = r[1]
            nombre_a_id = {v: k for k, v in campos_map.items()}

            def _norm(s):
                return (s or '').strip().lower()

            def _gid_de_grupo(cond):
                return next((g.get('gid') for g in grupos_config if cond(_norm(g.get('nombre', '')))), None)

            gid_pp  = _gid_de_grupo(lambda n: 'programado' in n and 'no programado' not in n)
            gid_pnp = _gid_de_grupo(lambda n: 'no programado' in n)

            campo_id_linea     = nombre_a_id.get('Línea de Producción', '')
            campo_id_causa_pp  = nombre_a_id.get('Código Paro Programado', '')
            campo_id_horas_pp  = nombre_a_id.get('Horas Paro Programado', '')
            campo_id_causa_pnp = nombre_a_id.get('Código Paro No Programado', '')
            campo_id_horas_pnp = nombre_a_id.get('Horas Paro No Programado', '')

            query  = "SELECT rf.respuestas FROM respuestas_formulario rf WHERE rf.formulario_id = %s"
            params = [formulario_id]
            if project_id:
                query += " AND rf.id_proyecto = %s"
                params.append(project_id)
            if desde:
                query += " AND CASE WHEN rf.respuestas->>'96' = '' THEN NULL ELSE (rf.respuestas->>'96')::date END >= %s"
                params.append(desde)
            if hasta:
                query += " AND CASE WHEN rf.respuestas->>'96' = '' THEN NULL ELSE (rf.respuestas->>'96')::date END <= %s"
                params.append(hasta)

            cursor.execute(query, params)
            registros = cursor.fetchall()

        acumulado_pp, acumulado_pnp = {}, {}

        for (respuestas,) in registros:
            resp = respuestas if isinstance(respuestas, dict) else {}

            if linea and campo_id_linea and resp.get(campo_id_linea, '') != linea:
                continue

            repeticiones = resp.get('__repeticiones') or {}

            for bloque in (repeticiones.get(gid_pp, []) if gid_pp else []):
                codigo = bloque.get(campo_id_causa_pp, '') or 'sin-codigo'
                causa  = bloque.get(campo_id_causa_pp + '_codigo', '') or f'Código {codigo}'
                horas  = float(bloque.get(campo_id_horas_pp, '') or 0)
                acumulado_pp.setdefault(codigo, {'causa': causa, 'horas': 0})
                acumulado_pp[codigo]['horas'] += horas

            for bloque in (repeticiones.get(gid_pnp, []) if gid_pnp else []):
                codigo = bloque.get(campo_id_causa_pnp, '') or 'sin-codigo'
                causa  = bloque.get(campo_id_causa_pnp + '_codigo', '') or f'Código {codigo}'
                horas  = float(bloque.get(campo_id_horas_pnp, '') or 0)
                acumulado_pnp.setdefault(codigo, {'causa': causa, 'horas': 0})
                acumulado_pnp[codigo]['horas'] += horas

        a_lista = lambda dic: [{'causa': v['causa'], 'horas': round(v['horas'], 2)} for v in dic.values()]

        return jsonify({'pp': a_lista(acumulado_pp), 'pnp': a_lista(acumulado_pnp)})

    except Exception as e:
        print(f"Error en analitica_pareto: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/analitica')
def analitica():
    if 'user_id' not in session or session.get('user_rol') != 'admin':
        return redirect(url_for('registros'))

    logo_actual    = None
    color_primario = '#FFAF33'
    try:
        with db_connection() as (conn, cursor):
            cursor.execute("""
                SELECT logo_url, color_primario FROM empresas
                WHERE id = %s
            """, (session.get('empresa_id'),))
            emp = cursor.fetchone()
            if emp:
                logo_actual    = emp[0]
                color_primario = emp[1] or '#FFAF33'
    except Exception as e:
        print(f"Error cargando empresa en analitica: {e}")

    return render_template('analitica.html',
                           empresa_id=session.get('empresa_id'),
                           logo_actual=logo_actual,
                           color_primario=color_primario)

@app.route('/analitica/paros')
def analitica_paros():
    if 'user_id' not in session or session.get('user_rol') != 'admin':
        return redirect(url_for('registros'))

    logo_actual    = None
    color_primario = '#FFAF33'
    try:
        with db_connection() as (conn, cursor):
            cursor.execute("""
                SELECT logo_url, color_primario FROM empresas
                WHERE id = %s
            """, (session.get('empresa_id'),))
            emp = cursor.fetchone()
            if emp:
                logo_actual    = emp[0]
                color_primario = emp[1] or '#FFAF33'
    except Exception as e:
        print(f"Error cargando empresa en analitica_paros: {e}")

    return render_template('analitica_paros.html',
                           logo_actual=logo_actual,
                           color_primario=color_primario)

@app.route('/analitica/tablero/<int:tablero_id>')
def analitica_tablero(tablero_id):
    if 'user_id' not in session or session.get('user_rol') != 'admin':
        return redirect(url_for('registros'))

    logo_actual    = None
    color_primario = '#FFAF33'
    try:
        with db_connection() as (conn, cursor):
            cursor.execute("""
                SELECT logo_url, color_primario FROM empresas
                WHERE id = %s
            """, (session.get('empresa_id'),))
            emp = cursor.fetchone()
            if emp:
                logo_actual    = emp[0]
                color_primario = emp[1] or '#FFAF33'
    except Exception as e:
        print(f"Error cargando empresa en analitica_tablero: {e}")

    return render_template('analitica_tablero.html',
                           tablero_id=tablero_id,
                           empresa_id=session.get('empresa_id'),
                           user_rol=session.get('user_rol'),
                           logo_actual=logo_actual,
                           color_primario=color_primario)

@app.route('/api/proyectos-usuario')
def api_proyectos_usuario():
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    proyectos = get_user_projects(session['user_id'])
    return jsonify({'proyectos': proyectos})

@app.route('/transcribe-audio', methods=['POST'])
def transcribe_audio():
    try:
        if 'audio' not in request.files:
            print("🔴 [WHISPER] No se recibió archivo de audio.")
            return jsonify({"error": "No se envió el archivo de audio"}), 400

        file = request.files['audio']
        print(f"📥 [WHISPER] Recibido archivo: {file.filename}")

        # Guardar el archivo temporalmente
        temp_input = tempfile.NamedTemporaryFile(delete=False, suffix=".webm")
        file.save(temp_input.name)
        print(f"💾 [WHISPER] Guardado en: {temp_input.name}")

        temp_wav = tempfile.NamedTemporaryFile(delete=False, suffix=".wav")
        formato_detectado = None

        try:
            print("🔍 [WHISPER] Intentando decodificar como webm...")
            audio = AudioSegment.from_file(temp_input.name, format="webm")
            print("✅ [WHISPER] Decodificado como webm.")
            formato_detectado = "webm"
        except Exception as e_webm:
            print("⚠️ [WHISPER] Falla al decodificar como webm:", str(e_webm))
            try:
                print("🔁 [WHISPER] Intentando decodificar como mp4...")
                audio = AudioSegment.from_file(temp_input.name, format="mp4")
                print("✅ [WHISPER] Decodificado como mp4.")
                formato_detectado = "mp4"
            except Exception as e_mp4:
                print("❌ [WHISPER] Fallo total al decodificar audio.")
                traceback.print_exc()
                return jsonify({
                    "error": "No se pudo procesar el audio.",
                    "error_webm": str(e_webm),
                    "error_mp4": str(e_mp4)
                }), 500

        # Exportar a WAV (Whisper acepta mp3, mp4, mpeg, mpga, m4a, wav, webm)
        audio.export(temp_wav.name, format="wav")
        # Duración real del audio (pydub la expone en ms)
        duracion_audio_ms = len(audio)
        print(f"⏱️ [WHISPER] Duración del audio: {duracion_audio_ms} ms")
        print("🔄 [WHISPER] Exportado a WAV:", temp_wav.name)

        # Transcribir con Whisper API de OpenAI
        try:
            t_inicio = time.time()
            with open(temp_wav.name, 'rb') as f:
                transcription = openai_client.audio.transcriptions.create(
                    model="gpt-4o-transcribe",
                    file=f,
                    language="es",
                    prompt="Bitácora industrial IAC. Registros de campo con proyectos, formularios, clientes, contratistas, evidencias, observaciones y actividades. Se dictan correos electrónicos en formato nombre.apellido@empresa.com, ana_lopez@gmail.com, contacto@iac.com.co, carlos-mesa@outlook.com. Nombres colombianos comunes: Juan Pérez, María García, Andrés López, Camilo Ramírez, Santiago Giraldo. Cuando se escuche arroba se escribe @, punto se escribe ., guion bajo se escribe _, guion medio se escribe -. Dominios frecuentes: .com, .co, .com.co, .org."
                )
            ms_transcripcion = int((time.time() - t_inicio) * 1000)
            texto = transcription.text
            print(f"✅ [WHISPER] {len(texto)} caracteres en {ms_transcripcion} ms")

            return jsonify({
                "text":              texto,
                "formato_detectado": formato_detectado,
                "servicio":          "whisper-api",
                # Métricas para el log de calidad
                "duracion_audio_ms": duracion_audio_ms,
                "ms_transcripcion":  ms_transcripcion,
                "modelo_transcribe": "gpt-4o-transcribe"
            })

        except Exception as e_whisper:
            print("❌ [WHISPER] Error al llamar a Whisper API:", str(e_whisper))
            traceback.print_exc()
            return jsonify({
                "error": "Error al transcribir con Whisper.",
                "detalle": str(e_whisper)
            }), 500

        finally:
            # Limpiar archivos temporales
            try:
                if os.path.exists(temp_input.name):
                    os.remove(temp_input.name)
                if os.path.exists(temp_wav.name):
                    os.remove(temp_wav.name)
            except Exception as e_cleanup:
                print(f"⚠️ [WHISPER] Error limpiando temporales: {e_cleanup}")

    except Exception as e:
        print("❌ [WHISPER] Error general en transcribe_audio:", str(e))
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

def _log_transcripcion(data, resultado, ms_interpretacion):
    """Registra la transcripción y lo que el modelo extrajo de ella.
    Nunca interrumpe el flujo: si falla, solo se imprime."""
    try:
        texto = (data.get('transcripcion') or '').strip()
        if not texto:
            return

        extraido = json.dumps({
            'campos':       resultado.get('campos', {}),
            'grupos':       resultado.get('grupos', {}),
            'ambiguedades': resultado.get('ambiguedades', []),
            'faltantes':    resultado.get('faltantes', [])
        }, ensure_ascii=False)

        empresa_id = data.get('_empresa_id') or session.get('empresa_id')
        user_id    = data.get('_user_id')    or session.get('user_id')

        with db_connection() as (conn, cursor):
            cursor.execute("""
                INSERT INTO transcripciones_log
                    (empresa_id, formulario_id, user_id, sesion_uuid, orden,
                     texto, extraido, duracion_audio_ms, formato_audio,
                     ms_transcripcion, ms_interpretacion,
                     modelo_transcribe, modelo_interpreta)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                empresa_id,
                data.get('formulario_id'),
                user_id,
                data.get('sesion_uuid') or str(uuid.uuid4()),
                data.get('orden') or 1,
                texto,
                extraido,
                data.get('duracion_audio_ms'),
                data.get('formato_audio'),
                data.get('ms_transcripcion'),
                ms_interpretacion,
                data.get('modelo_transcribe'),
                'gpt-4o-mini'
            ))
            conn.commit()
    except Exception as e:
        print(f"[LOG-TRANS] No se pudo registrar la transcripción: {e}")

@app.route('/api/distribuir-campos', methods=['POST'])
def distribuir_campos():
    # La web valida por sesión; la lógica vive en distribuir_campos_core.
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    data = request.get_json()
    t_inicio = time.time()
    resultado, codigo = distribuir_campos_core(data)
    ms_interpretacion = int((time.time() - t_inicio) * 1000)

    if codigo == 200:
        _log_transcripcion(data, resultado, ms_interpretacion)

    return jsonify(resultado), codigo


def distribuir_campos_core(data):
    """Lógica pura de distribución: recibe el dict data y devuelve
    (resultado_dict, codigo_http). No usa session ni request; la llaman
    la ruta web y el blueprint del APK. Sin duplicar la lógica de IA."""

    respuesta_texto = ''
    try:
        transcripcion = data.get('transcripcion', '')
        sueltos       = data.get('sueltos', [])
        grupos        = data.get('grupos', [])

        # Compatibilidad: si llega el formato viejo, se trata como campos sueltos
        if not sueltos and not grupos and data.get('campos'):
            sueltos = data.get('campos', [])

        if not transcripcion or (not sueltos and not grupos):
            return {'error': 'Faltan datos'}, 400

        def describir(c):
            desc = f'- {c["id"]}: "{c["nombre"]}" (tipo: {c["tipo"]})'
            if c.get('opciones'):
                opciones_desc = []
                conf = c.get('configuracion', {})
                etiq_izq = conf.get('etiqueta_izquierda', 'Opción')
                etiq_der = conf.get('etiqueta_derecha', 'Valor')
                for op in c['opciones']:
                    if isinstance(op, dict) and 'izquierda' in op:
                        opciones_desc.append(f'{etiq_izq}: {op["izquierda"]} → {etiq_der}: {op["derecha"]}')
                    elif isinstance(op, dict) and 'nombre' in op:
                        opciones_desc.append(f'{op["nombre"]} [código: {op.get("codigo", "")}]')
                    else:
                        opciones_desc.append(str(op))
                desc += f' (opciones: {"; ".join(opciones_desc)})'
            if c.get('requerido'):
                desc += ' [obligatorio]'
            return desc

        sueltos_texto = '\n'.join(describir(c) for c in sueltos) or '(ninguno)'

        grupos_desc = []
        for g in grupos:
            lineas = [f'GRUPO gid="{g["gid"]}" nombre="{g["nombre"]}"']
            lineas += ['  ' + describir(c) for c in g.get('campos', [])]
            lineas.append('  Bloques que ya existen en pantalla:')
            for b in g.get('bloques_actuales', []):
                vals = b.get('valores') or {}
                resumen = ', '.join(f'{k}="{v}"' for k, v in vals.items()) if vals else '(vacío)'
                lineas.append(f'    bloque {b["bloque"]}: {resumen}')
            grupos_desc.append('\n'.join(lineas))
        grupos_texto = '\n\n'.join(grupos_desc) or '(ninguno)'

        fecha_hoy = date.today().strftime('%Y-%m-%d')

        prompt_sistema = """Eres un asistente que extrae información de una transcripción de voz de un reporte de obra y la distribuye en los campos de un formulario.

El formulario tiene dos partes:
- CAMPOS SUELTOS: un solo valor cada uno.
- GRUPOS: conjuntos de campos que se repiten en BLOQUES. El usuario pudo haber hablado de varios (por ejemplo, varias actividades del día). Cada bloque es una ocurrencia distinta.

REGLAS GENERALES:
1. Devuelve SOLO un JSON válido, sin explicaciones ni markdown.
2. Las claves de los campos son SOLO el ID numérico como string ("1", "3", "15"). Nunca prefijos como "campo_".
3. NO incluyas los campos de los que el usuario no habló. Si no escuchaste nada para un campo, OMÍTELO del JSON. Nunca devuelvas "" para rellenar.
4. Para tipo numero, moneda o porcentaje: solo el número, sin símbolos.
5. Para tipo seleccion o seleccion_unica: elige la opción más cercana. Si ninguna coincide, omite el campo.
6. Para tipo booleano: true o false.
7. Para tipo fecha: formato YYYY-MM-DD. La fecha actual es {fecha_hoy}. 
   Si el usuario menciona una fecha SIN especificar el año (ej: "el dos de agosto", "el 15 de marzo"), 
   usa SIEMPRE el año de la fecha actual. 
   Si el usuario especifica el año explícitamente (ej: "dos de febrero del 2025", "el 3 de enero de 2024"), 
   usa ese año exacto sin modificarlo.
8. Extrae de forma inteligente: el usuario puede no decir el nombre exacto del campo pero sí dar el dato.
9. Limpia el texto: capitaliza nombres propios, corrige puntuación básica.
10. Para texto_largo u observaciones: resume y redacta de forma profesional y concisa, pero NUNCA omitas datos específicos como correos, nombres de personas, teléfonos, cantidades o direcciones.
11. Para campos de selección con valor asociado, el usuario puede mencionar cualquiera de los dos valores (izquierda o derecha). Si el campo muestra códigos en el desplegable y el usuario dice un código, devuelve el CÓDIGO. Si dice el nombre, busca qué código le corresponde y devuelve el CÓDIGO. Siempre devuelve el valor que corresponde a lo que se muestra en el desplegable.

REGLAS DE LOS GRUPOS (lo más importante):
12. Un grupo representa algo de lo que el usuario pudo hablar varias veces. Devuelve un elemento por cada ocurrencia REAL que mencione.
13. Cada elemento lleva "bloque": el número que el usuario mencionó cuando se refiere explícitamente a un bloque del formulario ("en la actividad 2", "el registro 3"). Si el usuario menciona varias ocurrencias distintas de forma seguida SIN referirse a bloques del formulario (ej: "el primero son albañiles, el segundo maestros, el tercero peones"), usa "bloque": "nuevo" para cada una — nunca repitas "bloque": 1 para ocurrencias distintas.
14. Si el usuario se refiere a un bloque por su contenido en vez de por número (ej. "el del vaciado"), usa el número de ese bloque según los bloques que ya existen en pantalla.
15. En "valores" pon SOLO los campos que el usuario mencionó para ese bloque. Los que no mencionó se omiten: lo que ya está escrito en pantalla se conserva solo.
16. Si el usuario NO usó marcadores ni números y no está claro que sean varias ocurrencias, trátalo como UNA SOLA: un elemento con "bloque": 1. Ante la duda, NO partas.
17. NUNCA inventes ocurrencias para rellenar. Si habló de una actividad, devuelve una.

AMBIGÜEDADES:
- Repórtalas SOLO para campos sueltos, y solo cuando un valor claramente dicho podría ir en más de un campo (ej. "Correo del cliente" vs "Correo del supervisor"). En ese caso no lo asignes a ninguno.
- NO reportes ambigüedad si por contexto es claro.
- NUNCA preguntes cuántas actividades son ni si algo es una o dos ocurrencias. Aplica la regla 15.

FORMATO DE RESPUESTA:
{
  "campos": { "9": "valor" },
  "grupos": {
    "gid_del_grupo": [
      { "bloque": 1, "valores": { "4": "Sergio Gómez" } },
      { "bloque": "nuevo", "valores": { "3": "Armado de columnas", "4": "Cuadrilla B" } }
    ]
  },
  "ambiguedades": [
    { "valor": "el valor ambiguo",
      "campos_posibles": [ {"id": "1", "nombre": "Campo 1"}, {"id": "2", "nombre": "Campo 2"} ],
      "pregunta": "pregunta corta y natural" }
  ],
  "faltantes": [
    { "campo_id": "id", "nombre": "nombre del campo", "pregunta": "pregunta natural y corta" }
  ]
}

- "faltantes": SOLO campos sueltos que tienen la etiqueta [obligatorio] en su descripción Y que quedaron sin valor tras analizar la transcripción. Si un campo NO tiene la etiqueta [obligatorio], jamás lo incluyas en faltantes aunque esté vacío — es opcional y el usuario eligió no mencionarlo. NO reportes faltantes de campos dentro de un grupo."
- Si no hay grupos, "grupos" debe ser {}. Si no hay ambigüedades o faltantes, arrays vacíos.
- Las preguntas: naturales, cortas y en español colombiano."""

        prompt_usuario = f"""Fecha actual: {fecha_hoy}

CAMPOS SUELTOS:
{sueltos_texto}

GRUPOS:
{grupos_texto}

Transcripción del usuario:
"{transcripcion}"

Devuelve SOLO el JSON con los valores extraídos."""

        response = openai_client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": prompt_sistema},
                {"role": "user", "content": prompt_usuario}
            ],
            temperature=0.1,
            max_tokens=2000
        )

        respuesta_texto = response.choices[0].message.content.strip()
        respuesta_texto = respuesta_texto.replace('```json', '').replace('```', '').strip()

        resultado = json.loads(respuesta_texto)

        if 'campos' in resultado or 'grupos' in resultado:
            campos_resultado = resultado.get('campos', {}) or {}
            grupos_resultado = resultado.get('grupos', {}) or {}
            ambiguedades     = resultado.get('ambiguedades', []) or []
            faltantes        = resultado.get('faltantes', []) or []
        else:
            campos_resultado = resultado
            grupos_resultado = {}
            ambiguedades     = []
            faltantes        = []

        # Descartar grupos que no existen en el formulario
        gids_validos     = {g['gid'] for g in grupos}
        grupos_resultado = {k: v for k, v in grupos_resultado.items()
                            if k in gids_validos and isinstance(v, list)}

        print(f"[DISTRIBUIR] Grupos resultado: {grupos_resultado}")
        bloques = sum(len(v) for v in grupos_resultado.values())
        print(f"[DISTRIBUIR] Sueltos: {len([v for v in campos_resultado.values() if v])}/{len(sueltos)}. "
              f"Grupos: {len(grupos_resultado)} ({bloques} bloques). "
              f"Ambigüedades: {len(ambiguedades)}. Faltantes: {len(faltantes)}")

        return {
            'success':      True,
            'campos':       campos_resultado,
            'grupos':       grupos_resultado,
            'ambiguedades': ambiguedades,
            'faltantes':    faltantes
        }, 200

    except json.JSONDecodeError as e:
        print(f"[DISTRIBUIR] Error parseando JSON: {e}")
        print(f"[DISTRIBUIR] Respuesta raw: {respuesta_texto}")
        return {'error': 'Error al interpretar la respuesta de la IA'}, 500
    except Exception as e:
        print(f"[DISTRIBUIR] Error: {e}")
        return {'error': str(e)}, 500

#Exportar registros seleccionados a Excel
@app.route('/exportar-registros-excel', methods=['POST'])
def exportar_registros_excel():
    registro_ids = request.form.getlist('registro_ids')
    project_id = request.form.get('project_id')

    if not registro_ids and not project_id:
        return "No se seleccionaron registros ni proyecto", 400

    try:
        conn, cursor = get_db_connection()

        if not registro_ids:
            cursor.execute("""
                SELECT id_registro, zona_intervencion, items, metros_lineales, proximas_tareas, foto_base64
                FROM registrosbitacoraeqing
                WHERE id_proyecto = %s
                ORDER BY id_registro DESC
            """, (project_id,))
        else:
            format_ids = tuple(map(int, registro_ids))
            cursor.execute("""
                SELECT id_registro, zona_intervencion, items, metros_lineales, proximas_tareas, foto_base64
                FROM registrosbitacoraeqing
                WHERE id_registro IN %s
                ORDER BY id_registro DESC
            """, (format_ids,))

        rows = cursor.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Registros"

        # Encabezado
        ws.append(["ID", "Zona de Intervención", "Ítems", "Metros Lineales", "Próximas Tareas", "Foto"])

        row_index = 2  # Comienza después del encabezado

        for row in rows:
            id_registro, zona, items, metros, tareas, foto_base64 = row
            ws.append([id_registro, zona, items, metros, tareas, ""])  # celda para imagen

            if foto_base64:
                try:
                    header, base64_data = foto_base64.split(',', 1) if ',' in foto_base64 else ('', foto_base64)
                    image_data = base64.b64decode(base64_data)
                    img = Image.open(io.BytesIO(image_data))
                    img.thumbnail((120, 120))  # redimensiona para celda
                    image_io = io.BytesIO()
                    img.save(image_io, format='PNG')
                    image_io.seek(0)

                    img_excel = ExcelImage(image_io)
                    img_excel.anchor = f"F{row_index}"
                    ws.add_image(img_excel)

                    # Ajustar altura de fila
                    ws.row_dimensions[row_index].height = 90
                except Exception as img_err:
                    print(f"Error al procesar imagen para registro {id_registro}: {img_err}")

            row_index += 1

        # Ajuste de anchos de columnas
        ws.column_dimensions['A'].width = 12  # ID
        ws.column_dimensions['B'].width = 30  # Zona de intervención
        ws.column_dimensions['C'].width = 25  # Ítems
        ws.column_dimensions['D'].width = 20  # Metros lineales
        ws.column_dimensions['E'].width = 35  # Próximas tareas
        ws.column_dimensions['F'].width = 18  # Imagen

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(output,
                         download_name="registros_bitacora.xlsx",
                         as_attachment=True,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        print(f"Error al exportar: {e}")
        return "Error al exportar", 500
    finally:
        if conn:
            cursor.close()
            connection_pool.putconn(conn)

@app.route('/exportar-proyectos-pdf', methods=['POST'])
def exportar_proyectos_pdf():
    project_ids = request.form.getlist('project_ids')
    if not project_ids:
        return "No se seleccionaron proyectos", 400

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)

    try:
        conn, cursor = get_db_connection()

        for pid in project_ids:
            # 1. Info del proyecto (Tabla: proyectos)
            cursor.execute("""
                SELECT nombre_proyecto, cliente, contratista, orden_de_trabajo, ubicacion, fecha_inicio 
                FROM proyectos WHERE id_proyecto = %s
            """, (pid,))
            proyecto = cursor.fetchone()
            if not proyecto: continue

            nombre, cliente, contratista, ot, ubicacion, f_inicio = proyecto
            pdf.add_page()
            
            # --- ENCABEZADO TÉCNICO (190mm Total) ---
            y_inicial = pdf.get_y()
            
            # Celda Logo (40mm)
            pdf.rect(10, y_inicial, 40, 20)
            logo_path = os.path.join('static', 'logo.png')
            if os.path.exists(logo_path):
                pdf.image(logo_path, x=15, y=y_inicial + 2, w=30)
            
            # Celda Título (150mm)
            pdf.set_xy(50, y_inicial)
            pdf.set_font("Arial", 'B', 14)
            pdf.cell(150, 20, "REPORTE TÉCNICO DE ACTIVIDADES", border=1, ln=True, align='C')

            # Filas de Información (Ancho total 190mm)
            pdf.set_font("Arial", 'B', 9)
            pdf.cell(30, 8, "PROYECTO:", border=1)
            pdf.set_font("Arial", '', 9)
            pdf.cell(65, 8, f"{nombre or ''}", border=1)
            pdf.set_font("Arial", 'B', 9)
            pdf.cell(40, 8, "CLIENTE:", border=1)
            pdf.set_font("Arial", '', 9)
            pdf.cell(55, 8, f"{cliente or ''}", border=1, ln=True)

            pdf.set_font("Arial", 'B', 9)
            pdf.cell(30, 8, "UBICACIÓN:", border=1)
            pdf.set_font("Arial", '', 9)
            pdf.cell(65, 8, f"{ubicacion or ''}", border=1)
            pdf.set_font("Arial", 'B', 9)
            pdf.cell(40, 8, "ORDEN DE TRABAJO:", border=1)
            pdf.set_font("Arial", '', 9)
            pdf.cell(55, 8, f"{ot or ''}", border=1, ln=True)
            
            pdf.ln(10)

            # --- REGISTROS DE ACTIVIDAD ---
            cursor.execute("""
                SELECT id_registro, actividad, descripcion_actividad, estado, porcentaje_avance, fecha
                FROM registros
                WHERE id_proyecto = %s ORDER BY fecha DESC
            """, (pid,))
            registros = cursor.fetchall()

            for reg in registros:
                id_reg, actividad, desc, estado, avance, fecha_reg = reg
                
                # Encabezado Actividad
                pdf.set_fill_color(255, 240, 220)
                pdf.set_font("Arial", 'B', 11)
                pdf.cell(190, 8, f"FECHA: {fecha_reg} - ACTIVIDAD: {actividad or 'Sin actividad'}", ln=True, fill=True, border='T')
                
                # Descripción
                pdf.set_font("Arial", '', 10)
                pdf.multi_cell(190, 6, f"Descripción: {desc or 'Sin descripción'}", border='LR')

                # FIX: Forzar X al margen izquierdo después del multi_cell
                pdf.set_x(10)

                # ESTADO Y AVANCE (95mm + 95mm = 190mm)
                pdf.set_font("Arial", 'B', 10)
                pdf.set_fill_color(245, 245, 245)
                pdf.cell(95, 8, f" ESTADO: {estado or 'N/A'}", border=1, fill=True)
                pdf.cell(95, 8, f" AVANCE: {avance or 0}%", border=1, fill=True, ln=True)
                
                # --- SECCIÓN DE EVIDENCIA FOTOGRÁFICA ---
                cursor.execute("SELECT imagen_base64, description FROM fotos_registro WHERE id_registro = %s", (id_reg,))
                fotos = cursor.fetchall()
                
                if fotos:
                    pdf.ln(2)
                    pdf.set_font("Arial", 'B', 10)
                    pdf.cell(0, 8, "EVIDENCIA:", ln=True)
                    
                    img_w = 60
                    current_x = 10
                    
                    for i, (foto_data, foto_desc) in enumerate(fotos):
                        # Salto de página preventivo
                        if pdf.get_y() > 220:
                            pdf.add_page()
                            pdf.ln(5)
                            current_x = 10

                        # Fila de 3 fotos
                        if i > 0 and i % 3 == 0:
                            pdf.set_y(pdf.get_y() + 50)
                            current_x = 10

                        try:
                            header, encoded = foto_data.split(",", 1) if "," in foto_data else ("", foto_data)
                            img_bytes = base64.b64decode(encoded)
                            
                            with NamedTemporaryFile(delete=False, suffix=".png") as tmp:
                                tmp.write(img_bytes)
                                tmp_path = tmp.name
                            
                            # Imagen
                            pdf.image(tmp_path, x=current_x, y=pdf.get_y(), w=img_w, h=40)
                            
                            # Descripción (Pie de foto opcional)
                            pdf.set_xy(current_x, pdf.get_y() + 41)
                            pdf.set_font("Arial", 'I', 7)
                            desc_txt = (foto_desc[:40]) if foto_desc else ""
                            pdf.cell(img_w, 4, desc_txt, align='C')
                            
                            current_x += img_w + 5
                            pdf.set_y(pdf.get_y() - 41)

                        except Exception as e:
                            print(f"Error procesando imagen: {e}")
                    
                    pdf.set_y(pdf.get_y() + 55)
                else:
                    pdf.ln(5)

        response_pdf = pdf.output(dest='S')
        return send_file(
            io.BytesIO(response_pdf),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"Reporte_Terranovus_{datetime.now().strftime('%Y%m%d')}.pdf"
        )

    except Exception as e:
        print(f"Error PDF: {e}")
        return f"Error: {str(e)}", 500
    finally:
        if conn:
            cursor.close()
            connection_pool.putconn(conn)


@app.route('/tablero-bi')
def tablero_bi():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    conn = None
    try:
        conn, cursor = get_db_connection()

        # 1. Estadísticas Generales de Proyectos
        cursor.execute('SELECT COUNT(*) FROM proyectos WHERE user_id = %s', (session['user_id'],))
        total_proyectos = cursor.fetchone()[0]

        # 2. Avance promedio y total de registros
        cursor.execute("""
            SELECT 
                COUNT(r.id_registro), 
                AVG(r.porcentaje_avance) 
            FROM registros r
            JOIN proyectos p ON r.id_proyecto = p.id_proyecto
            WHERE p.user_id = %s
        """, (session['user_id'],))
        stats = cursor.fetchone()
        total_registros = stats[0] or 0
        promedio_avance = round(stats[1], 2) if stats[1] else 0

        # 3. Conteo por Estados (para gráfico de torta)
        cursor.execute("""
            SELECT estado, COUNT(*) 
            FROM registros r
            JOIN proyectos p ON r.id_proyecto = p.id_proyecto
            WHERE p.user_id = %s
            GROUP BY estado
        """, (session['user_id'],))
        estados_raw = cursor.fetchall()
        
        # Convertimos a diccionario para fácil manejo en JS
        datos_estados = {row[0]: row[1] for row in estados_raw}

        return render_template('tableroBI.html', 
                               total_p=total_proyectos,
                               total_r=total_registros,
                               promedio=promedio_avance,
                               datos_estados=datos_estados)

    except Exception as e:
        print(f"Error en Tablero BI: {e}")
        return redirect(url_for('history'))
    finally:
        if conn:
            cursor.close()
            connection_pool.putconn(conn)


@app.route('/exportar-proyectos-excel', methods=['POST'])
def exportar_proyectos_excel():
    project_ids = request.form.getlist('project_ids')
    
    if not project_ids:
        return "No se seleccionaron proyectos", 400

    try:
        conn, cursor = get_db_connection()
        wb = Workbook()
        wb.remove(wb.active)  # Eliminar hoja por defecto

        for pid in project_ids:
            try:
                pid_int = int(pid)
            except:
                continue

            # Obtener info del proyecto
            cursor.execute("""
                SELECT nombre_proyecto, fecha_inicio, fecha_fin, director_obra, ubicacion, coordenadas
                FROM proyectos WHERE id_proyecto = %s
            """, (pid_int,))
            proyecto = cursor.fetchone()
            if not proyecto:
                continue

            nombre, fecha_inicio, fecha_fin, director, ubicacion, coordenadas = proyecto
            sheet_title = (nombre[:30] or f"Proyecto {pid_int}").strip()
            ws = wb.create_sheet(title=sheet_title)

            # Encabezado de proyecto
            ws.append(["Nombre del Proyecto:", nombre])
            ws.append(["Fecha de Inicio:", str(fecha_inicio)])
            ws.append(["Fecha de Finalización:", str(fecha_fin)])
            ws.append(["Director del Proyecto:", director])
            ws.append(["Ubicación:", ubicacion])
            ws.append(["Coordenadas:", coordenadas])
            ws.append([])

            # Encabezado de registros
            ws.append(["ID", "Zona de Intervención", "Ítems Instalados", "Metros Lineales", "Próximas Tareas", "Foto"])

            # Obtener registros
            cursor.execute("""
                SELECT id_registro, zona_intervencion, items, metros_lineales, proximas_tareas, foto_base64
                FROM registrosbitacoraeqing
                WHERE id_proyecto = %s
                ORDER BY id_registro DESC
            """, (pid_int,))
            registros = cursor.fetchall()

            row_index = 9
            for registro in registros:
                idr, zona, items, metros, tareas, foto = registro
                ws.append([idr, zona, items, metros, tareas, ""])

                if foto:
                    try:
                        header, base64_data = foto.split(',', 1) if ',' in foto else ('', foto)
                        img_data = base64.b64decode(base64_data)
                        img = Image.open(io.BytesIO(img_data))
                        img.thumbnail((120, 120))
                        img_io = io.BytesIO()
                        img.save(img_io, format='PNG')
                        img_io.seek(0)

                        img_excel = ExcelImage(img_io)
                        img_excel.anchor = f"F{row_index}"
                        ws.add_image(img_excel)

                        ws.row_dimensions[row_index].height = 90
                    except Exception as e:
                        print(f"Error en imagen de registro {idr}: {e}")
                row_index += 1

            # Ajustes de columnas
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 35
            ws.column_dimensions['F'].width = 18

        # Generar archivo
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(output,
                         download_name="proyectos_exportados.xlsx",
                         as_attachment=True,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        print(f"Error exportando proyectos: {e}")
        return "Error interno al exportar", 500
    finally:
        if conn:
            cursor.close()
            connection_pool.putconn(conn)

@app.route('/exportar-contactos-pdf', methods=['POST'])
def exportar_contactos_pdf():
    if 'user_id' not in session:
        return redirect(url_for('principalscreen'))

    id_proyecto = request.form.get('id_proyecto')
    if not id_proyecto:
        return "No se especifico proyecto", 400

    conn = None
    try:
        conn, cursor = get_db_connection()

        # Info del proyecto
        cursor.execute("""
            SELECT nombre_proyecto, cliente, contratista, orden_de_trabajo, ubicacion
            FROM proyectos WHERE id = %s
        """, (id_proyecto,))
        proyecto = cursor.fetchone()
        if not proyecto:
            return "Proyecto no encontrado", 404
        nombre_proy, cliente, contratista, ot, ubicacion = proyecto

        # Info de empresa (logo y colores)
        cursor.execute("""
            SELECT logo_url, color_primario
            FROM empresas WHERE id = %s
        """, (session.get('empresa_id'),))
        empresa_row = cursor.fetchone()
        logo_url    = empresa_row[0] if empresa_row else None
        color_hex   = empresa_row[1] if empresa_row else '#FFAF33'

        # Convertir color hex a RGB
        color_hex = color_hex.lstrip('#')
        cr, cg, cb = tuple(int(color_hex[i:i+2], 16) for i in (0, 2, 4))

        # Todos los contactos del proyecto
        cursor.execute("""
            SELECT c.id, c.nombre, c.empresa, c.cargo,
                   c.telefono, c.email, c.ciudad, c.notas,
                   c.created_at, u.name, u.apellido
            FROM contactos c
            LEFT JOIN usuario u ON u.user_id = c.user_id
            WHERE c.id_proyecto = %s
            ORDER BY c.created_at DESC
        """, (id_proyecto,))
        contactos = cursor.fetchall()

        # Fotos en una sola consulta
        ids_contactos = [c[0] for c in contactos]
        fotos_por_contacto = {}
        if ids_contactos:
            cursor.execute("""
                SELECT contacto_id, imagen_url
                FROM contacto_imagenes
                WHERE contacto_id = ANY(%s)
            """, (ids_contactos,))
            for f in cursor.fetchall():
                cid, url = f
                if cid not in fotos_por_contacto:
                    fotos_por_contacto[cid] = []
                if url:
                    fotos_por_contacto[cid].append(url)

        # Construir PDF
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()

        # ── ENCABEZADO TIPO TABLA ──
        y_enc = pdf.get_y()

        # Celda logo (40mm)
        pdf.rect(10, y_enc, 40, 22)
        if logo_url:
            try:
                resp = requests.get(logo_url, timeout=5)
                if resp.status_code == 200:
                    tmp_logo = NamedTemporaryFile(delete=False, suffix='.png')
                    tmp_logo.write(resp.content)
                    tmp_logo.close()
                    pdf.image(tmp_logo.name, x=12, y=y_enc + 2, h=18)
                    os.unlink(tmp_logo.name)
            except:
                pass

        # Celda titulo (150mm)
        pdf.set_fill_color(cr, cg, cb)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Arial", 'B', 14)
        pdf.set_xy(50, y_enc)
        pdf.cell(150, 22, "REPORTE DE CONTACTOS CAPTURADOS", border=1, ln=True, align='C', fill=True)
        pdf.set_text_color(0, 0, 0)

        # Filas info proyecto
        pdf.set_font("Arial", 'B', 9)
        pdf.set_fill_color(245, 245, 245)
        pdf.cell(30, 8, "PROYECTO:", border=1, fill=True)
        pdf.set_font("Arial", '', 9)
        pdf.cell(75, 8, nombre_proy or '', border=1)
        pdf.set_font("Arial", 'B', 9)
        pdf.cell(30, 8, "CLIENTE:", border=1, fill=True)
        pdf.set_font("Arial", '', 9)
        pdf.cell(55, 8, cliente or '', border=1, ln=True)

        pdf.set_font("Arial", 'B', 9)
        pdf.cell(30, 8, "CONTRATISTA:", border=1, fill=True)
        pdf.set_font("Arial", '', 9)
        pdf.cell(75, 8, contratista or '', border=1)
        pdf.set_font("Arial", 'B', 9)
        pdf.cell(30, 8, "UBICACION:", border=1, fill=True)
        pdf.set_font("Arial", '', 9)
        pdf.cell(55, 8, ubicacion or '', border=1, ln=True)

        pdf.ln(8)

        # ── CONTACTOS ──
        colombia_tz = pytz.timezone('America/Bogota')
        for contacto in contactos:
            id_c, nombre, empresa, cargo, telefono, email, ciudad, notas, created_at, u_name, u_apellido = contacto

            if created_at:
                created_at_col = created_at.replace(tzinfo=timezone.utc).astimezone(colombia_tz)
                fecha_str = created_at_col.strftime('%d/%m/%Y %I:%M %p')
            else:
                fecha_str = 'S/F'

            # Encabezado del contacto
            pdf.set_fill_color(cr, cg, cb)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Arial", 'B', 10)
            pdf.cell(190, 7, f"  {nombre or 'Sin nombre'}  -  {fecha_str}", ln=True, fill=True)
            pdf.set_text_color(0, 0, 0)

            # Datos del contacto
            pdf.set_fill_color(250, 250, 250)
            pdf.set_font("Arial", '', 9)
            pdf.cell(40, 6, "Empresa:", fill=True, border='L')
            pdf.cell(55, 6, empresa or '', border='R')
            pdf.cell(35, 6, "Cargo:", fill=True, border='L')
            pdf.cell(60, 6, cargo or '', border='R', ln=True)

            pdf.cell(40, 6, "Telefono:", fill=True, border='L')
            pdf.cell(55, 6, telefono or '', border='R')
            pdf.cell(35, 6, "Email:", fill=True, border='L')
            pdf.cell(60, 6, email or '', border='R', ln=True)

            pdf.cell(40, 6, "Ciudad:", fill=True, border='L')
            pdf.cell(150, 6, ciudad or '', border='R', ln=True)

            if notas:
                pdf.set_font("Arial", 'I', 9)
                pdf.multi_cell(190, 5, f"Notas: {notas}", border='LRB')
                pdf.set_x(10)

            # Registrado por
            registrado = f"{u_name or ''} {u_apellido or ''}".strip() or 'Sin asignar'
            pdf.set_font("Arial", '', 8)
            pdf.set_text_color(100, 100, 100)
            pdf.cell(190, 5, f"  Registrado por: {registrado}", ln=True)
            pdf.set_text_color(0, 0, 0)

            # Fotos
            fotos = fotos_por_contacto.get(id_c, [])
            if fotos:
                pdf.set_font("Arial", 'B', 8)
                pdf.cell(190, 5, "  Evidencia fotografica:", ln=True)
                x_foto = 10
                for foto_url in fotos[:3]:
                    try:
                        resp = requests.get(foto_url, timeout=5)
                        if resp.status_code == 200:
                            tmp = NamedTemporaryFile(delete=False, suffix='.jpg')
                            tmp.write(resp.content)
                            tmp.close()
                            pdf.image(tmp.name, x=x_foto, y=pdf.get_y(), h=30)
                            os.unlink(tmp.name)
                            x_foto += 35
                    except:
                        pass
                if fotos:
                    pdf.ln(32)

            pdf.ln(4)

        # Pie de pagina
        pdf.set_y(-15)
        pdf.set_font("Arial", 'I', 8)
        pdf.set_text_color(150, 150, 150)
        pdf.cell(0, 10, f"Generado por Bitacora IAC - {datetime.now().strftime('%d/%m/%Y %H:%M')}", align='C')

        output = io.BytesIO()
        pdf.output(output)
        output.seek(0)

        return send_file(output,
                         download_name=f"contactos_{nombre_proy or id_proyecto}.pdf",
                         as_attachment=True,
                         mimetype='application/pdf')
    except Exception as e:
        print(f"Error exportando PDF: {e}")
        traceback.print_exc()
        return "Error interno al exportar", 500
    finally:
        if conn:
            cursor.close()
            connection_pool.putconn(conn)


@app.route('/exportar-contactos-excel', methods=['POST'])
def exportar_contactos_excel():
    if 'user_id' not in session:
        return redirect(url_for('principalscreen'))

    id_proyecto = request.form.get('id_proyecto')
    if not id_proyecto:
        return "No se especificó proyecto", 400

    conn = None
    try:
        conn, cursor = get_db_connection()

        # Info del proyecto
        cursor.execute("""
            SELECT nombre_proyecto, cliente, contratista, orden_de_trabajo, ubicacion
            FROM proyectos WHERE id = %s
        """, (id_proyecto,))
        proyecto = cursor.fetchone()
        if not proyecto:
            return "Proyecto no encontrado", 404
        nombre_proy, cliente, contratista, ot, ubicacion = proyecto

        # Todos los contactos
        cursor.execute("""
            SELECT c.id, c.nombre, c.empresa, c.cargo,
                   c.telefono, c.email, c.ciudad, c.notas,
                   c.created_at, u.name, u.apellido
            FROM contactos c
            LEFT JOIN usuario u ON u.user_id = c.user_id
            WHERE c.id_proyecto = %s
            ORDER BY c.created_at DESC
        """, (id_proyecto,))
        contactos = cursor.fetchall()

        # Fotos en una sola consulta
        ids_contactos = [c[0] for c in contactos]
        fotos_por_contacto = {}
        if ids_contactos:
            cursor.execute("""
                SELECT contacto_id, imagen_url
                FROM contacto_imagenes
                WHERE contacto_id = ANY(%s)
            """, (ids_contactos,))
            for f in cursor.fetchall():
                cid, url = f
                if cid not in fotos_por_contacto:
                    fotos_por_contacto[cid] = []
                if url:
                    fotos_por_contacto[cid].append(url)

        # Construir Excel
        wb = Workbook()
        ws = wb.active
        ws.title = (nombre_proy or 'Contactos')[:30]

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # Info del proyecto
        ws.append(["Proyecto:", nombre_proy])
        ws.append(["Cliente:", cliente])
        ws.append(["Contratista:", contratista])
        ws.append(["Orden de Trabajo:", ot])
        ws.append(["Ubicación:", ubicacion])
        ws.append([])

        # Encabezados
        headers = ["#", "Nombre", "Empresa", "Cargo", "Teléfono",
                   "Email", "Ciudad", "Notas", "Fecha", "Registrado por", "Fotos"]
        ws.append(headers)

        header_row = ws.max_row
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col)
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = PatternFill("solid", fgColor="FFAF33")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        colombia_tz = pytz.timezone('America/Bogota')
        row_index   = header_row + 1

        for i, contacto in enumerate(contactos, 1):
            id_c, nombre, empresa, cargo, telefono, email, ciudad, notas, created_at, u_name, u_apellido = contacto

            if created_at:
                created_at_col = created_at.replace(tzinfo=timezone.utc).astimezone(colombia_tz)
                fecha_str = created_at_col.strftime('%d/%m/%Y %I:%M %p')
            else:
                fecha_str = 'S/F'

            registrado = f"{u_name or ''} {u_apellido or ''}".strip() or 'Sin asignar'

            ws.append([
                i, nombre or '', empresa or '', cargo or '',
                telefono or '', email or '', ciudad or '',
                notas or '', fecha_str, registrado, ''
            ])

            ws.row_dimensions[row_index].height = 80

            # Fotos
            fotos = fotos_por_contacto.get(id_c, [])
            col_foto = 11
            for foto_url in fotos[:3]:
                try:
                    resp = requests.get(foto_url, timeout=5)
                    if resp.status_code == 200:
                        img     = Image.open(io.BytesIO(resp.content))
                        img.thumbnail((100, 100))
                        img_io  = io.BytesIO()
                        img.save(img_io, format='PNG')
                        img_io.seek(0)
                        img_excel        = ExcelImage(img_io)
                        img_excel.anchor = f"{get_column_letter(col_foto)}{row_index}"
                        ws.add_image(img_excel)
                        col_foto += 1
                except Exception as e:
                    print(f"Error foto contacto {id_c}: {e}")

            row_index += 1

        # Anchos de columna
        anchos = [5, 20, 20, 18, 15, 25, 15, 30, 18, 20, 15]
        for i, ancho in enumerate(anchos, 1):
            ws.column_dimensions[get_column_letter(i)].width = ancho

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(output,
                         download_name=f"contactos_{nombre_proy or id_proyecto}.xlsx",
                         as_attachment=True,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        print(f"Error exportando Excel: {e}")
        traceback.print_exc()
        return "Error interno al exportar", 500
    finally:
        if conn:
            cursor.close()
            connection_pool.putconn(conn)

app.register_blueprint(api_movil)

if __name__ == '__main__':
    app.run(debug=True)